"""
Combination Strategy Compilation - Multi-Symbol Portfolio Analysis

This script processes WFAlphaResults backtest data to generate comprehensive
portfolio analysis with correlation optimization across multiple symbols.

Output: Combination_Strategy_Compilation_{exchange}_{interval}_{YYYYMMDD}.xlsx

- Corr Summary: Correlation optimization results for all symbols
- {SYMBOL} Portfolio: Full timeseries portfolio data per symbol
- {SYMBOL} Corr: Correlation matrix per symbol
"""

import pandas as pd
import numpy as np
import os
from pathlib import Path
from datetime import datetime
import sys
import glob
import argparse

# Import utility functions
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from util_AQS_parallel import get_period_for_interval

# ============================================================================
# DIRECTORY PATHS
# ============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = SCRIPT_DIR  # Script is in project root
DATA_DIR = os.path.join(PROJECT_DIR, "GridSearch_Data")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "WFAlphaResults")

# ============================================================================
# CONFIGURATION (defaults, can be overridden via command-line)
# ============================================================================

EXCHANGE = "ibkr"
INTERVAL = "1h"

# Symbol filtering - user can modify this list
# None = process all symbols from WF_Short
# List = process only specified symbols (e.g., ["VIXY", "VIXM"])
SYMBOLS = ["VIXY", "VIXM"]  # Base symbols

# Worksheet name
WORKSHEET_NAME = "WF_Short"

# Correlation optimization thresholds
CORRELATION_THRESHOLDS = [0.7, 0.6, 0.5]


# ============================================================================
# ARGUMENT PARSING
# ============================================================================

def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Combination strategy compilation with correlation optimization"
    )

    parser.add_argument(
        "--exchange",
        type=str,
        default=EXCHANGE,
        help=f"Exchange name (default: {EXCHANGE})",
    )

    parser.add_argument(
        "--interval",
        type=str,
        default=INTERVAL,
        help=f"Time interval (default: {INTERVAL})",
    )

    parser.add_argument(
        "--symbols",
        type=str,
        default=",".join(SYMBOLS),
        help=f"Comma-separated list of symbols (default: {','.join(SYMBOLS)})",
    )

    # NEW: optional date override (YYYYMMDD)
    parser.add_argument(
        "--date",
        type=str,
        default=None,
        help="Override WFAlpha date in YYYYMMDD (default: today)",
    )

    return parser.parse_args()


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def log_error(message, error_log_path):
    """
    Log error message to file with timestamp.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] {message}\n"

    os.makedirs(os.path.dirname(error_log_path), exist_ok=True)
    with open(error_log_path, "a", encoding="utf-8") as f:
        f.write(log_entry)

    print(f" [ERROR] {message}")


def filter_symbols(df_wf_short, symbols_config):
    """
    Filter WF_Short by symbol configuration.
    """
    if symbols_config is None:
        symbols = sorted(df_wf_short["Symbol"].unique())
        print(f"Processing ALL symbols: {len(symbols)} total")
        print(f"Symbols: {', '.join(symbols)}")
    else:
        df_wf_short = df_wf_short[df_wf_short["Symbol"].isin(symbols_config)]
        symbols = sorted(df_wf_short["Symbol"].unique())
        print(f"Processing SELECTED symbols: {len(symbols)} total")
        print(f"Symbols: {', '.join(symbols)}")

        missing = set(symbols_config) - set(symbols)
        if missing:
            print(f"WARNING: Symbols not found in WF_Short: {', '.join(missing)}")

    return df_wf_short, symbols


def load_wf_short(input_file, worksheet_name):
    """
    Load WF_Short configurations from Excel.
    """
    print("\n" + "=" * 80)
    print("LOADING WF_SHORT CONFIGURATIONS")
    print("=" * 80)
    print(f"Input file: {input_file}")
    print(f"Worksheet: {worksheet_name}")

    try:
        df = pd.read_excel(input_file, sheet_name=worksheet_name)
        # strip any accidental whitespace in column names
        df.columns = df.columns.str.strip()
        print(f"Loaded {len(df)} configurations from worksheet")
        return df
    except Exception as e:
        print(f"ERROR: Failed to read Excel file: {e}")
        return pd.DataFrame()


def construct_backtest_path(
    exchange, symbol, interval, feature, model, strategy, base_dir, variant=None
):
    """
    Construct path to backtest.csv file.
    If variant is provided, constructs exact path. Otherwise uses glob pattern.
    """
    if variant:
        folder_name = f"merged_{exchange}_{symbol}_{interval}_{variant}"
        path = os.path.join(base_dir, folder_name, feature, model, strategy, "backtest.csv")
        # Debug
        # print(f"    [DEBUG] Path: {path}")
        # print(f"    [DEBUG] Exists: {os.path.exists(path)}")
        return path
    else:
        pattern = f"merged_{exchange}_{symbol}_{interval}_*"
        pattern_path = os.path.join(base_dir, pattern)
        matches = glob.glob(pattern_path)
        if not matches:
            return None
        folder_name = os.path.basename(matches[0])
        path = os.path.join(base_dir, folder_name, feature, model, strategy, "backtest.csv")
        return path


def load_strategy_backtest(backtest_path, strategy_name):
    """
    Load backtest CSV and extract datetime, pnl, signal columns.
    """
    if backtest_path is None or not os.path.exists(backtest_path):
        return None

    try:
        df = pd.read_csv(backtest_path)

        if "datetime" not in df.columns or "pnl" not in df.columns or "signal" not in df.columns:
            # print(f"    [DEBUG] Missing columns. Found: {df.columns.tolist()}")
            return None

        df_result = pd.DataFrame(
            {
                # handle DST / mixed offsets safely
                "datetime": pd.to_datetime(df["datetime"], utc=True).dt.tz_convert(None),
                strategy_name: df["pnl"],
                f"{strategy_name}.1": df["signal"],
            }
        )
        return df_result
    except Exception as e:
        # print(f"    [DEBUG] Exception in load_strategy_backtest: {e}")
        return None


# ============================================================================
# PORTFOLIO CONSTRUCTION
# ============================================================================

def build_symbol_portfolio(symbol, df_wf_short_symbol, base_dir, error_log_path, interval):
    """
    Build portfolio for a single symbol by merging all strategy backtests.
    """
    print(f"\n  Building portfolio for {symbol}...")
    print(f"  Strategies to process: {len(df_wf_short_symbol)}")

    merged_df = None
    strategy_metrics = {}
    strategy_names = []
    loaded_count = 0
    skipped_count = 0

    for _, row in df_wf_short_symbol.iterrows():
        feature = row["Data Point"]
        model = row["Model"]
        strategy = row["Entry / Exit Model"]
        variant = row.get("Variant", None) if "Variant" in row.index else None

        if variant and str(variant).lower() != "default":
            strategy_name = f"{symbol}_{feature}_{model}_{strategy}_{variant}"
        else:
            strategy_name = f"{symbol}_{feature}_{model}_{strategy}"

        strategy_names.append(strategy_name)

        backtest_path = construct_backtest_path(
            EXCHANGE, symbol, interval, feature, model, strategy, base_dir, variant
        )

        df_strategy = load_strategy_backtest(backtest_path, strategy_name)
        if df_strategy is None:
            err = f"Failed to load backtest for {symbol}: {feature}/{model}/{strategy}"
            log_error(err, error_log_path)
            skipped_count += 1
            continue

        if merged_df is None:
            merged_df = df_strategy
        else:
            if not merged_df["datetime"].equals(df_strategy["datetime"]):
                print(f"  WARNING: DateTime mismatch for {strategy_name}, using inner join")
                merged_df = pd.merge(merged_df, df_strategy, on="datetime", how="inner")
            else:
                merged_df = (
                    pd.concat(
                        [
                            merged_df.set_index("datetime"),
                            df_strategy.set_index("datetime").drop(
                                "datetime", axis=1, errors="ignore"
                            ),
                        ],
                        axis=1,
                    )
                    .reset_index()
                    .rename(columns={"index": "datetime"})
                )

        strategy_metrics[strategy_name] = {
            "Sharpe": row["Sharpe"],
            "MDD": row["MDD"],
            "Annual_Return": row["Annual Return"],
        }

        loaded_count += 1

    if merged_df is None or loaded_count == 0:
        print(f"  ERROR: No backtests loaded for {symbol}")
        return None, None, None

    print(f"  [OK] Loaded {loaded_count} strategies, skipped {skipped_count}")
    print(
        f"  Portfolio datetime range: {merged_df['datetime'].min()} to "
        f"{merged_df['datetime'].max()}"
    )
    print(f"  Total rows: {len(merged_df)}")

    return merged_df, strategy_metrics, strategy_names


def load_close_prices(symbol, interval):
    """
    Load close prices from original merged CSV for Buy & Hold calculation.
    """
    pattern = f"merged_{EXCHANGE}_{symbol}_{interval}_*.csv"
    pattern_path = os.path.join(DATA_DIR, pattern)
    matches = glob.glob(pattern_path)

    if not matches:
        print(f"  WARNING: Close price CSV not found: {pattern}")
        return None

    csv_path = matches[0]
    try:
        df = pd.read_csv(csv_path, usecols=["datetime", "close"])
        # df["datetime"] = pd.to_datetime(df["datetime"]).dt.tz_localize(None)
        df['datetime'] = pd.to_datetime(df['datetime'], utc=True).dt.tz_convert(None)  #Revision after debug
        return df
    except Exception as e:
        print(f"  WARNING: Failed to load close prices: {e}")
        return None


def calculate_portfolio_metrics(merged_df, strategy_names, symbol, interval):
    """
    Calculate portfolio metrics (equal-weighted) with Buy & Hold comparison.
    """
    period = get_period_for_interval(interval)
    pnl_columns = [col for col in strategy_names if not col.endswith(".1")]

    merged_df["portfolio_pnl"] = merged_df[pnl_columns].mean(axis=1)
    merged_df["cumulative_pnl"] = merged_df["portfolio_pnl"].cumsum()

    df_close = load_close_prices(symbol, interval)
    if df_close is not None:
        merged_df = pd.merge(merged_df, df_close, on="datetime", how="left")
        merged_df["bnh"] = (
            (merged_df["close"] / merged_df["close"].shift(1)) - 1
        ).cumsum().fillna(0)
        merged_df = merged_df.drop(columns=["close"])
    else:
        merged_df["bnh"] = np.nan
        print(f"  WARNING: Buy & Hold not calculated for {symbol} (close prices unavailable)")

    merged_df["drawdown"] = (
        merged_df["cumulative_pnl"] - merged_df["cumulative_pnl"].cummax()
    )

    portfolio_sharpe = (
        merged_df["portfolio_pnl"].mean()
        / merged_df["portfolio_pnl"].std()
        * np.sqrt(period)
    )
    portfolio_mdd = merged_df["drawdown"].min()
    portfolio_annual_return = merged_df["portfolio_pnl"].mean() * period
    portfolio_calmar = (
        portfolio_annual_return / abs(portfolio_mdd) if portfolio_mdd != 0 else np.nan
    )

    final_cumulative_pnl = merged_df["cumulative_pnl"].iloc[-1] if len(merged_df) > 0 else 0
    if len(merged_df) > 0 and not merged_df["bnh"].isna().all():
        final_bnh = merged_df["bnh"].iloc[-1]
    else:
        final_bnh = 0

    pnl_ratio = final_cumulative_pnl / final_bnh if final_bnh != 0 else np.nan

    portfolio_metrics = {
        "Sharpe": portfolio_sharpe,
        "MDD": portfolio_mdd,
        "Annual_Return": portfolio_annual_return,
        "Calmar": portfolio_calmar,
        "Cumulative_PnL": final_cumulative_pnl,
        "BnH": final_bnh,
        "PnL_Ratio": pnl_ratio,
    }

    return merged_df, portfolio_metrics


# ============================================================================
# CORRELATION ANALYSIS
# ============================================================================

def calculate_correlation_matrix(merged_df, strategy_names):
    pnl_columns = [col for col in strategy_names if not col.endswith(".1")]
    return merged_df[pnl_columns].corr()


def greedy_correlation_optimization(corr_matrix, strategy_metrics, threshold):
    sorted_strategies = sorted(
        strategy_metrics.keys(), key=lambda x: strategy_metrics[x]["Sharpe"], reverse=True
    )

    selected = [sorted_strategies[0]]
    for candidate in sorted_strategies[1:]:
        max_corr = max(abs(corr_matrix.loc[candidate, s]) for s in selected)
        if max_corr < threshold:
            selected.append(candidate)

    selected_corr_matrix = corr_matrix.loc[selected, selected]
    upper_tri = np.triu(selected_corr_matrix.values, k=1)
    upper_vals = upper_tri[upper_tri != 0]

    if len(upper_vals) > 0:
        avg_corr = np.mean(upper_vals)
        max_corr = np.max(upper_vals)
        min_corr = np.min(upper_vals)
    else:
        avg_corr = max_corr = min_corr = np.nan

    sharpes = [strategy_metrics[s]["Sharpe"] for s in selected]
    portfolio_sharpe = np.mean(sharpes) if sharpes else np.nan

    return {
        "selected_strategies": selected,
        "count": len(selected),
        "portfolio_sharpe": portfolio_sharpe,
        "avg_correlation": avg_corr,
        "max_correlation": max_corr,
        "min_correlation": min_corr,
    }


def calculate_optimized_portfolio_metrics(merged_df, selected_strategies, interval):
    period = get_period_for_interval(interval)

    portfolio_pnl = merged_df[selected_strategies].mean(axis=1)
    cumulative_pnl = portfolio_pnl.cumsum()
    drawdown = cumulative_pnl - cumulative_pnl.cummax()

    sharpe = (portfolio_pnl.mean() / portfolio_pnl.std()) * np.sqrt(period)
    mdd = drawdown.min()
    annual_return = portfolio_pnl.mean() * period
    calmar = annual_return / abs(mdd) if mdd != 0 else np.nan

    if len(cumulative_pnl) > 0:
        final_cumulative_pnl = cumulative_pnl.iloc[-1]
    else:
        final_cumulative_pnl = 0

    if "bnh" in merged_df.columns and len(merged_df) > 0 and not merged_df["bnh"].isna().all():
        final_bnh = merged_df["bnh"].iloc[-1]
    else:
        final_bnh = 0

    pnl_ratio = final_cumulative_pnl / final_bnh if final_bnh != 0 else np.nan

    return {
        "Sharpe": sharpe,
        "MDD": mdd,
        "Annual_Return": annual_return,
        "Calmar": calmar,
        "Cumulative_PnL": final_cumulative_pnl,
        "BnH": final_bnh,
        "PnL_Ratio": pnl_ratio,
    }


# ============================================================================
# WORKSHEET CREATION
# ============================================================================

def create_portfolio_worksheet(symbol, merged_df, strategy_names, strategy_metrics,
                               portfolio_metrics, writer):
    sheet_name = f"{symbol} Portfolio"
    merged_df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    from openpyxl.utils import get_column_letter

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def create_correlation_worksheet(symbol, corr_matrix, writer):
    sheet_name = f"{symbol} Corr"
    upper_tri = np.triu(corr_matrix.values, k=1)
    upper_vals = upper_tri[upper_tri != 0]

    if len(upper_vals) > 0:
        avg_corr = np.mean(upper_vals)
        max_corr = np.max(upper_vals)
        min_corr = np.min(upper_vals)
    else:
        avg_corr = max_corr = min_corr = np.nan

    corr_output = corr_matrix.copy()

    summary_row = pd.Series({col: np.nan for col in corr_output.columns}, name="---")
    corr_output = pd.concat([corr_output, summary_row.to_frame().T])

    stats_row = pd.Series(
        {corr_output.columns[0]: "STATISTICS", **{col: np.nan for col in corr_output.columns[1:]}} ,
        name="Statistics",
    )
    corr_output = pd.concat([corr_output, stats_row.to_frame().T])

    avg_row = pd.Series(
        {
            corr_output.columns[0]:
                f"Avg Correlation: {avg_corr:.4f}" if not np.isnan(avg_corr) else "Avg Correlation: N/A",
            **{col: np.nan for col in corr_output.columns[1:]},
        },
        name="Avg",
    )
    corr_output = pd.concat([corr_output, avg_row.to_frame().T])

    max_row = pd.Series(
        {
            corr_output.columns[0]:
                f"Max Correlation: {max_corr:.4f}" if not np.isnan(max_corr) else "Max Correlation: N/A",
            **{col: np.nan for col in corr_output.columns[1:]},
        },
        name="Max",
    )
    corr_output = pd.concat([corr_output, max_row.to_frame().T])

    min_row = pd.Series(
        {
            corr_output.columns[0]:
                f"Min Correlation: {min_corr:.4f}" if not np.isnan(min_corr) else "Min Correlation: N/A",
            **{col: np.nan for col in corr_output.columns[1:]},
        },
        name="Min",
    )
    corr_output = pd.concat([corr_output, min_row.to_frame().T])

    count_row = pd.Series(
        {
            corr_output.columns[0]: f"Strategies: {len(corr_matrix)}",
            **{col: np.nan for col in corr_output.columns[1:]},
        },
        name="Count",
    )
    corr_output = pd.concat([corr_output, count_row.to_frame().T])

    corr_output.to_excel(writer, sheet_name=sheet_name)

    ws = writer.sheets[sheet_name]
    from openpyxl.utils import get_column_letter

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def create_summary_worksheet(summary_data, writer):
    sheet_name = "Corr Summary"
    df_summary = pd.DataFrame(summary_data)
    df_summary = df_summary.sort_values(["Symbol", "Threshold"], ascending=[True, False])
    df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    from openpyxl.utils import get_column_letter

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)


# ============================================================================
# INTERACTIVE MODE
# ============================================================================

def interactive_mode():
    """Interactive CLI for running Stage 6: Combination Strategies"""
    print(
        """
╔══════════════════════════════════════════════════════════════╗
║    Stage 6: Combination Strategies - Interactive Mode        ║
║  Portfolio construction with correlation optimization        ║
╚══════════════════════════════════════════════════════════════╝
"""
    )
    try:
        symbols_input = input(
            "Enter symbols (comma-separated, e.g., NVDA, AAPL, TECL): "
        ).strip()
        if not symbols_input:
            print("Symbols cannot be empty")
            return

        symbols = [s.strip().upper() for s in symbols_input.split(",")]

        print("\nAvailable intervals:")
        print("  1min, 5min, 15min, 30min, 1h, 1d, 1w")
        interval = input("Enter interval (default: 1h): ").strip().lower()
        if not interval or interval not in ["1min", "5min", "15min", "30min", "1h", "1d", "1w"]:
            interval = "1h"
            print("  ✗ Invalid interval. Using default: 1h")

        date_override = input("Enter WFAlpha date YYYYMMDD (blank = today): ").strip()
        if date_override == "":
            date_override = None

        print("\n" + "=" * 60)
        print("Configuration Summary:")
        print(f"  Symbols:     {', '.join(symbols)}")
        print(f"  Interval:    {interval}")
        print(f"  Exchange:    {EXCHANGE}")
        print(f"  Date:        {date_override or datetime.now().strftime('%Y%m%d')}")
        print("=" * 60)

        proceed = input("\nProceed with execution? (y/n): ").strip().lower()
        if proceed != "y":
            print("\nExecution cancelled")
            return

        sys.argv = [
            sys.argv[0],
            "--symbols",
            ",".join(symbols),
            "--interval",
            interval,
            "--exchange",
            EXCHANGE,
        ]
        if date_override is not None:
            sys.argv.extend(["--date", date_override])

        main()

    except KeyboardInterrupt:
        print("\n\nCancelled by user")


# ============================================================================
# MAIN ORCHESTRATION
# ============================================================================

def main():
    """Main function to orchestrate combination strategy compilation."""
    args = parse_arguments()
    exchange = args.exchange
    interval = args.interval
    symbols_arg = [s.strip() for s in args.symbols.split(",")]

    # use provided date or today
    today = args.date if args.date is not None else datetime.now().strftime("%Y%m%d")

    input_file = os.path.join(
        OUTPUT_DIR, f"WFAlpha_Compilation_{exchange}_{interval}_{today}.xlsx"
    )
    output_file = f"Combination_Strategy_Compilation_{exchange}_{interval}_{today}.xlsx"
    error_log_file = os.path.join(OUTPUT_DIR, "combination_strategy_errors.txt")

    print("\n" + "=" * 80)
    print("COMBINATION STRATEGY COMPILATION (MULTI-SYMBOL)")
    print("=" * 80)
    print(f"Exchange: {exchange}")
    print(f"Interval: {interval}")
    print(f"Symbols: {', '.join(symbols_arg)}")
    print(f"Date: {today}")
    print(f"Output: {output_file}")
    print("=" * 80)

    if os.path.exists(error_log_file):
        os.remove(error_log_file)

    df_wf_short = load_wf_short(input_file, WORKSHEET_NAME)
    if df_wf_short.empty:
        print("\nERROR: Failed to load WF_Short!")
        return

    df_wf_short, symbols = filter_symbols(df_wf_short, symbols_arg)
    if len(symbols) == 0:
        print("\nERROR: No symbols to process!")
        return

    print(f"\nTotal strategies to process: {len(df_wf_short)}")
    print("=" * 80)

    symbol_portfolios = {}
    symbol_correlations = {}
    summary_data = []

    for idx, symbol in enumerate(symbols):
        print(f"\n[{idx + 1}/{len(symbols)}] Processing {symbol}...")
        print("-" * 80)

        df_symbol = df_wf_short[df_wf_short["Symbol"] == symbol]

        merged_df, strategy_metrics, strategy_names = build_symbol_portfolio(
            symbol, df_symbol, OUTPUT_DIR, error_log_file, interval
        )

        if merged_df is None:
            print(f"  Skipping {symbol} due to errors")
            continue

        merged_df, portfolio_metrics = calculate_portfolio_metrics(
            merged_df, strategy_names, symbol, interval
        )
        print(f"  Portfolio Sharpe:        {portfolio_metrics['Sharpe']:.4f}")
        print(f"  Portfolio MDD:           {portfolio_metrics['MDD']:.4f}")
        print(f"  Portfolio Annual Return: {portfolio_metrics['Annual_Return']:.4f}")

        symbol_portfolios[symbol] = {
            "df": merged_df,
            "strategy_names": strategy_names,
            "strategy_metrics": strategy_metrics,
            "portfolio_metrics": portfolio_metrics,
        }

        pnl_columns = [col for col in strategy_names if not col.endswith(".1")]
        if len(pnl_columns) > 1:
            corr_matrix = calculate_correlation_matrix(merged_df, strategy_names)
            symbol_correlations[symbol] = corr_matrix

            for threshold in CORRELATION_THRESHOLDS:
                opt_result = greedy_correlation_optimization(
                    corr_matrix, strategy_metrics, threshold
                )

                if len(opt_result["selected_strategies"]) > 0:
                    opt_metrics = calculate_optimized_portfolio_metrics(
                        merged_df, opt_result["selected_strategies"], interval
                    )
                else:
                    opt_metrics = {
                        "Sharpe": np.nan,
                        "MDD": np.nan,
                        "Annual_Return": np.nan,
                        "Calmar": np.nan,
                        "Cumulative_PnL": np.nan,
                        "BnH": np.nan,
                        "PnL_Ratio": np.nan,
                    }

                summary_data.append(
                    {
                        "Symbol": symbol,
                        "Threshold": threshold,
                        "Total Strategies": len(pnl_columns),
                        "# Strategies Selected": opt_result["count"],
                        "Portfolio Sharpe": opt_metrics["Sharpe"],
                        "Avg Correlation": opt_result["avg_correlation"],
                        "Max Correlation": opt_result["max_correlation"],
                        "Min Correlation": opt_result["min_correlation"],
                        "Portfolio MDD": opt_metrics["MDD"],
                        "Portfolio Annual Return": opt_metrics["Annual_Return"],
                        "Portfolio Calmar Ratio": opt_metrics["Calmar"],
                        "Cumulative PnL": opt_metrics.get("Cumulative_PnL", np.nan),
                        "Buy & Hold": opt_metrics.get("BnH", np.nan),
                        "PnL Ratio": opt_metrics.get("PnL_Ratio", np.nan),
                        "Strategy List": ", ".join(opt_result["selected_strategies"]),
                    }
                )

                print(
                    f"  Optimization @ {threshold}: "
                    f"{opt_result['count']} strategies selected, "
                    f"Sharpe={opt_metrics['Sharpe']:.4f}"
                )
        else:
            print(f"  WARNING: Only 1 strategy for {symbol}, skipping correlation analysis")

    print("\n" + "=" * 80)
    print("SAVING TO EXCEL")
    print("=" * 80)
    output_path = os.path.join(OUTPUT_DIR, output_file)
    print(f"Output file: {output_path}")

    if len(symbol_portfolios) == 0:
        print("ERROR: No portfolios built. Nothing to save. Check errors above.")
        return

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if len(summary_data) > 0:
            print("  Writing Corr Summary...")
            create_summary_worksheet(summary_data, writer)

        for symbol in sorted(symbol_portfolios.keys()):
            print(f"  Writing {symbol} Portfolio...")
            portfolio_data = symbol_portfolios[symbol]
            create_portfolio_worksheet(
                symbol,
                portfolio_data["df"],
                portfolio_data["strategy_names"],
                portfolio_data["strategy_metrics"],
                portfolio_data["portfolio_metrics"],
                writer,
            )

            if symbol in symbol_correlations:
                print(f"  Writing {symbol} Corr...")
                create_correlation_worksheet(symbol, symbol_correlations[symbol], writer)

    print("\n" + "=" * 80)
    print("COMPILATION COMPLETE")
    print("=" * 80)
    print(f"Symbols processed: {len(symbol_portfolios)}")
    print(f"Total worksheets: {1 + len(symbol_portfolios) * 2}")
    print(f"Output: {output_path}")
    if os.path.exists(error_log_file):
        print(f"\nErrors logged to: {error_log_file}")
    print("=" * 80)


if __name__ == "__main__":
    if len(sys.argv) == 1:
        interactive_mode()
    else:
        main()
