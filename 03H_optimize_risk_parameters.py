"""
03H_optimize_risk_parameters.py
================================

Stage 8: Risk Parameter Optimization

Takes strategies from the Final_Alphas sheet (produced by Stage 7) and finds
the optimal Stop Loss (SL), Take Profit (TP), and Trailing Stop Loss (TSL)
for each strategy using a three-phase grid search.

Phase 1: SL/TP grid search (2,501 combinations)
Phase 2A: TSL/TP grid search (1,281 combinations)
Phase 2B: Bracket orders - SL(P1) + TP candidates + TSL (21-42 combinations)

Usage:
    # Interactive mode (no args — prompts for exchange, interval, input file, symbols)
    python 03H_optimize_risk_parameters.py

    # CLI mode (auto-detects latest Final_Compilation)
    python 03H_optimize_risk_parameters.py --interval 1h --exchange ibkr

    # CLI mode with symbol filter
    python 03H_optimize_risk_parameters.py --interval 1h --symbols VIXY,VIXM

    # Pipeline mode
    python 03_run_ibkr_gs_pipeline.py --start-stage 8 --end-stage 8
"""

import argparse
import glob
import os
import sys
import time
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
from joblib import Parallel, delayed

from util_AQS_parallel import (
    calculate_with_risk_params,
    filter_by_min_trade_count,
    get_period_for_interval,
    select_best_combination,
)

warnings.filterwarnings('ignore', category=FutureWarning)

# ============================================================================
# CONSTANTS
# ============================================================================

FEE = 0.06 / 100  # 6 basis points
TRADE_COUNT_THRESHOLD = 0.015  # 1.5% of total rows
BASELINE_TOLERANCE = 0.001  # 0.1% for baseline validation

# Grid search ranges
SL_RANGE = np.arange(0.0, 0.205, 0.005).tolist()   # 0% to 20%, step 0.5% -> 41 values
TP_RANGE = np.arange(0.0, 0.305, 0.005).tolist()    # 0% to 30%, step 0.5% -> 61 values
TSL_RANGE = np.arange(0.0, 0.105, 0.005).tolist()   # 0% to 10%, step 0.5% -> 21 values

# Required columns in Final_Alphas sheet
REQUIRED_COLUMNS = ['Symbol', 'Data Point', 'Model', 'Entry / Exit Model']

# ============================================================================
# CLI ARGUMENTS
# ============================================================================

def parse_arguments():
    parser = argparse.ArgumentParser(
        description='Stage 8: Optimize SL/TP/TSL risk parameters for Final Alpha strategies'
    )
    parser.add_argument('--exchange', type=str, default='ibkr',
                        help='Exchange name (default: ibkr)')
    parser.add_argument('--interval', type=str, default='1h',
                        help='Data interval (default: 1h)')
    parser.add_argument('--symbols', type=str, default=None,
                        help='Comma-separated symbol list to filter (default: all symbols)')
    parser.add_argument('--input', type=str, default=None,
                        help='Explicit path to Final_Compilation Excel (overrides auto-detect)')
    parser.add_argument('--output', type=str, default=None,
                        help='Explicit output path for Risk_Optimized_Compilation Excel')
    parser.add_argument('--n-jobs', type=int, default=-1,
                        help='Number of parallel jobs for grid search (default: -1 = all cores)')
    return parser.parse_args()

# ============================================================================
# INTERACTIVE MODE
# ============================================================================

VALID_INTERVALS = ['1min', '5min', '15min', '30min', '1h', '4h', '1d', '1w']


def interactive_mode():
    """
    Interactive CLI to gather user inputs when no CLI args are provided.

    Returns:
    --------
    dict or None : Configuration dict, or None if user cancels.
    """
    print()
    print("=" * 60)
    print("  STAGE 8: Risk Parameter Optimization - Setup")
    print("=" * 60)

    # --- Prompt 1: Exchange ---
    exchange = input("\n  Exchange [ibkr]: ").strip().lower() or 'ibkr'

    # --- Prompt 2: Interval ---
    while True:
        interval = input(f"  Interval [1h]: ").strip().lower() or '1h'
        if interval in VALID_INTERVALS:
            break
        print(f"    Invalid interval. Choose from: {', '.join(VALID_INTERVALS)}")

    # --- Prompt 3: Input file selection ---
    pattern = f"WFAlphaResults/Final_Compilation_{exchange}_{interval}_*.xlsx"
    matches = sorted(glob.glob(pattern))

    if not matches:
        print(f"\n  ERROR: No Final_Compilation files found matching:")
        print(f"    {pattern}")
        print(f"  Run Stage 7 first.")
        return None

    print(f"\n  Available Final_Compilation files:")
    for i, path in enumerate(matches, 1):
        print(f"    [{i}] {os.path.basename(path)}")

    while True:
        choice = input(f"  Select file number (Enter = latest [{len(matches)}]): ").strip()
        if choice == '':
            input_path = matches[-1]
            break
        try:
            idx = int(choice)
            if 1 <= idx <= len(matches):
                input_path = matches[idx - 1]
                break
            print(f"    Enter a number between 1 and {len(matches)}")
        except ValueError:
            print(f"    Enter a valid number")

    print(f"  Selected: {os.path.basename(input_path)}")

    # --- Prompt 4: Symbol selection ---
    # Load Final_Alphas to show available symbols
    try:
        df_alphas = load_final_alphas(input_path)
    except Exception as e:
        print(f"\n  ERROR: Failed to read Final_Alphas sheet: {e}")
        return None

    available_symbols = sorted(df_alphas['Symbol'].unique().tolist())
    total_strategies = len(df_alphas)

    print(f"\n  Symbols found in Final_Alphas ({total_strategies} strategies):")
    print(f"    {', '.join(available_symbols)}")

    symbols_filter = None
    while True:
        sym_input = input("  Process [A]ll symbols or specify (comma-separated) [A]: ").strip()
        if sym_input == '' or sym_input.upper() == 'A':
            symbols_filter = None
            break

        requested = [s.strip().upper() for s in sym_input.split(',') if s.strip()]
        invalid = [s for s in requested if s not in available_symbols]
        if invalid:
            print(f"    Unknown symbols: {', '.join(invalid)}")
            print(f"    Available: {', '.join(available_symbols)}")
            continue

        if not requested:
            print(f"    No valid symbols entered")
            continue

        symbols_filter = requested
        break

    # Count strategies for selected symbols
    if symbols_filter:
        n_strategies = len(df_alphas[df_alphas['Symbol'].isin(symbols_filter)])
        sym_display = ', '.join(symbols_filter) + f" ({n_strategies} strategies)"
    else:
        n_strategies = total_strategies
        sym_display = f"All ({n_strategies} strategies)"

    # --- Prompt 5: Confirmation ---
    print()
    print("  " + "-" * 56)
    print("  Configuration Summary:")
    print(f"    Exchange:    {exchange}")
    print(f"    Interval:    {interval}")
    print(f"    Input:       {input_path}")
    print(f"    Symbols:     {sym_display}")
    print(f"    Grid:        Phase1={len(SL_RANGE)*len(TP_RANGE)} | "
          f"Phase2A={len(TSL_RANGE)*len(TP_RANGE)} | Phase2B=21-42")
    print("  " + "-" * 56)

    confirm = input("  Proceed? (y/n) [y]: ").strip().lower() or 'y'
    if confirm != 'y':
        return None

    return {
        'exchange': exchange,
        'interval': interval,
        'input_path': input_path,
        'symbols': symbols_filter,
    }


# ============================================================================
# INPUT DISCOVERY
# ============================================================================

def find_latest_final_compilation(exchange, interval):
    """Find the most recent Final_Compilation Excel via glob."""
    pattern = f"WFAlphaResults/Final_Compilation_{exchange}_{interval}_*.xlsx"
    matches = sorted(glob.glob(pattern))
    if not matches:
        return None
    return matches[-1]  # Latest by filename date


def load_final_alphas(excel_path):
    """Load the Final_Alphas sheet from a Final_Compilation Excel."""
    df = pd.read_excel(excel_path, sheet_name='Final_Alphas')

    # Validate required columns
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise KeyError(f"Final_Alphas sheet missing required columns: {missing}")

    return df

# ============================================================================
# DATA LOADING
# ============================================================================

# Cache for GridSearch_data CSV row counts (per symbol)
_row_count_cache = {}


def get_total_rows_cached(symbol, interval, exchange):
    """
    Get total row count from the original GridSearch_data CSV.
    Results are cached per symbol to avoid re-reading.

    Returns:
    --------
    int or None : Row count, or None if CSV not found
    """
    cache_key = f"{exchange}_{symbol}_{interval}"
    if cache_key in _row_count_cache:
        return _row_count_cache[cache_key]

    pattern = f"GridSearch_data/merged_{exchange}_{symbol}_{interval}_*.csv"
    matches = glob.glob(pattern)
    if not matches:
        return None

    csv_path = matches[-1]
    # Read only first column to count rows efficiently
    try:
        row_count = len(pd.read_csv(csv_path, usecols=[0]))
    except Exception:
        return None

    _row_count_cache[cache_key] = row_count
    return row_count


def construct_strategy_folder(row, exchange):
    """
    Construct the WFAlphaResults folder path for a strategy row.

    Path: WFAlphaResults/merged_{exchange}_{Symbol}_{Interval}_{Variant}/{Data_Point}/{Model}/{Entry_Exit_Model}/
    """
    symbol = row['Symbol']
    interval = row['Interval'] if 'Interval' in row.index else '1h'
    variant = row['Variant'] if 'Variant' in row.index and pd.notna(row.get('Variant')) else None
    data_point = row['Data Point']
    model = row['Model']
    entry_exit = row['Entry / Exit Model']

    # Find matching folder via glob (variant may have date suffix)
    if variant:
        folder_pattern = f"WFAlphaResults/merged_{exchange}_{symbol}_{interval}_{variant}"
    else:
        folder_pattern = f"WFAlphaResults/merged_{exchange}_{symbol}_{interval}_*"

    matches = glob.glob(folder_pattern)
    if not matches:
        return None

    base_folder = matches[0]
    strategy_folder = os.path.join(base_folder, data_point, model, entry_exit)
    return strategy_folder


def load_backtest_data(strategy_folder):
    """Load backtest.csv from a strategy folder with timezone normalization."""
    backtest_path = os.path.join(strategy_folder, 'backtest.csv')
    if not os.path.exists(backtest_path):
        return None

    df = pd.read_csv(backtest_path)
    if 'datetime' in df.columns:
        try:
            df['datetime'] = pd.to_datetime(df['datetime'], utc=True).dt.tz_convert(None)
        except Exception:
            df['datetime'] = pd.to_datetime(df['datetime'])
    return df


def load_baseline_metrics(strategy_folder):
    """Load metrics.csv (Metric/Value format) from a strategy folder."""
    metrics_path = os.path.join(strategy_folder, 'metrics.csv')
    if not os.path.exists(metrics_path):
        return None

    try:
        df = pd.read_csv(metrics_path)
        metrics = dict(zip(df['Metric'], df['Value']))
        return metrics
    except Exception:
        return None

# ============================================================================
# BASELINE VALIDATION
# ============================================================================

def validate_baseline(computed_metrics, stored_metrics, strategy_id):
    """
    Compare computed baseline metrics against stored Stage 7 metrics.

    Returns:
    --------
    list of str : Warning messages (empty if all match within tolerance)
    """
    warnings_list = []
    if stored_metrics is None:
        warnings_list.append(f"  [{strategy_id}] metrics.csv not found — skipping baseline validation")
        return warnings_list

    check_keys = ['Sharpe Ratio', 'Max Drawdown', 'Trade Count', 'Annualized Return']
    for key in check_keys:
        computed_val = computed_metrics.get(key, 0)
        stored_val = float(stored_metrics.get(key, 0))

        if stored_val == 0:
            if abs(computed_val) > BASELINE_TOLERANCE:
                warnings_list.append(
                    f"  [{strategy_id}] {key}: computed={computed_val:.6f}, stored=0.0"
                )
        else:
            pct_diff = abs((computed_val - stored_val) / stored_val)
            if pct_diff > BASELINE_TOLERANCE:
                warnings_list.append(
                    f"  [{strategy_id}] {key}: computed={computed_val:.6f}, stored={stored_val:.6f} "
                    f"(diff={pct_diff*100:.2f}%)"
                )

    return warnings_list

# ============================================================================
# GRID SEARCH PHASES
# ============================================================================

def _test_sl_tp(df_slim, fee, period, sl, tp):
    """Test a single SL/TP combination. Used by joblib workers."""
    _, metrics = calculate_with_risk_params(df_slim, fee, period, sl_pct=sl, tp_pct=tp, tsl_pct=0)
    return {'SL': sl, 'TP': tp, 'TSL': 0, **metrics}


def _test_tsl_tp(df_slim, fee, period, tsl, tp):
    """Test a single TSL/TP combination. Used by joblib workers."""
    _, metrics = calculate_with_risk_params(df_slim, fee, period, sl_pct=0, tp_pct=tp, tsl_pct=tsl)
    return {'SL': 0, 'TP': tp, 'TSL': tsl, **metrics}


def _test_bracket(df_slim, fee, period, sl, tp, tsl):
    """Test a single bracket combination. Used by joblib workers."""
    _, metrics = calculate_with_risk_params(df_slim, fee, period, sl_pct=sl, tp_pct=tp, tsl_pct=tsl)
    return {'SL': sl, 'TP': tp, 'TSL': tsl, **metrics}


def optimize_phase1_sl_tp(df_slim, fee, period, min_trades, baseline_metrics, n_jobs):
    """
    Phase 1: Grid search over SL × TP combinations.

    Returns:
    --------
    dict with keys: best (Series or None), grid_df (DataFrame), phase_name
    """
    combos = [(sl, tp) for sl in SL_RANGE for tp in TP_RANGE]

    results = Parallel(n_jobs=n_jobs, prefer="threads")(
        delayed(_test_sl_tp)(df_slim, fee, period, sl, tp)
        for sl, tp in combos
    )

    grid_df = pd.DataFrame(results)
    filtered = filter_by_min_trade_count(grid_df, min_trades)
    best = select_best_combination(filtered, tiebreak_cols=['SL', 'TP'])

    return {'best': best, 'grid_df': grid_df, 'phase_name': 'SL/TP'}


def optimize_phase2a_tsl_tp(df_slim, fee, period, min_trades, baseline_metrics, n_jobs):
    """
    Phase 2A: Grid search over TSL × TP combinations (no SL).

    Returns:
    --------
    dict with keys: best (Series or None), grid_df (DataFrame), phase_name
    """
    combos = [(tsl, tp) for tsl in TSL_RANGE for tp in TP_RANGE]

    results = Parallel(n_jobs=n_jobs, prefer="threads")(
        delayed(_test_tsl_tp)(df_slim, fee, period, tsl, tp)
        for tsl, tp in combos
    )

    grid_df = pd.DataFrame(results)
    filtered = filter_by_min_trade_count(grid_df, min_trades)
    best = select_best_combination(filtered, tiebreak_cols=['TSL', 'TP'])

    return {'best': best, 'grid_df': grid_df, 'phase_name': 'TSL/TP'}


def optimize_phase2b_bracket(df_slim, fee, period, sl_p1, tp_p1, tp_2a,
                              min_trades, baseline_metrics, n_jobs):
    """
    Phase 2B: Bracket orders — SL from Phase 1 + TP candidates + all TSL values.

    Smart TP check: if |tp_p1 - tp_2a| < 0.001, only test tp_p1 (halves work).

    Returns:
    --------
    dict with keys: best (Series or None), grid_df (DataFrame), phase_name
    """
    tp_candidates = [tp_p1]
    if abs(tp_p1 - tp_2a) >= 0.001:
        tp_candidates.append(tp_2a)

    combos = [(sl_p1, tp, tsl) for tp in tp_candidates for tsl in TSL_RANGE]

    results = Parallel(n_jobs=n_jobs, prefer="threads")(
        delayed(_test_bracket)(df_slim, fee, period, sl, tp, tsl)
        for sl, tp, tsl in combos
    )

    grid_df = pd.DataFrame(results)
    filtered = filter_by_min_trade_count(grid_df, min_trades)
    best = select_best_combination(filtered, tiebreak_cols=['TSL', 'TP', 'SL'])

    return {'best': best, 'grid_df': grid_df, 'phase_name': 'Bracket'}

# ============================================================================
# BEST APPROACH SELECTION
# ============================================================================

def select_best_approach(baseline_metrics, phase1, phase2a, phase2b, min_trades):
    """
    Select the single best approach across Baseline, SL/TP, TSL/TP, and Bracket.

    Priority order (tie-breaking): Baseline > SL/TP > TSL/TP > Bracket
    Baseline always eligible (bypasses min_trades filter).

    Returns:
    --------
    dict: selected approach info with keys:
        approach, SL, TP, TSL, Sharpe Ratio, Max Drawdown, Trade Count,
        Annualized Return, Calmar Ratio
    """
    candidates = []

    # Candidate 1: Baseline (always eligible)
    candidates.append({
        'approach': 'Baseline',
        'priority': 1,
        'SL': 0.0, 'TP': 0.0, 'TSL': 0.0,
        **baseline_metrics
    })

    # Candidate 2: Phase 1 best (SL/TP)
    if phase1['best'] is not None:
        p1 = phase1['best']
        if p1['Trade Count'] >= min_trades:
            candidates.append({
                'approach': 'SL/TP',
                'priority': 2,
                'SL': p1['SL'], 'TP': p1['TP'], 'TSL': 0.0,
                'Sharpe Ratio': p1['Sharpe Ratio'],
                'Max Drawdown': p1['Max Drawdown'],
                'Trade Count': p1['Trade Count'],
                'Annualized Return': p1['Annualized Return'],
                'Calmar Ratio': p1['Calmar Ratio'],
            })

    # Candidate 3: Phase 2A best (TSL/TP)
    if phase2a['best'] is not None:
        p2a = phase2a['best']
        if p2a['Trade Count'] >= min_trades:
            candidates.append({
                'approach': 'TSL/TP',
                'priority': 3,
                'SL': 0.0, 'TP': p2a['TP'], 'TSL': p2a['TSL'],
                'Sharpe Ratio': p2a['Sharpe Ratio'],
                'Max Drawdown': p2a['Max Drawdown'],
                'Trade Count': p2a['Trade Count'],
                'Annualized Return': p2a['Annualized Return'],
                'Calmar Ratio': p2a['Calmar Ratio'],
            })

    # Candidate 4: Phase 2B best (Bracket)
    if phase2b['best'] is not None:
        p2b = phase2b['best']
        if p2b['Trade Count'] >= min_trades:
            candidates.append({
                'approach': 'Bracket',
                'priority': 4,
                'SL': p2b['SL'], 'TP': p2b['TP'], 'TSL': p2b['TSL'],
                'Sharpe Ratio': p2b['Sharpe Ratio'],
                'Max Drawdown': p2b['Max Drawdown'],
                'Trade Count': p2b['Trade Count'],
                'Annualized Return': p2b['Annualized Return'],
                'Calmar Ratio': p2b['Calmar Ratio'],
            })

    # Select: max Sharpe, then lowest priority number (simpler approach)
    candidates_df = pd.DataFrame(candidates)
    candidates_df = candidates_df.sort_values(
        by=['Sharpe Ratio', 'priority'],
        ascending=[False, True]
    )

    return candidates_df.iloc[0].to_dict()

# ============================================================================
# CSV OUTPUT (PER-STRATEGY)
# ============================================================================

def save_phase_csvs(strategy_folder, prefix, phase_result, df_full, fee, period):
    """
    Save 4 CSV files for a phase: backtest, grid results, metrics, summary.

    Parameters:
    -----------
    strategy_folder : str
        Path to strategy folder
    prefix : str
        File prefix: 'sl', 'tsl', or 'bracket'
    phase_result : dict
        Phase result with 'best' (Series) and 'grid_df' (DataFrame)
    df_full : pd.DataFrame
        Full backtest DataFrame (for generating backtest CSV with optimal params)
    fee : float
        Transaction fee
    period : int
        Annualization period
    """
    best = phase_result['best']
    grid_df = phase_result['grid_df']

    # Grid results CSV
    grid_path = os.path.join(strategy_folder, f'grid_{prefix}_results.csv')
    grid_df.to_csv(grid_path, index=False)

    if best is None:
        # No valid combination found — save empty metrics/summary
        metrics_path = os.path.join(strategy_folder, f'{prefix}_metrics.csv')
        pd.DataFrame({'Metric': ['Note'], 'Value': ['No valid combination found']}).to_csv(metrics_path, index=False)

        summary_path = os.path.join(strategy_folder, f'{prefix}_summary.csv')
        pd.DataFrame([{'SL': 0, 'TP': 0, 'TSL': 0, 'Sharpe Ratio': 0, 'Note': 'No valid combination'}]).to_csv(summary_path, index=False)

        # Empty backtest CSV
        backtest_path = os.path.join(strategy_folder, f'{prefix}_backtest.csv')
        pd.DataFrame().to_csv(backtest_path, index=False)
        return

    sl_val = best.get('SL', 0)
    tp_val = best.get('TP', 0)
    tsl_val = best.get('TSL', 0)

    # Backtest CSV with optimal params — pass full DataFrame so all columns are preserved
    df_bt, _ = calculate_with_risk_params(df_full.copy(), fee, period, sl_pct=sl_val, tp_pct=tp_val, tsl_pct=tsl_val, track_risk_levels=True)

    backtest_path = os.path.join(strategy_folder, f'{prefix}_backtest.csv')
    df_bt.to_csv(backtest_path, index=False)

    # Metrics CSV (Metric/Value format, same as Stage 7)
    metrics_data = {
        'Metric': ['Sharpe Ratio', 'Max Drawdown', 'Trade Count', 'Annualized Return',
                    'Calmar Ratio', 'SL', 'TP', 'TSL'],
        'Value': [best['Sharpe Ratio'], best['Max Drawdown'], best['Trade Count'],
                  best['Annualized Return'], best['Calmar Ratio'], sl_val, tp_val, tsl_val]
    }
    metrics_path = os.path.join(strategy_folder, f'{prefix}_metrics.csv')
    pd.DataFrame(metrics_data).to_csv(metrics_path, index=False)

    # Summary CSV (single-row)
    summary_row = {
        'SL': sl_val, 'TP': tp_val, 'TSL': tsl_val,
        'Sharpe Ratio': best['Sharpe Ratio'],
        'Max Drawdown': best['Max Drawdown'],
        'Trade Count': best['Trade Count'],
        'Annualized Return': best['Annualized Return'],
        'Calmar Ratio': best['Calmar Ratio'],
    }
    summary_path = os.path.join(strategy_folder, f'{prefix}_summary.csv')
    pd.DataFrame([summary_row]).to_csv(summary_path, index=False)

# ============================================================================
# PORTFOLIO SHEET CONSTRUCTION
# ============================================================================

def build_risk_optimized_portfolios(all_pnl_data, interval):
    """
    Build risk-optimized portfolios per symbol using equal-weight combination.
    Mirrors Stage 7 portfolio construction (OUTER merge, equal-weight mean).

    Parameters:
    -----------
    all_pnl_data : list of (strategy_id, symbol, pnl_df)
        Each pnl_df has columns ['datetime', 'pnl']
    interval : str
        Data interval for annualization period

    Returns:
    --------
    dict : {symbol: {'sharpe': float, 'mdd': float, 'annual_return': float,
                      'calmar': float, 'cumulative_pnl': float}}
    """
    period = get_period_for_interval(interval)

    # Group by symbol
    symbol_data = {}
    for strategy_id, symbol, pnl_df in all_pnl_data:
        if pnl_df is None or pnl_df.empty:
            continue
        if symbol not in symbol_data:
            symbol_data[symbol] = []
        symbol_data[symbol].append((strategy_id, pnl_df))

    results = {}
    for symbol, strategies in symbol_data.items():
        if not strategies:
            continue

        # Build merged DataFrame via OUTER merge on datetime (same as Stage 7)
        merged = None
        strategy_cols = []
        for strategy_id, pnl_df in strategies:
            col_name = strategy_id
            df_renamed = pnl_df[['datetime', 'pnl']].rename(columns={'pnl': col_name})
            if merged is None:
                merged = df_renamed
            else:
                merged = pd.merge(merged, df_renamed, on='datetime', how='outer')
            strategy_cols.append(col_name)

        merged = merged.sort_values('datetime').reset_index(drop=True)
        merged[strategy_cols] = merged[strategy_cols].fillna(0)

        # Equal-weight portfolio
        portfolio_pnl = merged[strategy_cols].mean(axis=1)
        cumulative_pnl = portfolio_pnl.cumsum()
        drawdown = cumulative_pnl - cumulative_pnl.cummax()

        # Metrics (same formulas as Stage 7)
        pnl_std = portfolio_pnl.std()  # ddof=1 (pandas default)
        sharpe = (portfolio_pnl.mean() / pnl_std) * np.sqrt(period) if pnl_std != 0 else 0
        mdd = drawdown.min()
        annual_return = portfolio_pnl.mean() * period
        calmar = annual_return / abs(mdd) if mdd != 0 else 0
        cumulative_pnl_final = cumulative_pnl.iloc[-1]

        results[symbol] = {
            'sharpe': sharpe,
            'mdd': mdd,
            'annual_return': annual_return,
            'calmar': calmar,
            'cumulative_pnl': cumulative_pnl_final,
        }

    return results


def build_portfolio_sheet(df_corr_summary, risk_opt_metrics):
    """
    Build the Portfolio sheet DataFrame with 22 columns.

    Parameters:
    -----------
    df_corr_summary : pd.DataFrame
        Corr Summary from Final_Compilation (15 columns)
    risk_opt_metrics : dict
        {symbol: {'sharpe', 'mdd', 'annual_return', 'calmar', 'cumulative_pnl'}}

    Returns:
    --------
    pd.DataFrame : Portfolio sheet with 22 columns
    """
    rows = []
    for _, row in df_corr_summary.iterrows():
        symbol = row['Symbol']
        rm = risk_opt_metrics.get(symbol, {})
        bnh = row['Buy & Hold']
        cum_pnl = rm.get('cumulative_pnl', np.nan)
        pnl_ratio = cum_pnl / bnh if (bnh != 0 and not np.isnan(bnh) and not np.isnan(cum_pnl)) else np.nan

        # Warn if risk-optimized Sharpe is lower than baseline portfolio Sharpe
        risk_sharpe = rm.get('sharpe', np.nan)
        base_sharpe = row['Portfolio Sharpe']
        if not np.isnan(risk_sharpe) and not np.isnan(base_sharpe) and risk_sharpe < base_sharpe:
            print(f"  WARNING: {symbol} risk-opt portfolio Sharpe ({risk_sharpe:.4f}) < baseline ({base_sharpe:.4f})")

        rows.append([
            row['Symbol'],                      # A
            row['Threshold'],                   # B
            row['Total Strategies'],            # C
            row['# Strategies Selected'],       # D
            rm.get('sharpe', np.nan),           # E
            rm.get('mdd', np.nan),              # F
            rm.get('annual_return', np.nan),     # G
            rm.get('calmar', np.nan),           # H
            cum_pnl,                            # I
            bnh,                                # J (Buy & Hold)
            pnl_ratio,                          # K
            row['Portfolio Sharpe'],            # L
            row['Portfolio MDD'],               # M
            row['Portfolio Annual Return'],     # N
            row['Portfolio Calmar Ratio'],      # O
            row['Cumulative PnL'],              # P
            bnh,                                # Q (Buy & Hold — same value, duplicate name)
            row['PnL Ratio'],                   # R
            row['Total Rows'],                  # S
            row['First Datetime'],              # T
            row['Last Datetime'],               # U
            row['Strategy List'],               # V
        ])

    columns = [
        'Symbol', 'Threshold', 'Total Strategies', '# Strategies Selected',
        'Risk Opt Portfolio Sharpe', 'Risk Opt Portfolio MDD',
        'Risk Opt Portfolio Annual Return', 'Risk Opt Portfolio Calmar Ratio',
        'Risk Opt Cumulative PnL', 'Buy & Hold', 'Risk Opt PnL Ratio',
        'Portfolio Sharpe', 'Portfolio MDD', 'Portfolio Annual Return',
        'Portfolio Calmar Ratio', 'Cumulative PnL', 'Buy & Hold',
        'PnL Ratio', 'Total Rows', 'First Datetime', 'Last Datetime', 'Strategy List'
    ]
    portfolio_df = pd.DataFrame(rows, columns=columns)
    return portfolio_df


# ============================================================================
# CONSOLIDATED EXCEL OUTPUT
# ============================================================================

def save_consolidated_excel(all_results, comparison_rows, exchange, interval, output_path=None, portfolio_df=None):
    """
    Generate the Risk_Optimized_Compilation Excel with up to 4 sheets.

    Parameters:
    -----------
    all_results : list of dict
        One dict per strategy with selected approach info
    comparison_rows : list of dict
        4 dicts per strategy (Baseline, SL/TP, TSL/TP, Bracket)
    exchange : str
    interval : str
    output_path : str or None
        Explicit path, or auto-generate
    portfolio_df : pd.DataFrame or None
        Portfolio sheet DataFrame (22 columns). If None, Portfolio sheet is omitted.
    """
    if not all_results:
        print("  WARNING: No successful strategies to compile — skipping Excel output")
        return None

    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d')
        output_path = f"WFAlphaResults/Risk_Optimized_Compilation_{exchange}_{interval}_{timestamp}.xlsx"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Apply column renames to Risk_Optimized_Strategies
    df_main = pd.DataFrame(all_results)
    rename_map = {'Data_Point': 'Data Point', 'Entry_Exit': 'Entry / Exit Model'}
    df_main = df_main.rename(columns=rename_map)

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Portfolio (if available)
            if portfolio_df is not None:
                portfolio_df.to_excel(writer, sheet_name='Portfolio', index=False)

            # Sheet 2: Risk_Optimized_Strategies
            df_main.to_excel(writer, sheet_name='Risk_Optimized_Strategies', index=False)

            # Sheet 3: Risk_Approach_Comparison
            df_compare = pd.DataFrame(comparison_rows)
            df_compare.to_excel(writer, sheet_name='Risk_Approach_Comparison', index=False)

            # Sheet 4: Aggregated_Statistics
            _write_aggregated_stats(writer, df_main, df_compare)

        # Apply conditional formatting (green highlight for best approach per strategy)
        _apply_excel_formatting(output_path)

        print(f"\n  Consolidated Excel saved: {output_path}")
        return output_path

    except Exception as e:
        print(f"  WARNING: Excel formatting failed ({e}) — saving without formatting")
        # Fallback: save without formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            if portfolio_df is not None:
                portfolio_df.to_excel(writer, sheet_name='Portfolio', index=False)
            df_main.to_excel(writer, sheet_name='Risk_Optimized_Strategies', index=False)
            pd.DataFrame(comparison_rows).to_excel(writer, sheet_name='Risk_Approach_Comparison', index=False)
        return output_path


def _write_aggregated_stats(writer, df_main, df_compare):
    """Write the Aggregated_Statistics sheet with cross-strategy summary."""
    stats_rows = []

    # Section 1: Performance Summary
    stats_rows.append({'Section': '1. Performance Summary', 'Metric': '', 'Value': ''})
    for col in ['Optimized_Sharpe', 'Optimized_MDD', 'Optimized_Annual_Return', 'Optimized_Trade_Count']:
        if col in df_main.columns:
            stats_rows.append({'Section': '', 'Metric': f'Avg {col}', 'Value': df_main[col].mean()})

    # Section 2: Improvement Analysis
    stats_rows.append({'Section': '2. Improvement Analysis', 'Metric': '', 'Value': ''})
    if 'Sharpe_Improvement_%' in df_main.columns:
        stats_rows.append({'Section': '', 'Metric': 'Avg Sharpe Improvement %',
                           'Value': df_main['Sharpe_Improvement_%'].mean()})
    if 'MDD_Improvement_%' in df_main.columns:
        stats_rows.append({'Section': '', 'Metric': 'Avg MDD Improvement %',
                           'Value': df_main['MDD_Improvement_%'].mean()})

    # Section 3: Approach Distribution
    stats_rows.append({'Section': '3. Approach Distribution', 'Metric': '', 'Value': ''})
    if 'Selected_Approach' in df_main.columns:
        dist = df_main['Selected_Approach'].value_counts()
        for approach, count in dist.items():
            stats_rows.append({'Section': '', 'Metric': approach, 'Value': count})

    # Section 4: Optimal Parameter Ranges
    stats_rows.append({'Section': '4. Optimal Parameter Ranges', 'Metric': '', 'Value': ''})
    for param in ['Optimal_SL', 'Optimal_TP', 'Optimal_TSL']:
        if param in df_main.columns:
            vals = df_main[param]
            stats_rows.append({'Section': '', 'Metric': f'{param} Min', 'Value': vals.min()})
            stats_rows.append({'Section': '', 'Metric': f'{param} Median', 'Value': vals.median()})
            stats_rows.append({'Section': '', 'Metric': f'{param} Max', 'Value': vals.max()})

    # Section 5: Exit Type Breakdown
    stats_rows.append({'Section': '5. Exit Type Breakdown', 'Metric': '', 'Value': ''})
    for col in ['SL_Exits', 'TP_Exits', 'TSL_Exits', 'Signal_Exits']:
        if col in df_main.columns:
            stats_rows.append({'Section': '', 'Metric': f'Total {col}', 'Value': df_main[col].sum()})

    pd.DataFrame(stats_rows).to_excel(writer, sheet_name='Aggregated_Statistics', index=False)


def _apply_excel_formatting(output_path):
    """Apply conditional formatting to the Excel file (green highlight for best approaches)."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(output_path)
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        # Highlight best approach per strategy in Risk_Approach_Comparison
        if 'Risk_Approach_Comparison' in wb.sheetnames:
            ws = wb['Risk_Approach_Comparison']
            headers = [cell.value for cell in ws[1]]

            strategy_col = headers.index('Strategy_ID') + 1 if 'Strategy_ID' in headers else None
            sharpe_col = headers.index('Sharpe_Ratio') + 1 if 'Sharpe_Ratio' in headers else None

            if strategy_col and sharpe_col:
                # Group rows by strategy, find max Sharpe in each group
                current_strategy = None
                group_rows = []
                all_groups = []

                for row_idx in range(2, ws.max_row + 1):
                    sid = ws.cell(row=row_idx, column=strategy_col).value
                    if sid != current_strategy:
                        if group_rows:
                            all_groups.append(group_rows)
                        group_rows = []
                        current_strategy = sid
                    group_rows.append(row_idx)
                if group_rows:
                    all_groups.append(group_rows)

                for group in all_groups:
                    best_row = None
                    best_sharpe = -float('inf')
                    for row_idx in group:
                        val = ws.cell(row=row_idx, column=sharpe_col).value
                        if val is not None and val > best_sharpe:
                            best_sharpe = val
                            best_row = row_idx
                    if best_row:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=best_row, column=col_idx).fill = green_fill

        wb.save(output_path)
    except Exception:
        pass  # Non-critical — formatting failure is acceptable

# ============================================================================
# ERROR LOGGING
# ============================================================================

def log_error(errors_list, strategy_id, error_type, error_message):
    """Append an error to the errors list for later CSV output."""
    errors_list.append({
        'Strategy_ID': strategy_id,
        'Error_Type': error_type,
        'Error_Message': str(error_message),
        'Timestamp': datetime.now().isoformat()
    })


def save_errors(errors_list):
    """Save accumulated errors to stage8_errors.csv."""
    if errors_list:
        df = pd.DataFrame(errors_list)
        error_path = 'WFAlphaResults/stage8_errors.csv'
        df.to_csv(error_path, index=False)
        print(f"  Errors logged to: {error_path}")

# ============================================================================
# MAIN PROCESSING
# ============================================================================

def build_strategy_id(row):
    """Build a unique strategy identifier from row data."""
    parts = [str(row.get('Symbol', '')), str(row.get('Data Point', '')),
             str(row.get('Model', '')), str(row.get('Entry / Exit Model', ''))]
    if 'Variant' in row.index and pd.notna(row.get('Variant')):
        parts.append(str(row['Variant']))
    return '_'.join(parts)


def count_exit_types(strategy_folder, prefix, best_params, df_full, fee, period):
    """Count exit types from the optimal backtest for a phase."""
    sl_val = best_params.get('SL', 0)
    tp_val = best_params.get('TP', 0)
    tsl_val = best_params.get('TSL', 0)

    df_slim = df_full[['signal', 'open', 'high', 'low', 'close']].copy()
    df_bt, _ = calculate_with_risk_params(df_slim, fee, period, sl_pct=sl_val, tp_pct=tp_val, tsl_pct=tsl_val)

    counts = {'SL_Exits': 0, 'TP_Exits': 0, 'TSL_Exits': 0, 'Signal_Exits': 0}
    if 'exit_type' in df_bt.columns:
        exit_counts = df_bt['exit_type'].value_counts()
        counts['SL_Exits'] = int(exit_counts.get('sl_exit', 0))
        counts['TP_Exits'] = int(exit_counts.get('tp_exit', 0))
        counts['TSL_Exits'] = int(exit_counts.get('tsl_exit', 0))
        counts['Signal_Exits'] = int(exit_counts.get('signal_exit', 0))
    return counts


def process_strategy(row, exchange, interval, period, n_jobs, errors_list):
    """
    Process a single strategy: load data, run 3-phase grid search, save outputs.

    Returns:
    --------
    tuple: (result_dict, comparison_rows, risk_pnl_df) or (None, None, None) on failure
        risk_pnl_df has columns ['datetime', 'pnl'] for portfolio construction
    """
    strategy_id = build_strategy_id(row)
    symbol = row['Symbol']

    print(f"\n  Processing: {strategy_id}")

    # --- LOAD DATA ---
    strategy_folder = construct_strategy_folder(row, exchange)
    if strategy_folder is None or not os.path.isdir(strategy_folder):
        log_error(errors_list, strategy_id, 'FileNotFoundError', f'Strategy folder not found')
        print(f"    SKIP: Strategy folder not found")
        return None, None, None

    df_full = load_backtest_data(strategy_folder)
    if df_full is None or df_full.empty:
        log_error(errors_list, strategy_id, 'EmptyDataFrameError', 'backtest.csv not found or empty')
        print(f"    SKIP: backtest.csv not found or empty")
        return None, None, None

    # Check required columns
    required_cols = ['signal', 'open', 'high', 'low', 'close']
    missing_cols = [c for c in required_cols if c not in df_full.columns]
    if missing_cols:
        log_error(errors_list, strategy_id, 'KeyError', f'Missing columns in backtest.csv: {missing_cols}')
        print(f"    SKIP: Missing columns: {missing_cols}")
        return None, None, None

    # Check for non-zero signals
    if df_full['signal'].abs().sum() == 0:
        log_error(errors_list, strategy_id, 'ZeroTradesError', 'All signals are 0')
        print(f"    SKIP: All signals are 0 — no trades to optimize")
        return None, None, None

    # --- MIN TRADES ---
    total_rows = get_total_rows_cached(symbol, interval, exchange)
    if total_rows is None:
        log_error(errors_list, strategy_id, 'FileNotFoundError', f'GridSearch_data CSV not found for {symbol}')
        print(f"    SKIP: GridSearch_data CSV not found for {symbol}")
        return None, None, None

    min_trades = int(total_rows * TRADE_COUNT_THRESHOLD)

    # --- BASELINE ---
    df_slim = df_full[['signal', 'open', 'high', 'low', 'close']].copy()
    _, baseline_metrics = calculate_with_risk_params(df_slim.copy(), FEE, period, 0, 0, 0)

    stored_metrics = load_baseline_metrics(strategy_folder)
    validation_warnings = validate_baseline(baseline_metrics, stored_metrics, strategy_id)
    for w in validation_warnings:
        print(f"    WARNING: {w}")

    print(f"    Baseline Sharpe: {baseline_metrics['Sharpe Ratio']:.4f}  |  min_trades: {min_trades}  |  rows: {len(df_slim)}")

    # --- PHASE 1: SL/TP ---
    t0 = time.time()
    phase1 = optimize_phase1_sl_tp(df_slim.copy(), FEE, period, min_trades, baseline_metrics, n_jobs)
    p1_sharpe = phase1['best']['Sharpe Ratio'] if phase1['best'] is not None else 'N/A'
    print(f"    Phase 1 (SL/TP): {len(phase1['grid_df'])} combos -> best Sharpe={p1_sharpe}  [{time.time()-t0:.1f}s]")

    # --- PHASE 2A: TSL/TP ---
    t0 = time.time()
    phase2a = optimize_phase2a_tsl_tp(df_slim.copy(), FEE, period, min_trades, baseline_metrics, n_jobs)
    p2a_sharpe = phase2a['best']['Sharpe Ratio'] if phase2a['best'] is not None else 'N/A'
    print(f"    Phase 2A (TSL/TP): {len(phase2a['grid_df'])} combos -> best Sharpe={p2a_sharpe}  [{time.time()-t0:.1f}s]")

    # --- PHASE 2B: BRACKET ---
    t0 = time.time()
    sl_p1 = phase1['best']['SL'] if phase1['best'] is not None else 0
    tp_p1 = phase1['best']['TP'] if phase1['best'] is not None else 0
    tp_2a = phase2a['best']['TP'] if phase2a['best'] is not None else 0

    phase2b = optimize_phase2b_bracket(df_slim.copy(), FEE, period, sl_p1, tp_p1, tp_2a,
                                        min_trades, baseline_metrics, n_jobs)
    p2b_sharpe = phase2b['best']['Sharpe Ratio'] if phase2b['best'] is not None else 'N/A'
    print(f"    Phase 2B (Bracket): {len(phase2b['grid_df'])} combos -> best Sharpe={p2b_sharpe}  [{time.time()-t0:.1f}s]")

    # --- BEST APPROACH SELECTION ---
    best = select_best_approach(baseline_metrics, phase1, phase2a, phase2b, min_trades)
    print(f"    Selected: {best['approach']} (Sharpe={best['Sharpe Ratio']:.4f})")

    # --- COMPUTE RISK-OPTIMIZED PnL FOR PORTFOLIO ---
    df_for_risk = df_full[['datetime', 'signal', 'open', 'high', 'low', 'close']].copy()
    df_risk_bt, _ = calculate_with_risk_params(
        df_for_risk, FEE, period,
        sl_pct=best['SL'], tp_pct=best['TP'], tsl_pct=best['TSL']
    )
    risk_pnl_df = df_risk_bt[['datetime', 'pnl']].copy()

    # --- SAVE 12 CSV FILES ---
    save_phase_csvs(strategy_folder, 'sl', phase1, df_full, FEE, period)
    save_phase_csvs(strategy_folder, 'tsl', phase2a, df_full, FEE, period)
    save_phase_csvs(strategy_folder, 'bracket', phase2b, df_full, FEE, period)

    # --- COUNT EXIT TYPES ---
    exit_counts = count_exit_types(strategy_folder, best['approach'], best, df_full, FEE, period)

    # --- BUILD RESULT DICT ---
    base_sharpe = baseline_metrics['Sharpe Ratio']
    opt_sharpe = best['Sharpe Ratio']
    sharpe_improvement = ((opt_sharpe - base_sharpe) / abs(base_sharpe) * 100) if base_sharpe != 0 else 0

    base_mdd = baseline_metrics['Max Drawdown']
    opt_mdd = best['Max Drawdown']
    mdd_improvement = ((opt_mdd - base_mdd) / abs(base_mdd) * 100) if base_mdd != 0 else 0

    result = {
        'Strategy_ID': strategy_id,
        'Symbol': symbol,
        'Data_Point': row.get('Data Point', ''),
        'Model': row.get('Model', ''),
        'Entry_Exit': row.get('Entry / Exit Model', ''),
        'Selected_Approach': best['approach'],
        'Optimal_SL': best['SL'],
        'Optimal_TP': best['TP'],
        'Optimal_TSL': best['TSL'],
        'Optimized_Sharpe': opt_sharpe,
        'Optimized_MDD': opt_mdd,
        'Optimized_Annual_Return': best['Annualized Return'],
        'Optimized_Trade_Count': best['Trade Count'],
        'Optimized_Calmar': best.get('Calmar Ratio', 0),
        'Base_Sharpe': base_sharpe,
        'Base_MDD': base_mdd,
        'Sharpe_Improvement_%': sharpe_improvement,
        'MDD_Improvement_%': mdd_improvement,
        **exit_counts,
    }

    # --- BUILD COMPARISON ROWS (4 per strategy) ---
    comparison = []
    approaches_data = [
        ('Baseline', baseline_metrics, 0, 0, 0),
    ]
    if phase1['best'] is not None:
        p1b = phase1['best']
        approaches_data.append(('SL/TP', {k: p1b[k] for k in ['Sharpe Ratio', 'Max Drawdown', 'Trade Count', 'Annualized Return', 'Calmar Ratio']}, p1b['SL'], p1b['TP'], 0))
    else:
        approaches_data.append(('SL/TP', baseline_metrics, 0, 0, 0))

    if phase2a['best'] is not None:
        p2ab = phase2a['best']
        approaches_data.append(('TSL/TP', {k: p2ab[k] for k in ['Sharpe Ratio', 'Max Drawdown', 'Trade Count', 'Annualized Return', 'Calmar Ratio']}, 0, p2ab['TP'], p2ab['TSL']))
    else:
        approaches_data.append(('TSL/TP', baseline_metrics, 0, 0, 0))

    if phase2b['best'] is not None:
        p2bb = phase2b['best']
        approaches_data.append(('Bracket', {k: p2bb[k] for k in ['Sharpe Ratio', 'Max Drawdown', 'Trade Count', 'Annualized Return', 'Calmar Ratio']}, p2bb['SL'], p2bb['TP'], p2bb['TSL']))
    else:
        approaches_data.append(('Bracket', baseline_metrics, 0, 0, 0))

    for approach_name, metrics, sl_v, tp_v, tsl_v in approaches_data:
        comparison.append({
            'Strategy_ID': strategy_id,
            'Approach': approach_name,
            'SL': sl_v, 'TP': tp_v, 'TSL': tsl_v,
            'Sharpe_Ratio': metrics['Sharpe Ratio'],
            'Max_Drawdown': metrics['Max Drawdown'],
            'Annual_Return': metrics['Annualized Return'],
            'Trade_Count': metrics['Trade Count'],
            'Calmar_Ratio': metrics.get('Calmar Ratio', 0),
        })

    return result, comparison, risk_pnl_df


# ============================================================================
# MAIN
# ============================================================================

def main():
    args = parse_arguments()

    # Detect mode: CLI args present -> CLI mode; else -> interactive
    has_cli_args = (
        args.exchange != 'ibkr' or
        args.interval != '1h' or
        args.input is not None or
        args.output is not None or
        args.symbols is not None or
        args.n_jobs != -1
    )

    if has_cli_args:
        # CLI mode — use args directly
        exchange = args.exchange
        interval = args.interval
        n_jobs = args.n_jobs
        input_path = args.input
        output_path = args.output
        symbols_filter = [s.strip() for s in args.symbols.split(',')] if args.symbols else None
    else:
        # Interactive mode
        config = interactive_mode()
        if config is None:
            print("  Cancelled.")
            sys.exit(0)
        exchange = config['exchange']
        interval = config['interval']
        n_jobs = -1
        input_path = config['input_path']
        output_path = None
        symbols_filter = config['symbols']

    print("\n" + "=" * 70)
    print("  STAGE 8: Risk Parameter Optimization")
    print("=" * 70)
    print(f"  Exchange: {exchange}  |  Interval: {interval}  |  Jobs: {n_jobs}")

    # Get annualization period
    try:
        period = get_period_for_interval(interval)
    except ValueError as e:
        print(f"  CRITICAL: {e}")
        sys.exit(1)

    print(f"  Annualization period: {period}")

    # Find input
    if input_path is None:
        input_path = find_latest_final_compilation(exchange, interval)

    if input_path is None or not os.path.exists(input_path):
        print(f"\n  CRITICAL: Final_Compilation not found. Run Stage 7 first.")
        print(f"  Searched: WFAlphaResults/Final_Compilation_{exchange}_{interval}_*.xlsx")
        sys.exit(1)

    print(f"  Input: {input_path}")

    # Load Final_Alphas
    try:
        df_alphas = load_final_alphas(input_path)
    except KeyError as e:
        print(f"\n  CRITICAL: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n  CRITICAL: Failed to read Final_Alphas sheet: {e}")
        sys.exit(1)

    # Load Corr Summary for Portfolio sheet
    try:
        df_corr_summary = pd.read_excel(input_path, sheet_name='Corr Summary')
        print(f"  Corr Summary loaded: {len(df_corr_summary)} symbols")
    except Exception as e:
        print(f"  WARNING: Could not load Corr Summary: {e}")
        df_corr_summary = None

    # Apply symbol filter
    if symbols_filter:
        df_alphas = df_alphas[df_alphas['Symbol'].isin(symbols_filter)]
        if len(df_alphas) == 0:
            print(f"\n  CRITICAL: No strategies found for symbols: {symbols_filter}")
            sys.exit(1)
        print(f"  Symbol filter: {', '.join(symbols_filter)}")

    total_strategies = len(df_alphas)
    print(f"  Strategies to process: {total_strategies}")
    print(f"  Grid: Phase1={len(SL_RANGE)*len(TP_RANGE)} | Phase2A={len(TSL_RANGE)*len(TP_RANGE)} | Phase2B=21-42")
    print("-" * 70)

    # Process all strategies
    all_results = []
    all_comparisons = []
    all_pnl_data = []
    errors_list = []
    start_time = time.time()

    for idx, row in df_alphas.iterrows():
        try:
            result, comparisons, risk_pnl = process_strategy(row, exchange, interval, period, n_jobs, errors_list)
            if result is not None:
                all_results.append(result)
                all_comparisons.extend(comparisons)
                all_pnl_data.append((result['Strategy_ID'], result['Symbol'], risk_pnl))
        except Exception as e:
            strategy_id = build_strategy_id(row)
            log_error(errors_list, strategy_id, type(e).__name__, str(e))
            print(f"    ERROR: {type(e).__name__}: {e}")

    elapsed = time.time() - start_time

    # --- SUMMARY ---
    print("\n" + "=" * 70)
    print("  STAGE 8 SUMMARY")
    print("=" * 70)
    success_count = len(all_results)
    skip_count = total_strategies - success_count
    print(f"  Completed: {success_count} / {total_strategies}  ({skip_count} skipped)")
    print(f"  Runtime: {elapsed:.1f}s")

    if all_results:
        avg_sharpe_imp = np.mean([r['Sharpe_Improvement_%'] for r in all_results])
        print(f"  Avg Sharpe Improvement: {avg_sharpe_imp:+.1f}%")

        approach_dist = pd.Series([r['Selected_Approach'] for r in all_results]).value_counts()
        print(f"  Approach Distribution:")
        for approach, count in approach_dist.items():
            print(f"    {approach}: {count}")

    # Build Portfolio sheet
    portfolio_df = None
    if df_corr_summary is not None and all_pnl_data:
        risk_opt_metrics = build_risk_optimized_portfolios(all_pnl_data, interval)
        portfolio_df = build_portfolio_sheet(df_corr_summary, risk_opt_metrics)
        print(f"  Portfolio sheet: {len(portfolio_df)} symbols, {len(portfolio_df.columns)} columns")

    # Save consolidated Excel
    save_consolidated_excel(all_results, all_comparisons, exchange, interval, output_path, portfolio_df=portfolio_df)

    # Save errors
    save_errors(errors_list)

    print("=" * 70)


if __name__ == '__main__':
    main()
