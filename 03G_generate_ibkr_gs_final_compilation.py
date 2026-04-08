"""
Final Portfolio Compilation with OUTER Merge
Regenerates symbol portfolios using 0.5 correlation threshold strategies with OUTER merge

This script:
1. Reads Corr Summary (0.5 threshold) from Combination_Strategy_Compilation
2. Extracts strategy lists for each symbol
3. Loads backtest files and merges using OUTER join (preserves all datetime rows)
4. Calculates portfolio metrics on full datetime range
5. Generates Final_Compilation_ibkr_1h_{YYYYMMDD}.xlsx

Key difference from compile_combination_strategies.py:
- Uses how='outer' merge instead of how='inner' to preserve maximum data
- Fills NaN with 0 for missing strategy data

Configuration:
- Exchange: ibkr
- Symbols: VIXY, VIXM (base symbols)
- Interval: 1h
- Annualization: 252 trading days
"""

import pandas as pd
import numpy as np
import os
import sys
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import glob
from datetime import datetime
import argparse
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

# Import utility functions
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from util_AQS_parallel import get_period_for_interval

# ============================================================================
# DIRECTORY PATHS
# ============================================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = SCRIPT_DIR  # Script is in project root
DATA_DIR = os.path.join(PROJECT_DIR, "GridSearch_Data")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "WFAlphaResults")
AQS_RESULTS_DIR = os.path.join(PROJECT_DIR, "AQS_SFGridResults")


# ============================================================================
# FILE DISCOVERY UTILITIES
# ============================================================================

def find_latest_excel(pattern: str) -> Optional[str]:
    """
    Find latest Excel file matching glob pattern (for multi-day pipeline support).

    Args:
        pattern: Glob pattern to match files

    Returns:
        Path to most recently modified matching file, or None if not found
    """
    matches = glob.glob(pattern)
    if not matches:
        return None
    return max(matches, key=os.path.getmtime)

# ============================================================================
# CONFIGURATION (defaults, can be overridden via command-line)
# ============================================================================

EXCHANGE = "ibkr"
INTERVAL = "1h"
CORRELATION_THRESHOLD = 0.5  # Fixed to 0.5 threshold

# Symbols to process
SYMBOLS = ["VIXY", "VIXM"]  # IBKR symbols


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description='Generate final portfolio compilation with outer merge')
    parser.add_argument('--exchange', type=str, default=EXCHANGE,
                        help=f'Exchange name (default: {EXCHANGE})')
    parser.add_argument('--interval', type=str, default=INTERVAL,
                        help=f'Time interval (default: {INTERVAL})')
    parser.add_argument('--symbols', type=str, default=','.join(SYMBOLS),
                        help=f'Comma-separated list of symbols (default: {",".join(SYMBOLS)})')
    return parser.parse_args()


# ============================================================================
# UTILITY FUNCTIONS (Reused from compile_combination_strategies.py)
# ============================================================================

def load_strategy_backtest(backtest_path: str, strategy_name: str) -> Optional[pd.DataFrame]:
    """
    Load strategy backtest file and return DataFrame with renamed columns.

    Args:
        backtest_path: Path to backtest.csv file
        strategy_name: Name for the strategy (used in column naming)

    Returns:
        DataFrame with columns: datetime, {strategy_name}, {strategy_name}.1
        or None if file cannot be loaded
    """
    try:
        df = pd.read_csv(backtest_path)

        # Check required columns
        if 'datetime' not in df.columns or 'pnl' not in df.columns or 'signal' not in df.columns:
            return None

        # Extract and rename columns
        # Remove timezone to avoid Excel compatibility issues
        df_result = pd.DataFrame({
            # 'datetime': pd.to_datetime(df['datetime'], utc=True).dt.tz_convert(None),
            'datetime': pd.to_datetime(df['datetime'], utc=True).dt.tz_convert(None),  #Revision after debug
            strategy_name: df['pnl'],
            f'{strategy_name}.1': df['signal']
        })

        return df_result

    except Exception as e:
        print(f"  WARNING: Error loading {backtest_path}: {e}")
        return None


def load_close_prices(symbol: str, interval: str) -> Optional[pd.DataFrame]:
    """
    Load close prices from original merged CSV for Buy & Hold calculation.

    Args:
        symbol: Base symbol (e.g., "MBT")
        interval: Time interval (e.g., "1h", "1d")

    Returns:
        DataFrame with datetime and close columns, or None if error
    """
    # Construct path pattern to find merged CSV in GridSearch_Data
    pattern = f"merged_{EXCHANGE}_{symbol}_{interval}_*.csv"
    pattern_path = os.path.join(DATA_DIR, pattern)
    matches = glob.glob(pattern_path)

    if not matches:
        print(f"  WARNING: Close price CSV not found: {pattern}")
        return None

    csv_path = matches[0]

    try:
        df = pd.read_csv(csv_path, usecols=['datetime', 'close'])
        # df['datetime'] = pd.to_datetime(df['datetime'], utc=True).dt.tz_convert(None)
        df['datetime'] = pd.to_datetime(df['datetime'], utc=True).dt.tz_convert(None)  #Revision after debug
        return df
    except Exception as e:
        print(f"  WARNING: Failed to load close prices: {e}")
        return None


def calculate_portfolio_metrics(merged_df: pd.DataFrame, strategy_names: List[str], symbol: str, interval: str) -> Tuple[pd.DataFrame, Dict]:
    """
    Calculate portfolio metrics from merged DataFrame with Buy & Hold comparison.

    Args:
        merged_df: DataFrame with datetime and strategy P&L columns
        strategy_names: List of strategy names (P&L columns)
        symbol: Base symbol for loading close prices (e.g., "MBT")
        interval: Time interval (e.g., "1h", "1d")

    Returns:
        Tuple of (modified DataFrame with bnh/drawdown columns, metrics dictionary)
    """
    # Get annualization period dynamically based on interval
    period = get_period_for_interval(interval)

    # Calculate equal-weighted portfolio P&L
    pnl_cols = [col for col in merged_df.columns if col in strategy_names]
    merged_df['portfolio_pnl'] = merged_df[pnl_cols].mean(axis=1)
    merged_df['cumulative_pnl'] = merged_df['portfolio_pnl'].cumsum()

    # Load close prices for Buy & Hold calculation
    df_close = load_close_prices(symbol, interval)

    if df_close is not None:
        # Merge close prices on datetime
        merged_df = pd.merge(merged_df, df_close, on='datetime', how='left')

        # Calculate Buy & Hold cumulative returns
        merged_df['bnh'] = ((merged_df['close'] / merged_df['close'].shift(1)) - 1).cumsum()

        # Fill first NaN with 0
        merged_df['bnh'] = merged_df['bnh'].fillna(0)

        # Drop close column (not needed in output)
        merged_df = merged_df.drop(columns=['close'])
    else:
        # If close prices not available, set bnh to NaN
        merged_df['bnh'] = np.nan
        print(f"  WARNING: Buy & Hold not calculated for {symbol} (close prices unavailable)")

    merged_df['drawdown'] = merged_df['cumulative_pnl'] - merged_df['cumulative_pnl'].cummax()

    # Calculate metrics
    portfolio_pnl = merged_df['portfolio_pnl']
    sharpe = (portfolio_pnl.mean() / portfolio_pnl.std() * np.sqrt(period)) if portfolio_pnl.std() != 0 else 0
    mdd = merged_df['drawdown'].min()
    annual_return = portfolio_pnl.mean() * period
    calmar = annual_return / abs(mdd) if mdd != 0 else 0

    # Calculate final cumulative PnL and Buy & Hold values
    final_cumulative_pnl = merged_df['cumulative_pnl'].iloc[-1] if len(merged_df) > 0 else 0
    final_bnh = merged_df['bnh'].iloc[-1] if len(merged_df) > 0 and not merged_df['bnh'].isna().all() else 0
    pnl_ratio = final_cumulative_pnl / final_bnh if final_bnh != 0 else np.nan

    metrics = {
        'Sharpe': sharpe,
        'MDD': mdd,
        'Annual_Return': annual_return,
        'Calmar': calmar,
        'Cumulative_PnL': final_cumulative_pnl,
        'BnH': final_bnh,
        'PnL_Ratio': pnl_ratio,
        'Total_Rows': len(merged_df),
        'First_Datetime': merged_df['datetime'].min(),
        'Last_Datetime': merged_df['datetime'].max()
    }

    return merged_df, metrics


def calculate_correlation_matrix(merged_df: pd.DataFrame, strategy_names: List[str]) -> pd.DataFrame:
    """
    Calculate correlation matrix for strategy P&Ls.

    Args:
        merged_df: DataFrame with strategy P&L columns
        strategy_names: List of strategy names

    Returns:
        Correlation matrix DataFrame
    """
    # Filter to P&L columns only (exclude .1 signal columns)
    pnl_cols = [col for col in merged_df.columns if col in strategy_names]
    corr_matrix = merged_df[pnl_cols].corr()

    return corr_matrix


# ============================================================================
# STRATEGY MAPPING FUNCTIONS
# ============================================================================

def load_corr_summary_strategies(corr_summary_path: str, threshold: float = 0.5) -> Dict[str, List[str]]:
    """
    Load Corr Summary and extract strategy lists for specified threshold.

    Args:
        corr_summary_path: Path to Combination_Strategy_Compilation Excel
        threshold: Correlation threshold to filter (default: 0.5)

    Returns:
        Dictionary mapping symbol → list of strategy names
    """
    print(f"\n{'='*80}")
    print(f"Loading Corr Summary (Threshold={threshold})")
    print(f"{'='*80}")

    df_corr = pd.read_excel(corr_summary_path, sheet_name='Corr Summary')

    # Filter for specified threshold
    df_filtered = df_corr[df_corr['Threshold'] == threshold].copy()

    print(f"Found {len(df_filtered)} symbols with threshold {threshold}")

    # Parse strategy lists
    symbol_strategies = {}
    for idx, row in df_filtered.iterrows():
        symbol = row['Symbol']
        strategy_list_str = row['Strategy List']

        # Parse comma-separated strategy names
        strategies = [s.strip() for s in strategy_list_str.split(',')]
        symbol_strategies[symbol] = strategies

        print(f"  {symbol}: {len(strategies)} strategies")

    return symbol_strategies


# ============================================================================
# STRATEGY KEY AND EXTRACTION FUNCTIONS
# ============================================================================

def make_strategy_key(row: pd.Series, col_mapping: dict, include_variant: bool = True) -> str:
    """
    Generate strategy key from row using column name mapping.

    Args:
        row: DataFrame row
        col_mapping: Dict mapping logical names to column names
            Required keys: 'symbol', 'data_point', 'model', 'entry_exit'
            Optional keys: 'variant'
        include_variant: Whether to include variant in key (default True)

    Returns:
        Strategy key string: "{Symbol}_{Data Point}_{Model}_{Entry/Exit Model}_{Variant}"
        or "{Symbol}_{Data Point}_{Model}_{Entry/Exit Model}" if no variant
    """
    symbol = row[col_mapping['symbol']]
    data_point = row[col_mapping['data_point']]
    model = row[col_mapping['model']]
    entry_exit = row[col_mapping['entry_exit']]

    # Build base key (without variant)
    base_key = f"{symbol}_{data_point}_{model}_{entry_exit}"

    # Append variant at END if present and requested
    if include_variant and 'variant' in col_mapping and col_mapping['variant'] in row.index:
        variant = row[col_mapping['variant']]
        if pd.notna(variant) and str(variant).strip() and str(variant).lower() != 'default':
            return f"{base_key}_{variant}"

    return base_key


def extract_final_alphas(df_wfalpha: pd.DataFrame, all_strategies: List[str]) -> pd.DataFrame:
    """
    Extract Final_Alphas from WFAlpha_Compilation/WF_Short.
    Filters for strategies that appear in the Strategy List from Corr Summary.

    Args:
        df_wfalpha: WF_Short DataFrame from WFAlpha_Compilation
        all_strategies: List of all strategy names from Corr Summary Strategy Lists

    Returns:
        Filtered DataFrame containing only strategies in the list
    """
    # Column mapping for WFAlpha WF_Short (same as Final_Alphas format)
    col_mapping = {
        'symbol': 'Symbol',
        'data_point': 'Data Point',
        'model': 'Model',
        'entry_exit': 'Entry / Exit Model'
    }

    # Add variant to mapping if column exists
    if 'Variant' in df_wfalpha.columns:
        col_mapping['variant'] = 'Variant'

    # Generate strategy key for each row
    df_wfalpha['_strategy_key'] = df_wfalpha.apply(
        lambda row: make_strategy_key(row, col_mapping), axis=1
    )

    # Filter for strategies in the list
    df_filtered = df_wfalpha[df_wfalpha['_strategy_key'].isin(all_strategies)].copy()

    # Remove temporary column
    df_filtered = df_filtered.drop(columns=['_strategy_key'])

    print(f"  Final_Alphas: {len(df_filtered)} rows (from {len(df_wfalpha)} total)")

    return df_filtered


def extract_alpha_short(df_alpha_gs: pd.DataFrame, final_alphas_df: pd.DataFrame) -> pd.DataFrame:
    """
    Extract Alpha_Short from Alpha_GS_Compilation/Alpha_Short.
    Filters for strategies that appear in Final_Alphas.

    Note: Alpha_Short uses different column names than Final_Alphas:
        - Model -> model
        - Entry / Exit Model -> buy_type
        - Length -> length
        - Entry -> entry_threshold
        - Exit -> exit_threshold

    Args:
        df_alpha_gs: Alpha_Short DataFrame from Alpha_GS_Compilation
        final_alphas_df: Final_Alphas DataFrame (for filtering)

    Returns:
        Filtered DataFrame containing only strategies in Final_Alphas
    """
    # Build set of strategy keys from Final_Alphas
    final_col_mapping = {
        'symbol': 'Symbol',
        'data_point': 'Data Point',
        'model': 'Model',
        'entry_exit': 'Entry / Exit Model'
    }
    # Add variant to mapping if column exists
    if 'Variant' in final_alphas_df.columns:
        final_col_mapping['variant'] = 'Variant'

    final_strategies = set(
        final_alphas_df.apply(lambda row: make_strategy_key(row, final_col_mapping), axis=1)
    )

    # Column mapping for Alpha_Short (different names!)
    alpha_col_mapping = {
        'symbol': 'Symbol',
        'data_point': 'Data Point',
        'model': 'model',
        'entry_exit': 'buy_type'
    }
    # Add variant to mapping if column exists
    if 'Variant' in df_alpha_gs.columns:
        alpha_col_mapping['variant'] = 'Variant'

    # Generate strategy key for each row
    df_alpha_gs['_strategy_key'] = df_alpha_gs.apply(
        lambda row: make_strategy_key(row, alpha_col_mapping), axis=1
    )

    # Filter for strategies in Final_Alphas
    df_filtered = df_alpha_gs[df_alpha_gs['_strategy_key'].isin(final_strategies)].copy()

    # Remove temporary column
    df_filtered = df_filtered.drop(columns=['_strategy_key'])

    # ADDITIONAL FILTER: Match on specific parameter configurations
    # Build full keys from Final_Alphas (strategy + parameters)
    final_full_keys = set()
    for idx, row in final_alphas_df.iterrows():
        symbol = row['Symbol']
        data_point = row['Data Point']
        model = row['Model']
        entry_exit = row['Entry / Exit Model']
        variant = row.get('Variant', None)
        length = int(row['Length'])
        entry = round(float(row['Entry']), 10)
        exit_t = round(float(row['Exit']), 10)

        # Build base key
        if pd.notna(variant) and str(variant).strip() and str(variant).lower() != 'default':
            base_key = f"{symbol}_{data_point}_{model}_{entry_exit}_{variant}"
        else:
            base_key = f"{symbol}_{data_point}_{model}_{entry_exit}"

        # Add parameters
        full_key = f"{base_key}|{length}|{entry}|{exit_t}"
        final_full_keys.add(full_key)

    # Build full keys for Alpha_Short rows
    def make_alpha_full_key(row):
        symbol = row['Symbol']
        data_point = row['Data Point']
        model = row['model']
        entry_exit = row['buy_type']
        variant = row.get('Variant', None)
        length = int(row['length'])
        entry = round(float(row['entry_threshold']), 10)
        exit_t = round(float(row['exit_threshold']), 10)

        # Build base key
        if pd.notna(variant) and str(variant).strip() and str(variant).lower() != 'default':
            base_key = f"{symbol}_{data_point}_{model}_{entry_exit}_{variant}"
        else:
            base_key = f"{symbol}_{data_point}_{model}_{entry_exit}"

        # Add parameters
        return f"{base_key}|{length}|{entry}|{exit_t}"

    df_filtered['_full_key'] = df_filtered.apply(make_alpha_full_key, axis=1)

    # Filter to only rows with matching full keys (strategy + parameters)
    df_filtered = df_filtered[df_filtered['_full_key'].isin(final_full_keys)].copy()
    df_filtered = df_filtered.drop(columns=['_full_key'])

    print(f"  Alpha_Short: {len(df_filtered)} rows (from {len(df_alpha_gs)} total, matched on strategy + parameters)")

    return df_filtered


def extract_wf_short(df_wf_gs: pd.DataFrame, final_alphas_df: pd.DataFrame) -> pd.DataFrame:
    """
    Extract WF_Short from WF_GS_Compilation/WF_Short.
    Filters for strategies that appear in Final_Alphas.

    Args:
        df_wf_gs: WF_Short DataFrame from WF_GS_Compilation
        final_alphas_df: Final_Alphas DataFrame (for filtering)

    Returns:
        Filtered DataFrame containing only strategies in Final_Alphas
    """
    # Build set of strategy keys from Final_Alphas
    final_col_mapping = {
        'symbol': 'Symbol',
        'data_point': 'Data Point',
        'model': 'Model',
        'entry_exit': 'Entry / Exit Model'
    }
    # Add variant to mapping if column exists
    if 'Variant' in final_alphas_df.columns:
        final_col_mapping['variant'] = 'Variant'

    final_strategies = set(
        final_alphas_df.apply(lambda row: make_strategy_key(row, final_col_mapping), axis=1)
    )

    # Column mapping for WF_GS WF_Short (same naming as Final_Alphas)
    wf_col_mapping = {
        'symbol': 'Symbol',
        'data_point': 'Data Point',
        'model': 'Model',
        'entry_exit': 'Entry / Exit Model'
    }
    # Add variant to mapping if column exists
    if 'Variant' in df_wf_gs.columns:
        wf_col_mapping['variant'] = 'Variant'

    # Generate strategy key for each row
    df_wf_gs['_strategy_key'] = df_wf_gs.apply(
        lambda row: make_strategy_key(row, wf_col_mapping), axis=1
    )

    # Filter for strategies in Final_Alphas
    df_filtered = df_wf_gs[df_wf_gs['_strategy_key'].isin(final_strategies)].copy()

    # Remove temporary column
    df_filtered = df_filtered.drop(columns=['_strategy_key'])

    print(f"  WF_Short: {len(df_filtered)} rows (from {len(df_wf_gs)} total)")

    return df_filtered


def map_strategy_to_path(strategy_name: str, df_wf_short: pd.DataFrame,
                         exchange: str, interval: str, base_dir: str) -> Optional[str]:
    """
    Map strategy name to backtest.csv path using WF_Short metadata.

    Handles variant-aware strategy names in format:
    - With variant: {symbol}_{variant}_{feature}_{model}_{strategy}
    - Without variant: {symbol}_{feature}_{model}_{strategy}

    Args:
        strategy_name: Strategy name (e.g., "AAPL_john_23Jan2026_close_QQQ_spread_zscore_trend_reverse")
        df_wf_short: WF_Short DataFrame for lookups
        exchange: Exchange name (e.g., "ibkr")
        interval: Interval (e.g., "1h")
        base_dir: Base directory for backtest files

    Returns:
        Path to backtest.csv or None if not found
    """
    # Extract symbol from strategy name (first part before underscore)
    parts = strategy_name.split('_')
    symbol = parts[0]  # e.g., "AAPL"

    # Filter WF_Short for this symbol
    df_symbol = df_wf_short[df_wf_short['Symbol'] == symbol].copy()

    if df_symbol.empty:
        print(f"  WARNING: No WF_Short entries found for {symbol}")
        return None

    # Column mapping for generating strategy keys
    col_mapping = {
        'symbol': 'Symbol',
        'data_point': 'Data Point',
        'model': 'Model',
        'entry_exit': 'Entry / Exit Model'
    }

    # Add variant to mapping if column exists
    has_variant = 'Variant' in df_wf_short.columns
    if has_variant:
        col_mapping['variant'] = 'Variant'

    # Iterate through WF_Short rows to find matching strategy
    for idx, row in df_symbol.iterrows():
        # Reconstruct strategy name from WF_Short columns (with variant if present)
        test_name = make_strategy_key(row, col_mapping)

        if test_name == strategy_name:
            # Found match! Construct path using variant-aware folder name
            feature = row['Data Point']
            model = row['Model']
            entry_exit = row['Entry / Exit Model']

            # Get variant for folder construction
            variant = None
            if has_variant and pd.notna(row.get('Variant', None)):
                variant = str(row['Variant']).strip()
                if variant.lower() == 'default' or not variant:
                    variant = None

            # Construct folder name with variant
            if variant:
                folder_name = f"merged_{exchange}_{symbol}_{interval}_{variant}"
            else:
                # Fallback to glob pattern for non-variant folders
                folder_pattern = f"merged_{exchange}_{symbol}_{interval}_*"
                folder_pattern_path = os.path.join(base_dir, folder_pattern)
                matches = glob.glob(folder_pattern_path)
                if not matches:
                    print(f"  WARNING: No folder found matching pattern: {folder_pattern}")
                    return None
                folder_name = os.path.basename(matches[0])

            path = os.path.join(base_dir, folder_name, feature, model, entry_exit, "backtest.csv")

            return path

    print(f"  WARNING: Strategy {strategy_name} not found in WF_Short")
    return None


# ============================================================================
# PORTFOLIO BUILDING WITH OUTER MERGE
# ============================================================================

def build_portfolio_outer_merge(symbol: str, strategy_names: List[str],
                                df_wf_short: pd.DataFrame, interval: str) -> Optional[pd.DataFrame]:
    """
    Build portfolio using OUTER merge to preserve all datetime rows.

    Args:
        symbol: Base symbol name (e.g., "MBT")
        strategy_names: List of strategy names to include
        df_wf_short: WF_Short DataFrame for path lookups
        interval: Time interval (e.g., "1h", "1d")

    Returns:
        Merged DataFrame with all strategies (NaN filled with 0) or None if no strategies loaded
    """
    print(f"\n{'='*80}")
    print(f"Building Portfolio for {symbol}: {len(strategy_names)} strategies")
    print(f"{'='*80}")

    merged_df = None
    loaded_strategies = []

    for i, strategy_name in enumerate(strategy_names, 1):
        print(f"  [{i}/{len(strategy_names)}] Loading {strategy_name}...")

        # Map strategy name to backtest path
        backtest_path = map_strategy_to_path(strategy_name, df_wf_short, EXCHANGE, interval, OUTPUT_DIR)

        if backtest_path is None:
            print(f"    SKIP: Could not map to path")
            continue

        if not os.path.exists(backtest_path):
            print(f"    SKIP: File not found: {backtest_path}")
            continue

        # Load strategy backtest
        df_strategy = load_strategy_backtest(backtest_path, strategy_name)

        if df_strategy is None:
            print(f"    SKIP: Failed to load")
            continue

        print(f"    OK: {len(df_strategy)} rows, {df_strategy['datetime'].min()} to {df_strategy['datetime'].max()}")

        # OUTER merge to preserve all datetime rows
        if merged_df is None:
            merged_df = df_strategy
        else:
            # ALWAYS use OUTER merge (key difference from original script!)
            merged_df = pd.merge(merged_df, df_strategy, on='datetime', how='outer', sort=True)
            print(f"    MERGE: OUTER join -> {len(merged_df)} rows")

        loaded_strategies.append(strategy_name)

    if merged_df is None or len(loaded_strategies) == 0:
        print(f"\n  ERROR: No strategies successfully loaded for {symbol}")
        return None

    # Fill NaN values with 0 (no position for missing data)
    print(f"\nFilling NaN values...")
    pnl_cols = [col for col in merged_df.columns if not col.endswith('.1') and col != 'datetime']
    signal_cols = [col for col in merged_df.columns if col.endswith('.1')]

    nan_before = merged_df[pnl_cols].isna().sum().sum()
    merged_df[pnl_cols] = merged_df[pnl_cols].fillna(0)
    merged_df[signal_cols] = merged_df[signal_cols].fillna(0)
    print(f"  Filled {nan_before} NaN values with 0")

    # Calculate portfolio metrics
    print(f"\nCalculating portfolio metrics...")
    merged_df, metrics = calculate_portfolio_metrics(merged_df, loaded_strategies, symbol, interval)

    print(f"  Portfolio Sharpe: {metrics['Sharpe']:.4f}")
    print(f"  Portfolio MDD: {metrics['MDD']:.4f}")
    print(f"  Annual Return: {metrics['Annual_Return']:.4f}")
    print(f"  Calmar Ratio: {metrics['Calmar']:.4f}")
    print(f"  Cumulative PnL: {metrics['Cumulative_PnL']:.4f}")
    print(f"  Buy & Hold: {metrics['BnH']:.4f}")
    print(f"  PnL Ratio: {metrics['PnL_Ratio']:.4f}")
    print(f"  Total Rows: {metrics['Total_Rows']:,}")
    print(f"  Date Range: {metrics['First_Datetime']} to {metrics['Last_Datetime']}")

    return merged_df


# ============================================================================
# EXCEL OUTPUT GENERATION
# ============================================================================

def generate_corr_summary_sheet(symbol_strategies: Dict[str, List[str]],
                                symbol_metrics: Dict[str, Dict]) -> pd.DataFrame:
    """
    Generate Corr Summary worksheet DataFrame.

    Args:
        symbol_strategies: Dict mapping symbol → strategy list
        symbol_metrics: Dict mapping symbol → metrics dict

    Returns:
        DataFrame for Corr Summary worksheet
    """
    rows = []

    for symbol, strategies in symbol_strategies.items():
        if symbol in symbol_metrics:
            metrics = symbol_metrics[symbol]
            rows.append({
                'Symbol': symbol,
                'Threshold': CORRELATION_THRESHOLD,
                'Total Strategies': len(strategies),
                '# Strategies Selected': len(strategies),
                'Portfolio Sharpe': metrics.get('Sharpe', 0),
                'Portfolio MDD': metrics.get('MDD', 0),
                'Portfolio Annual Return': metrics.get('Annual_Return', 0),
                'Portfolio Calmar Ratio': metrics.get('Calmar', 0),
                'Cumulative PnL': metrics.get('Cumulative_PnL', 0),
                'Buy & Hold': metrics.get('BnH', 0),
                'PnL Ratio': metrics.get('PnL_Ratio', 0),
                'Total Rows': metrics.get('Total_Rows', 0),
                'First Datetime': metrics.get('First_Datetime', ''),
                'Last Datetime': metrics.get('Last_Datetime', ''),
                'Strategy List': ', '.join(strategies)
            })

    return pd.DataFrame(rows)


# ============================================================================
# EXCEL FORMATTING HELPER FUNCTIONS
# ============================================================================

def rename_alpha_short_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename columns in Alpha_Short DataFrame for Excel output.

    Changes:
    - 'model' → 'Model'
    - 'buy_type' → 'Entry / Exit Model'

    Args:
        df: Alpha_Short DataFrame

    Returns:
        DataFrame with renamed columns
    """
    if df is None or df.empty:
        return df

    rename_mapping = {
        'model': 'Model',
        'buy_type': 'Entry / Exit Model'
    }

    # Only rename columns that exist
    existing_renames = {k: v for k, v in rename_mapping.items() if k in df.columns}

    return df.rename(columns=existing_renames)


def reorder_dataframe_columns(df: pd.DataFrame, sheet_type: str) -> pd.DataFrame:
    """
    Reorder DataFrame columns according to sheet-specific requirements.

    Target order for all sheets:
    #, Exchange, Symbol, Interval, Data Point, Model, Entry / Exit Model, Variant, [remaining columns]

    Args:
        df: DataFrame to reorder
        sheet_type: One of 'Alpha_Short', 'Final_Alphas', 'WF_Short'

    Returns:
        DataFrame with reordered columns
    """
    if df is None or df.empty:
        return df

    # Define priority column order
    priority_cols = [
        '#', 'Exchange', 'Symbol', 'Interval', 'Data Point',
        'Model', 'Entry / Exit Model', 'Variant'
    ]

    # Filter to only existing priority columns (in order)
    ordered_cols = [col for col in priority_cols if col in df.columns]

    # Add remaining columns (not in priority list)
    remaining_cols = [col for col in df.columns if col not in priority_cols]

    # Combine: priority columns first, then remaining
    final_order = ordered_cols + remaining_cols

    return df[final_order]


def format_worksheet_headers(worksheet, df: pd.DataFrame, sheet_name: str):
    """
    Apply header formatting to worksheet:
    - Font: Bold, 11pt, Calibri
    - Background: Theme colors with 0.8 tint (sheet-specific ranges)

    Args:
        worksheet: openpyxl worksheet object
        df: DataFrame that was written
        sheet_name: Name of the worksheet ('Alpha_Short', 'Final_Alphas', 'WF_Short')
    """
    # Define header font (bold, 11pt, Calibri)
    header_font = Font(name='Calibri', size=11, bold=True)

    # Define theme color fills (with 0.8 tint)
    fill_theme5 = PatternFill(patternType='solid', fgColor=Color(theme=5, tint=0.8))  # Purple
    fill_theme6 = PatternFill(patternType='solid', fgColor=Color(theme=6, tint=0.8))  # Cyan
    fill_theme7 = PatternFill(patternType='solid', fgColor=Color(theme=7, tint=0.8))  # Orange
    fill_theme9 = PatternFill(patternType='solid', fgColor=Color(theme=9, tint=0.8))  # Purple
    fill_index7 = PatternFill(patternType='solid', fgColor='FFC000')  # Fill Index 7 (orange)

    # Get column headers
    headers = list(df.columns)

    # Define color ranges by sheet type
    if sheet_name == 'Alpha_Short':
        # Cols A-K: Theme 5 (purple)
        # IS metrics: Theme 6 (cyan)
        # OOS metrics: Theme 7 (orange)
        # Degradation metrics: Theme 9 (purple)

        # Map column names to theme colors
        theme5_cols = ['#', 'Exchange', 'Symbol', 'Interval', 'Data Point',
                       'Model', 'Entry / Exit Model', 'Variant', 'length',
                       'entry_threshold', 'exit_threshold']
        theme6_cols = ['Sharpe Ratio_IS', 'Annualized Return_IS', 'Max Drawdown_IS',
                       'Calmar Ratio_IS', 'Trade Count_IS']
        theme7_cols = ['Sharpe Ratio_OOS', 'Annualized Return_OOS', 'Max Drawdown_OOS',
                       'Calmar Ratio_OOS', 'Trade Count_OOS']
        theme9_cols = ['Sharpe_Degradation_%', 'Return_Degradation_%',
                       'Drawdown_Degradation_%', 'Calmar_Degradation_%']

        color_mapping = {}
        for col in theme5_cols:
            if col in headers:
                color_mapping[col] = fill_theme5
        for col in theme6_cols:
            if col in headers:
                color_mapping[col] = fill_theme6
        for col in theme7_cols:
            if col in headers:
                color_mapping[col] = fill_theme7
        for col in theme9_cols:
            if col in headers:
                color_mapping[col] = fill_theme9

    elif sheet_name == 'Final_Alphas':
        # Cols A-K: Theme 5
        # Cols L-S: Theme 6
        # Cols T-W: No fill

        theme5_cols = ['#', 'Exchange', 'Symbol', 'Interval', 'Data Point',
                       'Model', 'Entry / Exit Model', 'Variant', 'Length',
                       'Entry', 'Exit']
        theme6_cols = ['Sharpe', 'MDD', 'Trade Count', 'Annual Return',
                       'Calmar Ratio', 'Cumulative PnL', 'Buy & Hold', 'PnL Ratio']
        no_fill_cols = ['Overfit_1hop', 'Overfit_2hop', 'Overfit_Final', 'Heatmap Checked']

        color_mapping = {}
        for col in theme5_cols:
            if col in headers:
                color_mapping[col] = fill_theme5
        for col in theme6_cols:
            if col in headers:
                color_mapping[col] = fill_theme6
        # no_fill_cols get no background (omitted from mapping)

    elif sheet_name == 'WF_Short':
        # Cols A-K: Theme 5
        # IS metrics: Theme 6
        # WF1 metrics: Theme 7
        # WF2 metrics: Theme 9
        # Degradation columns: Fill Index 7

        theme5_cols = ['#', 'Exchange', 'Symbol', 'Interval', 'Data Point',
                       'Model', 'Entry / Exit Model', 'Variant', 'Length',
                       'Entry', 'Exit']
        theme6_cols = ['IS Sharpe', 'IS MDD', 'IS Trade Count', 'IS Annual Return',
                       'IS Calmar Ratio']
        theme7_cols = ['WF1 Sharpe', 'WF1 MDD', 'WF1 Trade Count', 'WF1 Annual Return',
                       'WF1 Calmar Ratio']
        theme9_cols = ['WF2 Sharpe', 'WF2 MDD', 'WF2 Trade Count', 'WF2 Annual Return',
                       'WF2 Calmar Ratio']
        fill7_cols = ['WF1 Sharpe Degrade', 'WF1 MDD Degrade', 'WF1 Annual Return Degrade',
                      'WF1 Calmar Ratio Degrade', 'WF2 Sharpe Degrade', 'WF2 MDD Degrade',
                      'WF2 Annual Return Degrade', 'WF2 Calmar Ratio Degrade',
                      'L Sharpe Degrade', 'L MDD Degrade', 'L Annual Return Degrade',
                      'L Calmar Ratio Degrade']

        color_mapping = {}
        for col in theme5_cols:
            if col in headers:
                color_mapping[col] = fill_theme5
        for col in theme6_cols:
            if col in headers:
                color_mapping[col] = fill_theme6
        for col in theme7_cols:
            if col in headers:
                color_mapping[col] = fill_theme7
        for col in theme9_cols:
            if col in headers:
                color_mapping[col] = fill_theme9
        for col in fill7_cols:
            if col in headers:
                color_mapping[col] = fill_index7

    elif sheet_name == 'Corr Summary':
        # Corr Summary: Mixed theme colors
        # Columns A-D: Theme 3 (Light Gray)
        # Column E onwards: Theme 5 (Purple)

        fill_theme3 = PatternFill(patternType='solid', fgColor=Color(theme=3, tint=0.8))  # Light Gray

        color_mapping = {}
        for i in range(min(4, len(headers))):
            col_name = headers[i]
            color_mapping[col_name] = fill_theme3

        for i in range(4, len(headers)):
            col_name = headers[i]
            color_mapping[col_name] = fill_theme5

    elif ' Corr' in sheet_name or ' Portfolio' in sheet_name:
        # Portfolio and Correlation worksheets: All Theme 3 (Light Gray)
        fill_theme3 = PatternFill(patternType='solid', fgColor=Color(theme=3, tint=0.8))  # Light Gray

        color_mapping = {}
        for col_name in headers:
            color_mapping[col_name] = fill_theme3

    else:
        # Unknown sheet type - no coloring
        color_mapping = {}

    # Apply formatting to header row (row 1)
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = worksheet[f'{col_letter}1']

        # Apply font
        cell.font = header_font

        # Apply fill if defined
        if col_name in color_mapping:
            cell.fill = color_mapping[col_name]


def autofit_column_widths(worksheet, sheet_name: str = None):
    """
    Auto-fit column widths based on content length.
    Special handling for Correlation worksheets (narrow Column A).

    Args:
        worksheet: openpyxl worksheet object
        sheet_name: Optional name of the worksheet (for special handling)
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        # Special case: Correlation worksheets Column A (intentionally narrow)
        if sheet_name and ' Corr' in sheet_name and column_letter == 'A':
            worksheet.column_dimensions[column_letter].width = 3.0
            continue

        for cell in column:
            try:
                if cell.value:
                    # Calculate length of cell content
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set width with padding (min 10, max 80 for long strategy names)
        adjusted_width = min(max(max_length + 2, 10), 80)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def hide_columns(worksheet, sheet_name: str, df: pd.DataFrame):
    """
    Hide specific columns based on sheet type.

    Currently only hides columns in Final_Alphas worksheet:
    - Columns T, U, V, W (Overfit_1hop, Overfit_2hop, Overfit_Final, Heatmap Checked)

    Args:
        worksheet: openpyxl worksheet object
        sheet_name: Name of the worksheet
        df: DataFrame that was written (to determine column positions)
    """
    if sheet_name == 'Final_Alphas':
        # Columns to hide in Final_Alphas
        cols_to_hide = ['Overfit_1hop', 'Overfit_2hop', 'Overfit_Final', 'Heatmap Checked']

        headers = list(df.columns)

        for col_name in cols_to_hide:
            if col_name in headers:
                # Find column index (1-indexed)
                col_idx = headers.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                # Hide the column
                worksheet.column_dimensions[col_letter].hidden = True


def save_to_excel(symbol_portfolios: Dict[str, pd.DataFrame],
                 symbol_strategies: Dict[str, List[str]],
                 symbol_metrics: Dict[str, Dict],
                 output_path: str,
                 final_alphas_df: Optional[pd.DataFrame] = None,
                 alpha_short_df: Optional[pd.DataFrame] = None,
                 wf_short_df: Optional[pd.DataFrame] = None):
    """
    Save all portfolios and correlation matrices to Excel file.

    Revised output structure (in order):
    1. Final_Alphas (green tab)
    2. Alpha_Short (yellow tab)
    3. WF_Short (yellow tab)
    4. Corr Summary (yellow tab)
    5. {Symbol} Portfolio (orange tab)
    6. {Symbol} Corr (orange tab)

    Args:
        symbol_portfolios: Dict mapping symbol → merged DataFrame
        symbol_strategies: Dict mapping symbol → strategy list
        symbol_metrics: Dict mapping symbol → metrics dict
        output_path: Output Excel file path
        final_alphas_df: Final_Alphas DataFrame (filtered from WFAlpha_Compilation)
        alpha_short_df: Alpha_Short DataFrame (filtered from Alpha_GS_Compilation)
        wf_short_df: WF_Short DataFrame (filtered from WF_GS_Compilation)
    """
    from openpyxl import load_workbook

    print(f"\n{'='*80}")
    print(f"Saving to Excel: {output_path}")
    print(f"{'='*80}")

    # PRE-PROCESS: Rename and reorder DataFrames BEFORE writing
    print(f"\nPre-processing DataFrames for formatting...")

    # 1. Rename Alpha_Short columns
    if alpha_short_df is not None and len(alpha_short_df) > 0:
        alpha_short_df = rename_alpha_short_columns(alpha_short_df)
        print(f"  [OK] Alpha_Short: Renamed columns (model -> Model, buy_type -> Entry / Exit Model)")

    # 2. Reorder columns for all three sheets
    if final_alphas_df is not None and len(final_alphas_df) > 0:
        final_alphas_df = reorder_dataframe_columns(final_alphas_df, 'Final_Alphas')
        print(f"  [OK] Final_Alphas: Reordered columns")

    if alpha_short_df is not None and len(alpha_short_df) > 0:
        alpha_short_df = reorder_dataframe_columns(alpha_short_df, 'Alpha_Short')
        print(f"  [OK] Alpha_Short: Reordered columns")

    if wf_short_df is not None and len(wf_short_df) > 0:
        wf_short_df = reorder_dataframe_columns(wf_short_df, 'WF_Short')
        print(f"  [OK] WF_Short: Reordered columns")

    print(f"\nWriting to Excel with formatting...")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. Write Final_Alphas worksheet (first sheet)
        if final_alphas_df is not None and len(final_alphas_df) > 0:
            final_alphas_df.to_excel(writer, sheet_name='Final_Alphas', index=False)
            worksheet = writer.sheets['Final_Alphas']
            format_worksheet_headers(worksheet, final_alphas_df, 'Final_Alphas')
            autofit_column_widths(worksheet, 'Final_Alphas')
            hide_columns(worksheet, 'Final_Alphas', final_alphas_df)
            print(f"  [OK] Final_Alphas: {len(final_alphas_df)} rows (formatted)")

        # 2. Write Alpha_Short worksheet
        if alpha_short_df is not None and len(alpha_short_df) > 0:
            alpha_short_df.to_excel(writer, sheet_name='Alpha_Short', index=False)
            worksheet = writer.sheets['Alpha_Short']
            format_worksheet_headers(worksheet, alpha_short_df, 'Alpha_Short')
            autofit_column_widths(worksheet, 'Alpha_Short')
            hide_columns(worksheet, 'Alpha_Short', alpha_short_df)
            print(f"  [OK] Alpha_Short: {len(alpha_short_df)} rows (formatted)")

        # 3. Write WF_Short worksheet
        if wf_short_df is not None and len(wf_short_df) > 0:
            wf_short_df.to_excel(writer, sheet_name='WF_Short', index=False)
            worksheet = writer.sheets['WF_Short']
            format_worksheet_headers(worksheet, wf_short_df, 'WF_Short')
            autofit_column_widths(worksheet, 'WF_Short')
            hide_columns(worksheet, 'WF_Short', wf_short_df)
            print(f"  [OK] WF_Short: {len(wf_short_df)} rows (formatted)")

        # 4. Write Corr Summary worksheet
        df_summary = generate_corr_summary_sheet(symbol_strategies, symbol_metrics)
        df_summary.to_excel(writer, sheet_name='Corr Summary', index=False)
        worksheet = writer.sheets['Corr Summary']
        format_worksheet_headers(worksheet, df_summary, 'Corr Summary')
        # NOTE: Corr Summary uses default widths (not auto-fitted)
        print(f"  [OK] Corr Summary: {len(df_summary)} rows (formatted)")

        # 5-6. Write symbol-specific worksheets (Portfolio and Corr)
        for symbol, merged_df in symbol_portfolios.items():
            strategies = symbol_strategies[symbol]

            # Write Portfolio worksheet
            portfolio_sheet = f"{symbol} Portfolio"
            merged_df.to_excel(writer, sheet_name=portfolio_sheet, index=False)
            worksheet = writer.sheets[portfolio_sheet]
            format_worksheet_headers(worksheet, merged_df, portfolio_sheet)
            autofit_column_widths(worksheet, portfolio_sheet)
            print(f"  [OK] {portfolio_sheet}: {len(merged_df)} rows, {len(merged_df.columns)} columns (formatted)")

            # Calculate and write Correlation matrix worksheet
            corr_matrix = calculate_correlation_matrix(merged_df, strategies)

            # Add summary statistics (with guard for single-strategy symbols)
            corr_values = corr_matrix.values[np.triu_indices_from(corr_matrix.values, k=1)]

            if len(corr_values) > 0:
                avg_corr = corr_values.mean()
                max_corr = corr_values.max()
                min_corr = corr_values.min()
                avg_str = f'{avg_corr:.4f}'
                max_str = f'{max_corr:.4f}'
                min_str = f'{min_corr:.4f}'
            else:
                # Single strategy - no correlation pairs
                avg_str = 'N/A (single strategy)'
                max_str = 'N/A'
                min_str = 'N/A'

            # Append statistics to correlation matrix
            stats_df = pd.DataFrame({
                corr_matrix.columns[0]: ['---', 'STATISTICS', f'Avg Correlation: {avg_str}',
                                         f'Max Correlation: {max_str}', f'Min Correlation: {min_str}',
                                         f'Strategies: {len(strategies)}']
            })
            corr_with_stats = pd.concat([corr_matrix, stats_df], ignore_index=True)

            corr_sheet = f"{symbol} Corr"
            corr_with_stats.to_excel(writer, sheet_name=corr_sheet)
            worksheet = writer.sheets[corr_sheet]
            format_worksheet_headers(worksheet, corr_with_stats, corr_sheet)
            autofit_column_widths(worksheet, corr_sheet)
            print(f"  [OK] {corr_sheet}: {len(corr_matrix)}x{len(corr_matrix)} matrix (formatted)")

    # Apply tab colors after ExcelWriter closes
    print(f"\nApplying tab colors...")
    wb = load_workbook(output_path)

    # Tab color definitions
    TAB_COLOR_GREEN = 'FF92D050'   # Final_Alphas
    TAB_COLOR_YELLOW = 'FFFFFF00'  # Alpha_Short, WF_Short, Corr Summary
    TAB_COLOR_ORANGE = 'FFFFC000'  # Portfolio, Corr sheets

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == 'Final_Alphas':
            ws.sheet_properties.tabColor = TAB_COLOR_GREEN
        elif sheet_name in ['Alpha_Short', 'WF_Short', 'Corr Summary']:
            ws.sheet_properties.tabColor = TAB_COLOR_YELLOW
        elif 'Portfolio' in sheet_name or 'Corr' in sheet_name:
            ws.sheet_properties.tabColor = TAB_COLOR_ORANGE

    wb.save(output_path)
    print(f"  [OK] Tab colors applied")

    # Count total worksheets
    total_sheets = 0
    if final_alphas_df is not None and len(final_alphas_df) > 0:
        total_sheets += 1
    if alpha_short_df is not None and len(alpha_short_df) > 0:
        total_sheets += 1
    if wf_short_df is not None and len(wf_short_df) > 0:
        total_sheets += 1
    total_sheets += 1  # Corr Summary
    total_sheets += len(symbol_portfolios) * 2  # Portfolio + Corr per symbol

    print(f"\n{'='*80}")
    print(f"Excel file saved successfully!")
    print(f"  File: {output_path}")
    print(f"  Worksheets: {total_sheets}")
    print(f"{'='*80}")


# ============================================================================
# INTERACTIVE MODE
# ============================================================================

def interactive_mode():
    """Interactive CLI for running Stage 7: Final Compilation"""
    print("""
╔══════════════════════════════════════════════════════════════╗
║      Stage 7: Final Compilation - Interactive Mode           ║
║  Final portfolio assembly with OUTER merge                   ║
╚══════════════════════════════════════════════════════════════╝
    """)

    try:
        symbols_input = input("Enter symbols (comma-separated, e.g., NVDA, AAPL, TECL): ").strip()
        if not symbols_input:
            print("Symbols cannot be empty")
            return
        symbols = [s.strip().upper() for s in symbols_input.split(',')]

        print("\nAvailable intervals:")
        print("  1min, 5min, 15min, 30min, 1h, 1d, 1w")
        interval = input("Enter interval (default: 1h): ").strip().lower()
        if not interval or interval not in ['1min', '5min', '15min', '30min', '1h', '1d', '1w']:
            interval = '1h'
            print("  ✗ Invalid interval. Using default: 1h")

        print(f"\n{'='*60}")
        print("Configuration Summary:")
        print(f"  Symbols:     {', '.join(symbols)}")
        print(f"  Interval:    {interval}")
        print(f"  Exchange:    {EXCHANGE}")
        print(f"{'='*60}")

        proceed = input("\nProceed with execution? (y/n): ").strip().lower()
        if proceed != 'y':
            print("\nExecution cancelled")
            return

        sys.argv = [sys.argv[0], '--symbols', ','.join(symbols), '--interval', interval, '--exchange', EXCHANGE]
        main()

    except KeyboardInterrupt:
        print("\n\nCancelled by user")


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function."""
    # Parse command-line arguments
    args = parse_arguments()
    exchange = args.exchange
    interval = args.interval
    symbols = [s.strip() for s in args.symbols.split(',')]

    # Build dynamic paths using glob patterns (supports multi-day pipelines)
    today = datetime.now().strftime('%Y%m%d')

    # Existing source files (use glob to find latest)
    corr_summary_pattern = os.path.join(OUTPUT_DIR, f"Combination_Strategy_Compilation_{exchange}_{interval}_*.xlsx")
    wfalpha_pattern = os.path.join(OUTPUT_DIR, f"WFAlpha_Compilation_{exchange}_{interval}_*.xlsx")

    input_corr_summary = find_latest_excel(corr_summary_pattern)
    input_wfalpha = find_latest_excel(wfalpha_pattern)

    # New source files for revised output (from AQS_SFGridResults)
    alpha_gs_pattern = os.path.join(AQS_RESULTS_DIR, f"Alpha_GS_Compilation_{exchange}_{interval}_*.xlsx")
    wf_gs_pattern = os.path.join(AQS_RESULTS_DIR, f"WF_GS_Compilation_{exchange}_{interval}_*.xlsx")

    input_alpha_gs = find_latest_excel(alpha_gs_pattern)
    input_wf_gs = find_latest_excel(wf_gs_pattern)

    # Output file (use today's date)
    output_file = os.path.join(OUTPUT_DIR, f"Final_Compilation_{exchange}_{interval}_{today}.xlsx")

    print(f"\n{'='*80}")
    print(f"FINAL PORTFOLIO COMPILATION WITH OUTER MERGE (REVISED OUTPUT)")
    print(f"{'='*80}")
    print(f"Exchange: {exchange}")
    print(f"Interval: {interval}")
    print(f"Symbols: {', '.join(symbols)}")
    print(f"Correlation Threshold: {CORRELATION_THRESHOLD}")
    print(f"\nInput Files:")
    print(f"  Combination_Strategy_Compilation: {os.path.basename(input_corr_summary) if input_corr_summary else 'NOT FOUND'}")
    print(f"  WFAlpha_Compilation: {os.path.basename(input_wfalpha) if input_wfalpha else 'NOT FOUND'}")
    print(f"  Alpha_GS_Compilation: {os.path.basename(input_alpha_gs) if input_alpha_gs else 'NOT FOUND'}")
    print(f"  WF_GS_Compilation: {os.path.basename(input_wf_gs) if input_wf_gs else 'NOT FOUND'}")
    print(f"\nOutput: {output_file}")
    print(f"{'='*80}")

    # Validate required input files exist
    if not input_corr_summary:
        print(f"\nERROR: Combination_Strategy_Compilation not found matching: {corr_summary_pattern}")
        return
    if not input_wfalpha:
        print(f"\nERROR: WFAlpha_Compilation not found matching: {wfalpha_pattern}")
        return

    # 1. Load Corr Summary strategies (0.5 threshold)
    symbol_strategies = load_corr_summary_strategies(input_corr_summary, CORRELATION_THRESHOLD)

    if not symbol_strategies:
        print("\nERROR: No strategies found in Corr Summary")
        return

    # Filter symbols if specified
    if symbols:
        symbol_strategies = {k: v for k, v in symbol_strategies.items()
                           if k in symbols}
        print(f"\nFiltered to {len(symbol_strategies)} symbols: {', '.join(symbol_strategies.keys())}")

    # Collect all strategy names for filtering
    all_strategies = []
    for strategies in symbol_strategies.values():
        all_strategies.extend(strategies)
    print(f"Total strategies to include: {len(all_strategies)}")

    # 2. Load WF_Short from WFAlpha_Compilation for strategy mapping (existing functionality)
    print(f"\nLoading WF_Short from {os.path.basename(input_wfalpha)}...")
    df_wf_short = pd.read_excel(input_wfalpha, sheet_name='WF_Short')
    print(f"  Loaded {len(df_wf_short)} rows")

    # 3. Extract new sheets for revised output
    print(f"\n{'='*80}")
    print(f"Extracting sheets for revised output...")
    print(f"{'='*80}")

    # 3a. Extract Final_Alphas (from WFAlpha_Compilation/WF_Short)
    final_alphas_df = extract_final_alphas(df_wf_short, all_strategies)

    # 3b. Extract Alpha_Short (from Alpha_GS_Compilation)
    alpha_short_df = None
    if input_alpha_gs:
        print(f"\nLoading Alpha_Short from {os.path.basename(input_alpha_gs)}...")
        df_alpha_gs_full = pd.read_excel(input_alpha_gs, sheet_name='Alpha_Short')
        print(f"  Loaded {len(df_alpha_gs_full)} rows from source")
        alpha_short_df = extract_alpha_short(df_alpha_gs_full, final_alphas_df)
    else:
        print(f"\n  WARNING: Alpha_GS_Compilation not found, skipping Alpha_Short extraction")

    # 3c. Extract WF_Short (from WF_GS_Compilation)
    wf_short_df = None
    if input_wf_gs:
        print(f"\nLoading WF_Short from {os.path.basename(input_wf_gs)}...")
        df_wf_gs_full = pd.read_excel(input_wf_gs, sheet_name='WF_Short')
        print(f"  Loaded {len(df_wf_gs_full)} rows from source")
        wf_short_df = extract_wf_short(df_wf_gs_full, final_alphas_df)
    else:
        print(f"\n  WARNING: WF_GS_Compilation not found, skipping WF_Short extraction")

    # 4. Build portfolios with OUTER merge (existing functionality)
    print(f"\n{'='*80}")
    print(f"Building portfolios...")
    print(f"{'='*80}")

    symbol_portfolios = {}
    symbol_metrics = {}

    for symbol, strategies in symbol_strategies.items():
        merged_df = build_portfolio_outer_merge(symbol, strategies, df_wf_short, interval)

        if merged_df is not None:
            # Calculate metrics (this also adds bnh and drawdown columns to merged_df)
            merged_df, metrics = calculate_portfolio_metrics(merged_df, strategies, symbol, interval)

            # Store updated DataFrame and metrics
            symbol_portfolios[symbol] = merged_df
            symbol_metrics[symbol] = metrics

    if not symbol_portfolios:
        print("\nERROR: No portfolios successfully built")
        return

    # 5. Save to Excel with revised output format
    save_to_excel(
        symbol_portfolios,
        symbol_strategies,
        symbol_metrics,
        output_file,
        final_alphas_df=final_alphas_df,
        alpha_short_df=alpha_short_df,
        wf_short_df=wf_short_df
    )

    # 6. Print summary
    print(f"\n{'='*80}")
    print(f"EXECUTION COMPLETE")
    print(f"{'='*80}")
    print(f"Symbols processed: {len(symbol_portfolios)}/{len(symbol_strategies)}")
    print(f"Output file: {output_file}")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    if len(sys.argv) == 1:
        interactive_mode()
    else:
        main()
