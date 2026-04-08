"""
AQS Trading Journal Generator
==============================
Scans WFAlphaResults for Risk_Optimized_Compilation Excel files (Stage 8),
extracts strategies from the 'Risk_Optimized_Strategies' worksheet, and
builds/appends to AQS_Trading_Journal.xlsx.  Falls back to Final_Compilation
(Stage 7 / 'Final_Alphas') when no risk-optimized files are found.

Features:
- Prefers Risk_Optimized_Compilation_*.xlsx (includes SL/TP/TSL columns)
- Falls back to Final_Compilation_*.xlsx when Stage 8 has not been run
- Extracts strategies and deduplicates across runs
- Scans ibkr_deployment/trades/ for live/dry-run execution records
- Reads deploy_ibkr_portfolio/state/ for current positions & risk state
- Creates a polished multi-sheet journal with:
    • Dashboard         – fleet overview & key stats
    • Active Strategy   – live deployed strategy monitor (trades + positions)
    • Strategy Registry – all validated alpha strategies
    • Performance Summary – per-symbol aggregated metrics
    • Changelog         – record of every import run
- If journal already exists, only appends NEW strategies (no duplicates)
- Delivers journal via email (Gmail SMTP) and/or Telegram Bot API
"""

import os
import re
import sys
import json
import glob
import shutil
import smtplib
import argparse
import warnings
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import requests

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ============================================================================
# CONFIGURATION
# ============================================================================
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = SCRIPT_DIR

WF_RESULTS_DIR      = os.path.join(PROJECT_DIR, "WFAlphaResults")
TRADES_DIR          = os.path.join(PROJECT_DIR, "ibkr_deployment", "trades")
PORTFOLIO_STATE_DIR = os.path.join(PROJECT_DIR, "deploy_ibkr_portfolio", "state")
JOURNAL_PATH        = os.path.join(PROJECT_DIR, "AQS_Trading_Journal.xlsx")
TELEGRAM_CONFIG_PATH = os.path.join(PROJECT_DIR, "ibkr_deployment", "telegram_config.json")
EMAIL_CONFIG_PATH    = os.path.join(PROJECT_DIR, "ibkr_deployment", "email_config.json")

FINAL_COMP_GLOB  = "Final_Compilation_*.xlsx"
SHEET_FINAL      = "Final_Alphas"            # Stage 7 worksheet
RISK_OPT_GLOB    = "Risk_Optimized_Compilation_*.xlsx"
RISK_OPT_SHEET   = "Risk_Optimized_Strategies"  # Stage 8 worksheet
SNAPSHOT_FILE    = "_snapshot_all_trades.csv"
TRADES_GLOB      = "trades_*_all.csv"

# Strategy dedup key — uniquely identifies a strategy regardless of source file.
# Length/Entry/Exit are Stage 7 optimization outputs not carried forward to Stage 8
# (Risk_Optimized_Compilation), so they are excluded from the key permanently.
DEDUP_COLS = [
    "Exchange", "Symbol", "Interval", "Data Point",
    "Model", "Entry / Exit Model",
]

# Decimal places used when rounding float key columns before deduplication.
# Must be consistent between compile_all_strategies and merge_new_strategies.
DEDUP_PRECISION = 8

# orderRef values that are internal bookkeeping, not real strategy tags
EXCLUDED_ORDER_REFS: frozenset = frozenset({"UNTAGGED", "ActivityMonitor", "", None})

# ── Risk_Optimized_Compilation column name normalisation ─────────────────────
# Stage 8 files use Optimized_*/Optimal_* prefixes; the rest of the codebase
# expects the plain names from Stage 7 (Final_Compilation).
# df.rename() silently skips keys that are absent — this is intentional so the
# same map can be applied without knowing which variant is present.
RISK_OPT_COL_MAP: dict = {
    # Performance metrics (Optimized_ prefix)
    "Optimized_Sharpe":         "Sharpe",
    "Optimized_MDD":            "MDD",
    "Optimized_Annual_Return":  "Annual Return",
    "Optimized_Trade_Count":    "Trade Count",
    "Optimized_Calmar":         "Calmar Ratio",
    "Optimized_Calmar_Ratio":   "Calmar Ratio",
    "Optimized_Cumulative_PnL": "Cumulative PnL",
    "Optimized_PnL_Ratio":      "PnL Ratio",
    "Optimized_Buy_Hold":       "Buy & Hold",
    "Optimized_Buy_&_Hold":     "Buy & Hold",
    # Risk parameters (Optimal_ prefix)
    "Optimal_SL":               "SL",
    "Optimal_TP":               "TP",
    "Optimal_TSL":              "TSL",
    # Alternative spellings seen in some pipeline versions
    "Optimized_SL":             "SL",
    "Optimized_TP":             "TP",
    "Optimized_TSL":            "TSL",
    "Optimal_Sharpe":           "Sharpe",
    "Optimal_MDD":              "MDD",
    "Optimal_Annual_Return":    "Annual Return",
}

# ============================================================================
# COLOUR PALETTE  (dark quant / Bloomberg-style)
# ============================================================================
C = {
    # Backgrounds
    "navy":       "0D1B2A",   # deep navy  (dashboard header)
    "midnight":   "1A2744",   # midnight   (section headers)
    "slate":      "2C3E5D",   # slate blue (sub-headers)
    "charcoal":   "1E2B3C",   # charcoal   (alt row 1)
    "dark_row":   "141F2E",   # dark row   (alt row 2)
    "panel":      "0F172A",   # panel bg   (dashboard cards)
    "border":     "2A4066",   # border     (cell borders)

    # Accents
    "gold":       "D4A843",   # AQ gold    (highlights, badges)
    "teal":       "00B4D8",   # teal       (metrics)
    "green":      "22C55E",   # green      (positive / active)
    "red":        "EF4444",   # red        (negative / retired)
    "amber":      "F59E0B",   # amber      (watchlist)
    "silver":     "94A3B8",   # silver     (secondary text)

    # Text
    "white":      "FFFFFF",
    "light_gray": "CBD5E1",
    "dim":        "64748B",
}

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="FFFFFF", name="Calibri") -> Font:
    return Font(name=name, size=size, bold=bold, color=color)

def _border(color="2A4066") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _fmt_ts(cell, val: object) -> None:
    """Write a datetime value into a cell, silently skipping unparseable values."""
    if pd.notna(val):
        try:
            cell.value = pd.Timestamp(val).strftime("%Y-%m-%d %H:%M")
        except Exception as e:
            print(f"[WARN] Could not format timestamp {val!r}: {e}")


# ============================================================================
# DEDUP KEY HELPER
# ============================================================================

def _make_dedup_key(df: pd.DataFrame, key_cols: list[str]) -> pd.Series:
    """Build a tuple key for strategy deduplication.

    Converts all key columns to strings and returns a Series of tuples
    suitable for set membership tests.
    """
    tmp = df[key_cols].copy()
    return tmp.astype(str).apply(lambda r: tuple(r), axis=1)

# ============================================================================
# STEP 1 – SCAN & EXTRACT
# ============================================================================

def _search_dirs(wf_dir: str) -> list[str]:
    """Return wf_dir plus all immediate subfolders that are not merged_ibkr*."""
    dirs = [wf_dir]
    try:
        for entry in os.scandir(wf_dir):
            if entry.is_dir() and not entry.name.startswith("merged_ibkr"):
                dirs.append(entry.path)
    except OSError:
        pass
    return dirs


def find_source_files(wf_dir: str) -> tuple[list[str], str]:
    """Locate strategy source files, preferring Stage 8 (Risk_Optimized_Compilation).

    Searches wf_dir and all immediate non-merged_ibkr* subfolders.
    Returns (files, sheet_name).  Falls back to Final_Compilation / Final_Alphas
    when no risk-optimized files exist.
    """
    search_dirs = _search_dirs(wf_dir)

    risk_files = sorted(
        f for d in search_dirs for f in glob.glob(os.path.join(d, RISK_OPT_GLOB))
    )
    if risk_files:
        print(f"\n[SCAN] Found {len(risk_files)} Risk_Optimized_Compilation file(s) in WFAlphaResults/ (incl. subfolders)")
        for f in risk_files:
            rel = os.path.relpath(f, wf_dir)
            print(f"       • {rel}")
        return risk_files, RISK_OPT_SHEET

    final_files = sorted(
        f for d in search_dirs for f in glob.glob(os.path.join(d, FINAL_COMP_GLOB))
    )
    print(f"\n[SCAN] No Risk_Optimized_Compilation files found — falling back to Final_Compilation.")
    print(f"[SCAN] Found {len(final_files)} Final_Compilation file(s) in WFAlphaResults/ (incl. subfolders)")
    for f in final_files:
        rel = os.path.relpath(f, wf_dir)
        print(f"       • {rel}")
    return final_files, SHEET_FINAL


def extract_final_alphas(filepath: str, sheet_name: str) -> pd.DataFrame | None:
    """Read the specified worksheet from a compilation file.

    For Risk_Optimized_Compilation sheets (Stage 8), column names are
    normalised to the standard Stage 7 names expected by the rest of the
    codebase (e.g. Optimized_Sharpe → Sharpe, Optimal_SL → SL).
    """
    try:
        xl = pd.ExcelFile(filepath)
        if sheet_name not in xl.sheet_names:
            print(f"  [WARN] '{sheet_name}' not found in {os.path.basename(filepath)} — skipping")
            return None
        df = xl.parse(sheet_name)
        if df.empty:
            return None

        # Normalise Risk_Optimized column names to the standard set
        if sheet_name == RISK_OPT_SHEET:
            df = df.rename(columns=RISK_OPT_COL_MAP)
            if "Sharpe" in df.columns:
                print(f"  [MAP]  Column normalisation applied to {os.path.basename(filepath)}"
                      f"  — Sharpe/MDD/Annual Return now available.")
            else:
                print(f"  [WARN] 'Sharpe' still not found after normalisation in "
                      f"{os.path.basename(filepath)} — check column names in the sheet.")

            # Derive Exchange and Interval from the filename.
            # Pattern: Risk_Optimized_Compilation_{exchange}_{interval}_{date}_{symbols}.xlsx
            m = re.search(r'Risk_Optimized_Compilation_([^_]+)_([^_]+)_', os.path.basename(filepath), re.IGNORECASE)
            if m:
                df["Exchange"] = m.group(1)
                df["Interval"] = m.group(2)
            else:
                print(f"  [WARN] Could not parse Exchange/Interval from filename: {os.path.basename(filepath)}")

        df["_source_file"] = os.path.basename(filepath)
        df["_import_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return df
    except Exception as e:
        print(f"  [ERR] Could not read {os.path.basename(filepath)}: {e}")
        return None


def compile_all_strategies(files: list[str], sheet_name: str) -> pd.DataFrame:
    """Compile and deduplicate strategies across all compilation files."""
    frames = []
    for f in files:
        df = extract_final_alphas(f, sheet_name)
        if df is not None:
            frames.append(df)
            print(f"  [OK]  Extracted {len(df):>3} strategies from {os.path.basename(f)}")

    if not frames:
        print("[WARN] No strategies extracted — nothing to write.")
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)

    # Drop the '#' column if present; we'll renumber later
    if "#" in combined.columns:
        combined = combined.drop(columns=["#"])

    # Dedup: keep LAST occurrence so the most-recent file wins
    key_cols = [c for c in DEDUP_COLS if c in combined.columns]
    combined = combined.drop_duplicates(subset=key_cols, keep="last")
    combined = combined.reset_index(drop=True)

    print(f"\n[COMPILE] Total unique strategies: {len(combined)}")
    return combined


# ============================================================================
# STEP 2 – JOURNAL MERGE  (append-only)
# ============================================================================

REGISTRY_SHEET    = "Strategy Registry"
DASHBOARD_SHEET   = "Dashboard"
ACTIVE_SHEET      = "Active Strategy"
PERF_SHEET        = "Performance Summary"
CHANGE_SHEET      = "Changelog"
COMPARISON_SHEET  = "Backtest vs Execution"


def load_existing_registry(journal_path: str) -> pd.DataFrame:
    """Load existing Strategy Registry if journal already exists.

    The sheet has a banner row at row 1 and column headers at row 2,
    so we skip the first row (header=1).
    """
    if not os.path.exists(journal_path):
        return pd.DataFrame()
    try:
        # header=1 → row index 1 (the second row) is used as column headers
        df = pd.read_excel(journal_path, sheet_name=REGISTRY_SHEET, header=1)
        # Drop any fully-empty rows (can appear at end of formatted sheets)
        df = df.dropna(how="all")
        print(f"[JOURNAL] Loaded {len(df)} existing strategies from journal.")
        return df
    except Exception as e:
        print(f"[WARN] Could not load existing registry: {e}")
        return pd.DataFrame()


def merge_new_strategies(existing: pd.DataFrame, incoming: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Append new strategies and upsert risk columns for existing ones.

    New strategies (not in existing by MERGE_KEY_COLS) are appended.
    Existing strategies that match a row in incoming have their SL/TP/TSL
    and Exchange/Interval columns updated — so re-running after Stage 8
    enriches the journal with risk parameters without requiring a full reset.

    Returns (merged_df, new_count).
    """
    # Columns used to match incoming rows against the existing journal.
    # Exchange and Interval are intentionally excluded: old journals may lack
    # them entirely, and they may carry stale/differently-cased values from a
    # previous backfill.  They are instead back-filled via UPSERT below.
    # Symbol + Data Point + Model + Entry/Exit Model uniquely identifies a
    # strategy and is stable across all code versions.
    MERGE_KEY_COLS = ["Symbol", "Data Point", "Model", "Entry / Exit Model"]

    # Enrich incoming with journal metadata columns
    if "Status" not in incoming.columns:
        incoming = incoming.copy()
        incoming["Status"] = "Active"
    if "Date Added" not in incoming.columns:
        incoming["Date Added"] = incoming.get("_import_date",
                                               datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    if "Source File" not in incoming.columns:
        incoming["Source File"] = incoming.get("_source_file", "")

    if existing.empty:
        return incoming, len(incoming)

    key_cols = [c for c in MERGE_KEY_COLS
                if c in existing.columns and c in incoming.columns]

    if not key_cols:
        # Can't dedup — just stack everything
        merged = pd.concat([existing, incoming], ignore_index=True)
        return merged, len(incoming)

    existing_keys = set(_make_dedup_key(existing, key_cols))
    incoming = incoming.copy()
    incoming["_key"] = _make_dedup_key(incoming, key_cols)
    new_rows = incoming[~incoming["_key"].isin(existing_keys)].drop(columns=["_key"])

    # ── Upsert columns for already-existing strategies ────────────────────────
    # Includes risk parameters (SL/TP/TSL) AND performance metrics (Sharpe etc.)
    # so that re-running after Stage 8 always refreshes the journal with the
    # latest optimised values — without requiring a full registry reset.
    UPSERT_COLS = [
        # Structural identity columns derived from the filename (Stage 8 files).
        # These are absent from the existing registry on first run and must be
        # back-filled so that the BT-match lookup can use a direct prefix match.
        "Exchange", "Interval",
        # Risk parameters and performance metrics from Stage 8 optimisation.
        "SL", "TP", "TSL",
        "Sharpe", "MDD", "Annual Return", "Trade Count",
        "Calmar Ratio", "Cumulative PnL", "Buy & Hold", "PnL Ratio",
    ]
    upsert_cols_present = [c for c in UPSERT_COLS if c in incoming.columns]
    if upsert_cols_present:
        # Build a lookup: dedup key → latest values from incoming
        upsert_lookup = (
            incoming[incoming["_key"].isin(existing_keys)][["_key"] + upsert_cols_present]
            .drop_duplicates(subset=["_key"], keep="last")
            .set_index("_key")
        )
        if not upsert_lookup.empty:
            existing = existing.copy()
            existing["_key"] = _make_dedup_key(existing, key_cols)
            for col in upsert_cols_present:
                if col not in existing.columns:
                    existing[col] = np.nan
                # Prefer incoming value; keep existing only where incoming is NaN
                incoming_vals = existing["_key"].map(upsert_lookup[col])
                existing[col] = incoming_vals.where(incoming_vals.notna(), existing[col])
            existing = existing.drop(columns=["_key"])
            n_updated = upsert_lookup.shape[0]
            print(f"[MERGE]  Updated {n_updated} existing strategy row(s) "
                  f"with columns: {', '.join(upsert_cols_present)}.")

    incoming = incoming.drop(columns=["_key"])
    merged = pd.concat([existing, new_rows], ignore_index=True)
    return merged, len(new_rows)


def clean_registry(df: pd.DataFrame) -> pd.DataFrame:
    """Remove internal columns, renumber, and order columns."""
    internal = ["_source_file", "_import_date", "_key"]
    df = df.drop(columns=[c for c in internal if c in df.columns], errors="ignore")

    # Drop raw Optimized_*/Optimal_* columns that may have leaked from earlier
    # imports (before column normalisation was applied).  Their values are now
    # carried by the standard-named columns (Sharpe, MDD, SL, etc.).
    raw_prefixes = ("Optimized_", "Optimal_")
    residue = [c for c in df.columns if c.startswith(raw_prefixes)]
    if residue:
        df = df.drop(columns=residue)

    # Always drop # before we re-number
    df = df.drop(columns=["#"], errors="ignore")

    # Preferred column order (# excluded — added below)
    front = [
        "Status", "Exchange", "Symbol", "Interval", "Data Point",
        "Model", "Entry / Exit Model", "Variant", "Length", "Entry", "Exit",
        "SL", "TP", "TSL",
        "Sharpe", "MDD", "Trade Count", "Annual Return",
        "Calmar Ratio", "Cumulative PnL", "Buy & Hold", "PnL Ratio",
        "Date Added", "Source File",
    ]
    existing_front = [c for c in front if c in df.columns]
    other_cols = [c for c in df.columns if c not in front]
    ordered = existing_front + other_cols

    df = df[[c for c in ordered if c in df.columns]]

    # Sort alphabetically by Symbol before numbering
    if "Symbol" in df.columns:
        df = df.sort_values("Symbol", ascending=True, kind="stable")

    df.insert(0, "#", range(1, len(df) + 1))

    return df.reset_index(drop=True)


def audit_and_dedupe_registry(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Final-pass duplicate check on the combined registry.

    Uses the full DEDUP_COLS key (Exchange + Symbol + Interval + Data Point +
    Model + Entry/Exit Model).  By the time this runs, UPSERT has already
    back-filled Exchange/Interval onto all rows, so the 6-col key is reliable.
    Keeps the LAST occurrence of any duplicate (most recent import wins).
    Returns (deduplicated_df, n_duplicates_removed).
    """
    if df.empty:
        return df, 0

    key_cols = [c for c in DEDUP_COLS if c in df.columns]
    if not key_cols:
        return df, 0

    keys = _make_dedup_key(df, key_cols)

    n_before = len(df)
    df_clean = df[~keys.duplicated(keep="last")].reset_index(drop=True)
    dupes = n_before - len(df_clean)

    if dupes > 0:
        print(f"[AUDIT]  {dupes} duplicate row(s) removed  "
              f"({n_before} -> {len(df_clean)} strategies)")
    else:
        print(f"[AUDIT]  No duplicates found in registry ({n_before} rows, all unique)")

    return df_clean, dupes


BACKUP_KEEP = 7   # number of most-recent backups to retain

def backup_journal(journal_path: str) -> str | None:
    """Copy the existing journal to a timestamped backup, then prune old backups.

    Backup is stored alongside the journal:
        AQS_Trading_Journal_backup_YYYYMMDD_HHMMSS.xlsx

    Only the most-recent BACKUP_KEEP backups are kept; older ones are deleted.
    Returns the backup path, or None if no existing journal was found.
    """
    if not os.path.exists(journal_path):
        return None

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(journal_path)
    backup_path = f"{base}_backup_{ts}{ext}"
    shutil.copy2(journal_path, backup_path)
    print(f"[BACKUP] Journal backed up -> {os.path.basename(backup_path)}")

    # Prune: keep only the BACKUP_KEEP most-recent backups
    backup_dir  = os.path.dirname(journal_path)
    backup_glob = os.path.basename(base) + "_backup_*.xlsx"
    existing_backups = sorted(
        glob.glob(os.path.join(backup_dir, backup_glob))
    )
    for old in existing_backups[:-BACKUP_KEEP]:
        try:
            os.remove(old)
            print(f"[BACKUP] Pruned old backup -> {os.path.basename(old)}")
        except OSError as e:
            print(f"[WARN] Could not prune backup {old}: {e}")

    return backup_path


# ============================================================================
# STEP 3 – BUILD PERFORMANCE SUMMARY
# ============================================================================

def _avg_cost_pnl(group: pd.DataFrame) -> float:
    """Realized P&L for a single (strategy × symbol) trade series.

    Uses the average cost basis method — identical to how 05_monitor.py tracks P&L:
      - BOT  → increases long position (or covers short), updates running average cost
      - SLD  → closes long position at average cost (or opens/adds to short)
      - Commission split proportionally on zero-crossing trades: only the closing
        fraction is charged to realized P&L; the opening fraction enters cost basis
      - Open (unrealised) positions are NOT counted as gains or losses

    Handles short positions and zero-crossing trades (long→short and short→long).
    """
    pnl        = 0.0
    position   = 0.0
    avg_cost   = 0.0
    total_cost = 0.0

    for _, row in group.sort_values("datetime").iterrows():
        qty   = abs(float(row.get("quantity",   0)))
        price =     float(row.get("price",      0))
        comm  = abs(float(row.get("commission", 0)))
        side  =  str(row.get("side", ""))

        if side == "BOT":
            if position < 0:
                # Covering a short position
                close_qty  = min(qty, abs(position))
                close_comm = comm * (close_qty / qty) if qty > 0 else comm
                pnl       += close_qty * (avg_cost - price) - close_comm
                position  += close_qty
                remaining  = qty - close_qty
                if position == 0:
                    total_cost = avg_cost = 0.0
                else:
                    total_cost = abs(position) * avg_cost
                if remaining > 0:
                    # Crossed zero: excess qty opens a new long position.
                    # Carry the remaining commission fraction into cost basis.
                    open_comm  = comm - close_comm
                    total_cost = remaining * price + open_comm
                    position   = remaining
                    avg_cost   = total_cost / position
            else:
                # Adding to (or opening) a long position
                total_cost += qty * price
                position   += qty
                avg_cost    = total_cost / position if position > 0 else 0.0
                pnl        -= comm

        elif side == "SLD":
            if position > 0:
                # Closing (or reducing) a long position
                close_qty  = min(qty, position)
                close_comm = comm * (close_qty / qty) if qty > 0 else comm
                pnl       += close_qty * (price - avg_cost) - close_comm
                position  -= close_qty
                remaining  = qty - close_qty
                if position == 0:
                    total_cost = avg_cost = 0.0
                else:
                    total_cost = position * avg_cost
                if remaining > 0:
                    # Crossed zero: excess qty opens a new short position.
                    # Carry the remaining commission fraction into cost basis.
                    open_comm  = comm - close_comm
                    total_cost = remaining * price + open_comm
                    position   = -remaining
                    avg_cost   = total_cost / remaining
            else:
                # Opening or adding to a short position
                pnl        -= comm
                position   -= qty
                total_cost += qty * price
                avg_cost    = abs(total_cost / position) if position != 0 else 0.0

    return round(pnl, 2)


def build_performance_summary(registry: pd.DataFrame,
                               trades: pd.DataFrame = None) -> pd.DataFrame:
    """Aggregate key backtest metrics per Symbol and cross-reference trade logs.

    Deployment columns added (from trades):
      Deployed       – Yes / No flag
      Deployed Strats – count of unique orderRefs for that symbol in trades
      Live Trades    – count of live (non-dry-run) trades for that symbol
      Dry-Run Trades – count of dry-run trades for that symbol
      Net P&L ($)    – cumulative net P&L across all strategies for that symbol
    """
    if registry.empty:
        return pd.DataFrame()

    metric_cols = {
        "Sharpe":        "Avg Sharpe",
        "MDD":           "Avg MDD",
        "Annual Return": "Avg Annual Return",
        "Calmar Ratio":  "Avg Calmar",
        "Cumulative PnL":"Avg Cumulative PnL",
        "Trade Count":   "Avg Trade Count",
        "PnL Ratio":     "Avg PnL Ratio",
    }
    available = {k: v for k, v in metric_cols.items() if k in registry.columns}

    agg_funcs = {col: "mean" for col in available}
    agg_funcs["#"] = "count"

    grp = registry.groupby("Symbol", sort=False).agg(agg_funcs).reset_index()
    grp = grp.rename(columns={"#": "Strategy Count", **available})

    # Best strategy per symbol
    if "Sharpe" in registry.columns:
        best = (
            registry.sort_values("Sharpe", ascending=False)
            .drop_duplicates(subset=["Symbol"])
            [["Symbol", "Model", "Entry / Exit Model", "Sharpe"]]
            .rename(columns={"Sharpe": "Best Sharpe", "Model": "Best Model",
                             "Entry / Exit Model": "Best Entry/Exit"})
        )
        grp = grp.merge(best, on="Symbol", how="left")

    # ── Deployment cross-reference (from trade logs) ──────────────────────────
    if trades is not None and not trades.empty and "symbol" in trades.columns:
        tagged = trades[~trades["orderRef"].isin(EXCLUDED_ORDER_REFS)].dropna(subset=["orderRef"])

        # Normalise trade symbols to upper-case so they match the registry
        trades_norm  = trades.copy()
        tagged_norm  = tagged.copy()
        trades_norm["symbol"] = trades_norm["symbol"].astype(str).str.strip().str.upper()
        tagged_norm["symbol"] = tagged_norm["symbol"].astype(str).str.strip().str.upper()

        deploy_rows = []
        for sym in grp["Symbol"]:
            sym_trades  = tagged_norm[tagged_norm["symbol"] == sym]
            sym_all     = trades_norm[trades_norm["symbol"] == sym]
            n_strats    = sym_trades["orderRef"].nunique()
            n_live      = (sym_all["trade_mode"] == "Live").sum()   if not sym_all.empty else 0
            n_dry       = (sym_all["trade_mode"] == "Dry Run").sum() if not sym_all.empty else 0

            # Realized P&L — sum across each strategy that trades this symbol,
            # computed per (strategy × symbol) using average cost basis.
            # This matches Streamlit's "Realized P&L" and avoids inflated
            # negative values from open positions.
            if not sym_trades.empty:
                net = sum(
                    _avg_cost_pnl(strat_grp)
                    for _, strat_grp in sym_trades.groupby("orderRef")
                )
            else:
                net = 0.0

            deploy_rows.append({
                "Symbol":          sym,
                "Deployed":        "YES" if n_strats > 0 else "no",
                "Deployed Strats": int(n_strats),
                "Live Trades":     int(n_live),
                "Dry-Run Trades":  int(n_dry),
                "Net P&L ($)":     round(net, 2),
            })

        deploy_df = pd.DataFrame(deploy_rows)
        grp = grp.merge(deploy_df, on="Symbol", how="left")
    else:
        grp["Deployed"]        = "no"
        grp["Deployed Strats"] = 0
        grp["Live Trades"]     = 0
        grp["Dry-Run Trades"]  = 0
        grp["Net P&L ($)"]     = 0.0

    # Sort: deployed first, then by Avg Sharpe desc
    grp["_deploy_sort"] = (grp["Deployed"] == "YES").astype(int)
    sort_cols = ["_deploy_sort"]
    if "Avg Sharpe" in grp.columns:
        sort_cols.append("Avg Sharpe")
    grp = grp.sort_values(sort_cols, ascending=[False] + [False] * (len(sort_cols) - 1))
    grp = grp.drop(columns=["_deploy_sort"])

    grp.insert(0, "#", range(1, len(grp) + 1))
    return grp.reset_index(drop=True)


# ============================================================================
# STEP 4 – SHEET BUILDERS
# ============================================================================

# ---------- helpers -----------------------------------------------------------
def _status_fill(status: str) -> PatternFill:
    mapping = {
        "Active":    _fill(C["green"]),
        "Watchlist": _fill(C["amber"]),
        "Retired":   _fill(C["red"]),
    }
    return mapping.get(status, _fill(C["slate"]))


def _num_fmt(col: str) -> str:
    pct_cols  = {"MDD", "Annual Return", "PnL Ratio", "Avg MDD",
                 "Avg Annual Return", "Avg PnL Ratio",
                 "SL", "TP", "TSL"}
    rat_cols  = {"Sharpe", "Calmar Ratio", "Cumulative PnL", "Buy & Hold",
                 "Avg Sharpe", "Avg Calmar", "Avg Cumulative PnL", "Best Sharpe"}
    int_cols  = {"Trade Count", "Avg Trade Count", "Strategy Count",
                 "Active", "Watchlist", "Retired"}
    if col in pct_cols:
        return "0.00%"
    if col in rat_cols:
        return "0.0000"
    if col in int_cols:
        return "#,##0"
    return "General"

# ---------- DASHBOARD ---------------------------------------------------------

def build_dashboard(ws, registry: pd.DataFrame, perf: pd.DataFrame,
                    scan_files: list[str], new_count: int,
                    trades: pd.DataFrame = None):
    """Build the Dashboard sheet.

    Two distinct data sources are always shown separately:
      - REGISTRY  : all strategies compiled from Final_Alphas worksheets
                    (validated alpha pool — may not be deployed)
      - RUNNING   : strategies that appear in the trades folder
                    (actually deployed and executed at least one trade)
    """
    if trades is None:
        trades = pd.DataFrame()

    ws.sheet_view.showGridLines = False
    ws.tab_color = C["gold"]

    # ── Title banner ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:R3")
    title_cell = ws["A1"]
    title_cell.value = "AQS TRADING JOURNAL"
    title_cell.font  = Font(name="Calibri", size=28, bold=True, color=C["gold"])
    title_cell.fill  = _fill(C["navy"])
    title_cell.alignment = _center()

    n_running = _count_running_strategies(trades)
    ws.merge_cells("A4:R4")
    sub = ws["A4"]
    sub.value = (f"Alpha Quant System  |  "
                 f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                 f"Registry: {len(registry)} alphas  |  "
                 f"Running: {n_running} strategies")

    sub.font      = _font(size=11, color=C["silver"])
    sub.fill      = _fill(C["midnight"])
    sub.alignment = _center()

    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 22

    # ── KPI Row 1: Registry Universe ──────────────────────────────────────────
    _draw_section_header(ws, row=6, col=1, width=18,
                         text="STRATEGY REGISTRY  —  All Validated Alphas")
    registry_kpis = _build_registry_kpis(registry, new_count)
    _draw_kpi_cards(ws, registry_kpis, start_row=7, start_col=1)

    # ── KPI Row 2: Deployed / Running ────────────────────────────────────────
    kpi_end_row = 7 + 4 + 1          # start(7) + card_height(4) + gap(1)
    _draw_section_header(ws, row=kpi_end_row, col=1, width=18,
                         text="DEPLOYED / RUNNING  —  Strategies with Trade Activity")
    running_kpis = _build_running_kpis(trades)
    _draw_kpi_cards(ws, running_kpis, start_row=kpi_end_row + 1, start_col=1)

    # ── Registry summary table ────────────────────────────────────────────────
    reg_table_row = kpi_end_row + 1 + 4 + 2
    _draw_section_header(ws, row=reg_table_row, col=1, width=18,
                         text="REGISTRY BREAKDOWN BY SYMBOL")
    if not perf.empty:
        _draw_registry_symbol_table(ws, perf, start_row=reg_table_row + 1)
        reg_table_end = reg_table_row + 1 + len(perf) + 2
    else:
        reg_table_end = reg_table_row + 3

    # ── Running strategies table ──────────────────────────────────────────────
    _draw_section_header(ws, row=reg_table_end, col=1, width=18,
                         text="RUNNING / ACTIVE STRATEGIES  —  from Trade Logs")
    _draw_running_strategies_table(ws, trades, start_row=reg_table_end + 1)
    run_table_end = reg_table_end + 1 + max(n_running, 1) + 2

    # ── Source files table ────────────────────────────────────────────────────
    _draw_section_header(ws, row=run_table_end, col=1, width=18,
                         text="SCANNED SOURCE FILES")
    _draw_file_table(ws, scan_files, start_row=run_table_end + 1)

    # Column widths — uniform for dashboard
    for col_idx in range(1, 19):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


def _count_running_strategies(trades: pd.DataFrame) -> int:
    """Count unique strategies that have at least one trade."""
    if trades.empty or "orderRef" not in trades.columns:
        return 0
    tagged = trades[~trades["orderRef"].isin(EXCLUDED_ORDER_REFS)].dropna(subset=["orderRef"])
    return tagged["orderRef"].nunique()


def _build_registry_kpis(registry: pd.DataFrame, new_count: int) -> list[dict]:
    """KPI cards for the strategy registry (Final_Alphas universe)."""
    kpis = [
        {"label": "TOTAL ALPHAS",
         "value": str(len(registry)),
         "color": C["teal"]},
        {"label": "SYMBOLS",
         "value": str(registry["Symbol"].nunique() if "Symbol" in registry.columns else 0),
         "color": C["slate"]},
        {"label": "NEW THIS RUN",
         "value": str(new_count),
         "color": C["gold"]},
    ]
    if "Sharpe" in registry.columns:
        kpis += [
            {"label": "AVG SHARPE",
             "value": f"{registry['Sharpe'].mean():.3f}",
             "color": C["midnight"]},
            {"label": "BEST SHARPE",
             "value": f"{registry['Sharpe'].max():.3f}",
             "color": C["midnight"]},
        ]
    return kpis


def _build_running_kpis(trades: pd.DataFrame) -> list[dict]:
    """KPI cards for deployed/running strategies (from trade logs)."""
    if trades.empty:
        return [
            {"label": "RUNNING STRATEGIES", "value": "0", "color": C["green"]},
            {"label": "SYMBOLS ACTIVE",     "value": "0", "color": C["teal"]},
            {"label": "TOTAL TRADES",        "value": "0", "color": C["slate"]},
            {"label": "LIVE TRADES",         "value": "0", "color": C["teal"]},
            {"label": "DRY-RUN TRADES",      "value": "0", "color": C["amber"]},
        ]

    tagged = trades[~trades["orderRef"].isin(EXCLUDED_ORDER_REFS)].dropna(subset=["orderRef"])
    n_running = tagged["orderRef"].nunique()
    n_symbols = tagged["symbol"].nunique() if "symbol" in tagged.columns else 0
    n_live    = (trades["trade_mode"] == "Live").sum()
    n_dry     = (trades["trade_mode"] == "Dry Run").sum()

    return [
        {"label": "RUNNING STRATEGIES", "value": str(n_running),      "color": C["green"]},
        {"label": "SYMBOLS ACTIVE",     "value": str(n_symbols),      "color": C["teal"]},
        {"label": "TOTAL TRADES",       "value": str(len(trades)),     "color": C["slate"]},
        {"label": "LIVE TRADES",        "value": str(n_live),          "color": C["teal"]},
        {"label": "DRY-RUN TRADES",     "value": str(n_dry),           "color": C["amber"]},
    ]


def _draw_kpi_cards(ws, kpis: list[dict], start_row: int, start_col: int):
    cols_per_card = 2
    card_height   = 4

    for i, kpi in enumerate(kpis):
        c0 = start_col + i * cols_per_card
        c1 = c0 + cols_per_card - 1
        r0 = start_row
        r1 = start_row + card_height - 1

        ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c1)
        label_cell = ws.cell(row=r0, column=c0, value=kpi["label"])
        label_cell.font      = _font(bold=True, size=8, color=C["white"])
        label_cell.fill      = _fill(kpi["color"])
        label_cell.alignment = _center()

        ws.merge_cells(start_row=r0+1, start_column=c0, end_row=r1, end_column=c1)
        val_cell = ws.cell(row=r0+1, column=c0, value=kpi["value"])
        val_cell.font      = Font(name="Calibri", size=22, bold=True, color=C["white"])
        val_cell.fill      = _fill(C["panel"])
        val_cell.alignment = _center()

        for r in range(r0, r1+1):
            ws.row_dimensions[r].height = 16

    ws.row_dimensions[start_row].height = 14


def _draw_section_header(ws, row: int, col: int, width: int, text: str):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + width - 1)
    cell = ws.cell(row=row, column=col, value=f"  {text}")
    cell.font      = _font(bold=True, size=11, color=C["gold"])
    cell.fill      = _fill(C["midnight"])
    cell.alignment = _left()
    ws.row_dimensions[row].height = 20


def _draw_registry_symbol_table(ws, perf: pd.DataFrame, start_row: int):
    """Registry breakdown table — all validated alphas, no Status-based columns."""
    cols_show = ["Symbol", "Strategy Count", "Avg Sharpe", "Avg MDD",
                 "Avg Annual Return", "Best Sharpe", "Best Model"]
    cols_show = [c for c in cols_show if c in perf.columns]
    sub = perf[cols_show].copy()

    for ci, col_name in enumerate(sub.columns, start=1):
        cell = ws.cell(row=start_row, column=ci, value=col_name)
        cell.font      = _font(bold=True, size=10, color=C["white"])
        cell.fill      = _fill(C["slate"])
        cell.alignment = _center(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[start_row].height = 28

    for ri, (_, row) in enumerate(sub.iterrows(), start=1):
        r = start_row + ri
        fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])
        for ci, (col_name, val) in enumerate(zip(sub.columns, row), start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill          = fill
            cell.border        = _border()
            cell.number_format = _num_fmt(col_name)
            if col_name == "Symbol":
                cell.font      = _font(bold=True, size=11, color=C["gold"])
                cell.alignment = _center()
            elif "Sharpe" in col_name or "MDD" in col_name or "Return" in col_name:
                color = C["teal"] if "Sharpe" in col_name or "Return" in col_name else C["red"]
                cell.font      = _font(size=10, color=color)
                cell.alignment = _center()
            else:
                cell.font      = _font(size=10, color=C["light_gray"])
                cell.alignment = _center()
        ws.row_dimensions[r].height = 18


def _draw_running_strategies_table(ws, trades: pd.DataFrame, start_row: int):
    """Mini-table of strategies derived entirely from the trades folder."""
    if trades.empty:
        ws.cell(row=start_row, column=1,
                value="  No trade data found — check ibkr_deployment/trades/").font = \
            _font(size=9, color=C["dim"])
        return

    tagged = trades[~trades["orderRef"].isin(EXCLUDED_ORDER_REFS)].dropna(subset=["orderRef"])
    if tagged.empty:
        ws.cell(row=start_row, column=1,
                value="  No tagged strategy trades found.").font = _font(size=9, color=C["dim"])
        return

    # Aggregate per strategy — realized P&L only, matching Streamlit
    def _agg(grp):
        net = sum(
            _avg_cost_pnl(sym_grp)
            for _, sym_grp in grp.groupby("symbol")
        )
        modes = grp["trade_mode"].unique()
        mode  = modes[0] if len(modes) == 1 else "Hybrid"
        return pd.Series({
            "Symbol":       grp["symbol"].iloc[0],
            "Trades":       len(grp),
            "Net P&L ($)":  net,
            "Trade Mode":   mode,
            "Last Trade":   grp["datetime"].max(),
        })

    summary = (tagged.groupby("orderRef")
                     .apply(_agg)
                     .reset_index()
                     .rename(columns={"orderRef": "Strategy"})
                     .sort_values("Trades", ascending=False)
                     .reset_index(drop=True))
    summary.insert(0, "#", range(1, len(summary) + 1))

    headers  = list(summary.columns)
    hdr_fill = _fill(C["slate"])

    for ci, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=ci, value=col_name)
        cell.font      = _font(bold=True, size=10, color=C["white"])
        cell.fill      = hdr_fill
        cell.alignment = _center(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[start_row].height = 28

    trade_mode_colors = {"Live": C["teal"], "Dry Run": C["amber"]}
    for ri, (_, row) in enumerate(summary.iterrows(), start=1):
        r = start_row + ri
        fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])
        for ci, col_name in enumerate(headers, start=1):
            val  = row[col_name]
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill   = fill
            cell.border = _border()

            if col_name == "#":
                cell.font = _font(bold=True, size=9, color=C["gold"])
                cell.alignment = _center()
            elif col_name == "Strategy":
                cell.font      = _font(size=9, color=C["light_gray"])
                cell.alignment = _left()
            elif col_name == "Symbol":
                cell.font      = _font(bold=True, size=10, color=C["gold"])
                cell.alignment = _center()
            elif col_name == "Net P&L ($)":
                color = C["green"] if pd.notna(val) and val >= 0 else C["red"]
                cell.font          = _font(bold=True, size=10, color=color)
                cell.alignment     = _center()
                cell.number_format = '#,##0.00'
            elif col_name == "Trade Mode":
                c = trade_mode_colors.get(str(val), C["silver"])
                cell.font      = _font(bold=True, size=9, color=c)
                cell.alignment = _center()
            elif col_name == "Last Trade":
                _fmt_ts(cell, val)
                cell.font      = _font(size=9, color=C["silver"])
                cell.alignment = _center()
            else:
                cell.font      = _font(size=10, color=C["silver"])
                cell.alignment = _center()
        ws.row_dimensions[r].height = 18


def _draw_file_table(ws, files: list[str], start_row: int):
    headers = ["#", "File Name", "Date Modified"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = _font(bold=True, size=10, color=C["white"])
        cell.fill = _fill(C["slate"])
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[start_row].height = 24

    for ri, fpath in enumerate(files, start=1):
        r = start_row + ri
        mtime = datetime.fromtimestamp(os.path.getmtime(fpath)).strftime("%Y-%m-%d")
        fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])

        for ci, val in enumerate([ri, os.path.basename(fpath), mtime], start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = fill
            cell.font = _font(size=10, color=C["light_gray"])
            cell.alignment = _left() if ci == 2 else _center()
            cell.border = _border()
        ws.row_dimensions[r].height = 18


# ---------- STRATEGY REGISTRY ------------------------------------------------

def build_strategy_registry(ws, registry: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["teal"]
    ws.freeze_panes = "A3"

    n_cols = len(registry.columns)

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    banner = ws.cell(row=1, column=1,
                     value="  STRATEGY REGISTRY  —  All Validated Alpha Strategies")
    banner.font      = Font(name="Calibri", size=13, bold=True, color=C["gold"])
    banner.fill      = _fill(C["navy"])
    banner.alignment = _left()
    ws.row_dimensions[1].height = 28

    # ── Column headers ────────────────────────────────────────────────────────
    metric_hi = {"Sharpe", "Calmar Ratio", "Annual Return", "Cumulative PnL",
                 "PnL Ratio"}
    metric_lo = {"MDD"}
    id_cols   = {"#", "Exchange", "Symbol", "Interval", "Data Point",
                 "Model", "Entry / Exit Model", "Variant", "Length",
                 "Entry", "Exit"}

    for ci, col_name in enumerate(registry.columns, start=1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.font      = _font(bold=True, size=10, color=C["white"])
        cell.fill      = _fill(C["midnight"])
        cell.alignment = _center(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[2].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(registry.iterrows(), start=1):
        r = ri + 2
        base_fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])

        for ci, col_name in enumerate(registry.columns, start=1):
            val  = row[col_name]
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = _border()
            cell.number_format = _num_fmt(col_name)

            if col_name == "Status":
                cell.fill      = _status_fill(str(val))
                cell.font      = _font(bold=True, size=9, color=C["white"])
                cell.alignment = _center()
            elif col_name == "#":
                cell.fill      = _fill(C["slate"])
                cell.font      = _font(bold=True, size=9, color=C["gold"])
                cell.alignment = _center()
            elif col_name in id_cols:
                cell.fill      = base_fill
                cell.font      = _font(size=10, color=C["light_gray"])
                cell.alignment = _center()
            elif col_name in metric_hi:
                # Colour gradient: good = teal, bad = dim
                cell.fill      = base_fill
                cell.font      = _font(size=10, color=C["teal"])
                cell.alignment = _center()
            elif col_name in metric_lo:
                cell.fill      = base_fill
                cell.font      = _font(size=10, color=C["red"])
                cell.alignment = _center()
            else:
                cell.fill      = base_fill
                cell.font      = _font(size=10, color=C["silver"])
                cell.alignment = _center()

        ws.row_dimensions[r].height = 18

    # ── Auto column widths ────────────────────────────────────────────────────
    fixed_widths = {
        "#":               5,  "Status": 11,
        "Exchange":       10,  "Symbol":  9,  "Interval":  9,
        "Data Point":     13,  "Model":  22,
        "Entry / Exit Model": 22, "Variant": 11,
        "Length":          8,  "Entry":   9,  "Exit":       9,
        "SL":              7,  "TP":      7,  "TSL":         7,
        "Sharpe":         10,  "MDD":    10,  "Trade Count": 11,
        "Annual Return":  13,  "Calmar Ratio": 13,
        "Cumulative PnL": 14,  "Buy & Hold":  11, "PnL Ratio": 10,
        "Date Added":     16,  "Source File": 36,
    }
    for ci, col_name in enumerate(registry.columns, start=1):
        ltr = get_column_letter(ci)
        w = fixed_widths.get(col_name, 14)
        ws.column_dimensions[ltr].width = w

    # ── AutoFilter ────────────────────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A2:{get_column_letter(n_cols)}{len(registry) + 2}"
    )


# ---------- PERFORMANCE SUMMARY ----------------------------------------------

def build_perf_summary(ws, perf: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["green"]
    ws.freeze_panes = "A3"

    if perf.empty:
        ws.cell(row=1, column=1, value="No data available.")
        return

    n_cols = len(perf.columns)

    # Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    banner = ws.cell(row=1, column=1,
                     value="  PERFORMANCE SUMMARY  —  Aggregated Metrics by Symbol")
    banner.font      = Font(name="Calibri", size=13, bold=True, color=C["gold"])
    banner.fill      = _fill(C["navy"])
    banner.alignment = _left()
    ws.row_dimensions[1].height = 28

    # Classify columns for header fill colour:
    #   registry (backtest) columns → midnight blue
    #   deployment (live trade) columns → slate blue  (visually distinct group)
    deploy_cols = {"Deployed", "Deployed Strats", "Live Trades",
                   "Dry-Run Trades", "Net P&L ($)"}

    for ci, col_name in enumerate(perf.columns, start=1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        hdr_fill = _fill(C["slate"]) if col_name in deploy_cols else _fill(C["midnight"])
        cell.font      = _font(bold=True, size=10, color=C["white"])
        cell.fill      = hdr_fill
        cell.alignment = _center(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[2].height = 32

    # Data rows
    for ri, (_, row) in enumerate(perf.iterrows(), start=1):
        r    = ri + 2
        is_deployed = str(row.get("Deployed", "no")) == "YES"
        # Deployed rows get a slightly brighter base fill to stand out
        if is_deployed:
            base_fill = _fill(C["charcoal"] if ri % 2 == 0 else "1C2E42")
        else:
            base_fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])

        for ci, col_name in enumerate(perf.columns, start=1):
            val  = row[col_name]
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill      = base_fill
            cell.border    = _border()
            cell.number_format = _num_fmt(col_name)

            # ── Registry / backtest columns ───────────────────────────────
            if col_name == "Symbol":
                cell.font      = _font(bold=True, size=11, color=C["gold"])
                cell.alignment = _center()
            elif col_name == "Strategy Count":
                cell.font      = _font(bold=True, size=10, color=C["teal"])
                cell.alignment = _center()
            elif "Sharpe" in col_name or "Calmar" in col_name:
                cell.font      = _font(size=10, color=C["teal"])
                cell.alignment = _center()
            elif "MDD" in col_name:
                cell.font      = _font(size=10, color=C["red"])
                cell.alignment = _center()

            # ── Deployment columns (trade-log based) ──────────────────────
            elif col_name == "Deployed":
                dep_color = C["green"] if is_deployed else C["dim"]
                dep_fill  = _fill(C["panel"]) if is_deployed else base_fill
                cell.fill      = dep_fill
                cell.font      = _font(bold=True, size=10, color=dep_color)
                cell.alignment = _center()

            elif col_name == "Deployed Strats":
                color = C["teal"] if is_deployed else C["dim"]
                cell.font          = _font(bold=True, size=10, color=color)
                cell.alignment     = _center()
                cell.number_format = "#,##0"

            elif col_name in ("Live Trades", "Dry-Run Trades"):
                color = C["teal"] if col_name == "Live Trades" else C["amber"]
                cell.font          = _font(size=10, color=color if (pd.notna(val) and val > 0) else C["dim"])
                cell.alignment     = _center()
                cell.number_format = "#,##0"

            elif col_name == "Net P&L ($)":
                color = C["green"] if (pd.notna(val) and val >= 0) else C["red"]
                if not is_deployed:
                    color = C["dim"]
                cell.font          = _font(bold=True, size=10, color=color)
                cell.alignment     = _center()
                cell.number_format = "#,##0.00"

            else:
                cell.font      = _font(size=10, color=C["light_gray"])
                cell.alignment = _center()

        ws.row_dimensions[r].height = 20

    # Column widths — fixed for legibility
    fixed = {
        "#": 4, "Symbol": 9, "Strategy Count": 10,
        "Avg Sharpe": 11, "Avg MDD": 10, "Avg Annual Return": 14,
        "Avg Calmar": 11, "Avg Cumulative PnL": 16, "Avg Trade Count": 13,
        "Avg PnL Ratio": 12, "Best Sharpe": 11, "Best Model": 22,
        "Best Entry/Exit": 22,
        "Deployed": 10, "Deployed Strats": 14,
        "Live Trades": 11, "Dry-Run Trades": 13, "Net P&L ($)": 13,
    }
    for ci, col_name in enumerate(perf.columns, start=1):
        ltr = get_column_letter(ci)
        ws.column_dimensions[ltr].width = fixed.get(col_name, 14)

    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{len(perf) + 2}"


# ---------- CHANGELOG --------------------------------------------------------

def build_changelog(ws, new_count: int, scan_files: list[str], total_before: int,
                    dupes_removed: int = 0):
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["amber"]

    # Banner
    ws.merge_cells("A1:F1")
    banner = ws.cell(row=1, column=1, value="  CHANGELOG  —  Import History")
    banner.font      = Font(name="Calibri", size=13, bold=True, color=C["gold"])
    banner.fill      = _fill(C["navy"])
    banner.alignment = _left()
    ws.row_dimensions[1].height = 28

    # Header
    headers = ["Timestamp", "New Strategies", "Total Strategies",
               "Files Scanned", "Source Files", "Notes", "Dupes Removed"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = _font(bold=True, size=10, color=C["white"])
        cell.fill      = _fill(C["midnight"])
        cell.alignment = _center(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[2].height = 28

    # Find the last occupied row in column A (to append below it)
    last_data_row = 2
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            last_data_row = r

    new_row = last_data_row + 1

    file_list = "; ".join(os.path.basename(f) for f in scan_files)
    notes = f"Appended {new_count} new strategies."
    if dupes_removed > 0:
        notes += f" Removed {dupes_removed} duplicate(s)."
    row_data = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        new_count,
        total_before + new_count,
        len(scan_files),
        file_list,
        notes,
        dupes_removed if dupes_removed > 0 else "",
    ]

    fill = _fill(C["charcoal"])
    for ci, val in enumerate(row_data, start=1):
        cell = ws.cell(row=new_row, column=ci, value=val)
        cell.fill      = fill
        cell.font      = _font(size=10, color=C["light_gray"])
        cell.alignment = _left() if ci in (5, 6) else _center()
        cell.border    = _border()
    ws.row_dimensions[new_row].height = 18

    # Colour "Dupes Removed" cell red if > 0
    dupes_cell = ws.cell(row=new_row, column=7)
    if dupes_removed > 0:
        dupes_cell.font = _font(bold=True, size=10, color=C["amber"])

    # Column widths
    widths = [20, 16, 16, 14, 60, 34, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ============================================================================
# STEP 5b – ACTIVE STRATEGY DATA LOADING
# ============================================================================

def load_trades(trades_dir: str) -> pd.DataFrame:
    """Load all trade CSVs from ibkr_deployment/trades/. Uses the snapshot
    file when available; falls back to scanning individual per-symbol CSVs."""
    snapshot = os.path.join(trades_dir, SNAPSHOT_FILE)
    if os.path.exists(snapshot):
        try:
            df = pd.read_csv(snapshot)
            print(f"[TRADES] Loaded {len(df)} trades from snapshot.")
        except Exception as e:
            print(f"[WARN] Could not read snapshot: {e}")
            df = pd.DataFrame()
    else:
        df = pd.DataFrame()

    # Also scan individual files to catch anything newer than snapshot
    per_symbol_frames = []
    for fpath in sorted(glob.glob(os.path.join(trades_dir, TRADES_GLOB))):
        try:
            tmp = pd.read_csv(fpath)
            if not tmp.empty:
                per_symbol_frames.append(tmp)
        except Exception as e:
            print(f"[WARN] Could not parse {fpath}: {e}")

    if per_symbol_frames:
        per_sym_df = pd.concat(per_symbol_frames, ignore_index=True)
        if df.empty:
            df = per_sym_df
        else:
            # Combine and deduplicate by execId
            combined = pd.concat([df, per_sym_df], ignore_index=True)
            if "execId" in combined.columns:
                df = combined.drop_duplicates(subset=["execId"]).reset_index(drop=True)
            else:
                df = combined.reset_index(drop=True)

    if df.empty:
        print("[TRADES] No trade data found.")
        return pd.DataFrame()

    # Normalise datetime column — multiple formats exist in the wild
    df["datetime"] = pd.to_datetime(df["datetime"], errors="coerce",
                                    infer_datetime_format=True, utc=False)
    n_bad_dt = df["datetime"].isna().sum()
    if n_bad_dt > 0:
        print(f"[WARN] {n_bad_dt} trade row(s) had unparseable datetime and will be excluded.")
    df = df.dropna(subset=["datetime"]).sort_values("datetime", ascending=True).reset_index(drop=True)

    # Classify trade mode: live vs dry-run
    df["trade_mode"] = df["execId"].apply(
        lambda x: "Dry Run" if str(x).startswith("dry_") else "Live"
    )
    print(f"[TRADES] {len(df)} unique trades "
          f"({(df['trade_mode']=='Live').sum()} live, "
          f"{(df['trade_mode']=='Dry Run').sum()} dry-run)")
    return df


def load_portfolio_state(state_dir: str) -> dict:
    """Load all positions_*.json and risk_*.json from portfolio state directory.

    Returns:
        {
            "TQQQ": {
                "positions": {strategy_name: {signal, avg_cost, qty, ...}},
                "risk":      {date, daily_loss, peak_equity, ...}
            }, ...
        }
    """
    state = {}
    if not os.path.exists(state_dir):
        return state

    for pos_file in sorted(glob.glob(os.path.join(state_dir, "positions_*.json"))):
        portfolio = os.path.basename(pos_file).replace("positions_", "").replace(".json", "")
        try:
            with open(pos_file, "r") as f:
                positions = json.load(f)
            state.setdefault(portfolio, {})["positions"] = positions
        except Exception as e:
            print(f"[WARN] Could not load {pos_file}: {e}")

    for risk_file in sorted(glob.glob(os.path.join(state_dir, "risk_*.json"))):
        portfolio = os.path.basename(risk_file).replace("risk_", "").replace(".json", "")
        try:
            with open(risk_file, "r") as f:
                risk = json.load(f)
            state.setdefault(portfolio, {})["risk"] = risk
        except Exception as e:
            print(f"[WARN] Could not load {risk_file}: {e}")

    print(f"[STATE] Loaded portfolio state for: {', '.join(state.keys()) or 'none'}")
    return state


def build_strategy_pnl(trades: pd.DataFrame) -> pd.DataFrame:
    """Aggregate trade-level records into per-strategy P&L summary rows.

    P&L is computed as: net cash = SUM(SLD qty*price) - SUM(BOT qty*price) - commissions
    This equals realized P&L for closed positions and net cash flow for open ones.
    """
    if trades.empty:
        return pd.DataFrame()

    # Exclude untagged / activity-monitor records
    tagged = trades[~trades["orderRef"].isin(["UNTAGGED", "ActivityMonitor",
                                               "", None])].copy()
    tagged = tagged.dropna(subset=["orderRef"])

    def _agg(grp: pd.DataFrame) -> pd.Series:
        bot  = grp[grp["side"] == "BOT"]
        sld  = grp[grp["side"] == "SLD"]
        comm = grp["commission"].abs().sum()

        # Realized P&L per symbol within this strategy using avg cost basis.
        # Matches Streamlit — open positions are excluded from P&L.
        net_pnl = sum(
            _avg_cost_pnl(sym_grp)
            for _, sym_grp in grp.groupby("symbol")
        )

        modes = grp["trade_mode"].unique()
        mode  = modes[0] if len(modes) == 1 else "Hybrid"

        return pd.Series({
            "Symbol":         grp["symbol"].iloc[0],
            "Total Trades":   len(grp),
            "BOT Trades":     len(bot),
            "SLD Trades":     len(sld),
            "BOT Value ($)":  round((bot["quantity"] * bot["price"]).sum(), 2),
            "SLD Value ($)":  round((sld["quantity"] * sld["price"]).sum(), 2),
            "Net P&L ($)":    net_pnl,
            "Commission ($)": round(comm, 2),
            "Trade Mode":     mode,
            "First Trade":    grp["datetime"].min(),
            "Last Trade":     grp["datetime"].max(),
        })

    pnl = tagged.groupby("orderRef").apply(_agg).reset_index()
    pnl = pnl.rename(columns={"orderRef": "Strategy"})
    pnl = pnl.sort_values("Last Trade", ascending=False).reset_index(drop=True)
    pnl.insert(0, "#", range(1, len(pnl) + 1))
    return pnl


def build_position_table(portfolio_state: dict, strategy_pnl: pd.DataFrame) -> pd.DataFrame:
    """Build a flat position table from portfolio JSON state, enriched with trade stats."""
    rows = []
    for portfolio, data in portfolio_state.items():
        positions = data.get("positions", {})
        risk      = data.get("risk", {})
        for strategy, pos in positions.items():
            signal   = pos.get("signal", 0)
            avg_cost = pos.get("avg_cost", 0.0)
            qty      = pos.get("qty", 0)

            direction = {1: "LONG", -1: "SHORT", 0: "FLAT"}.get(signal, "FLAT")
            est_value = round(abs(qty) * avg_cost, 2) if qty and avg_cost else 0.0

            # Pull trade stats for this strategy
            trade_row = {}
            if not strategy_pnl.empty and "Strategy" in strategy_pnl.columns:
                match = strategy_pnl[strategy_pnl["Strategy"] == strategy]
                if not match.empty:
                    trade_row = match.iloc[0].to_dict()

            rows.append({
                "Portfolio":      portfolio,
                "Strategy":       strategy,
                "Symbol":         trade_row.get("Symbol", portfolio),
                "Signal":         signal,
                "Direction":      direction,
                "Qty":            qty,
                "Avg Cost":       avg_cost,
                "Est. Value ($)": est_value,
                "Total Trades":   trade_row.get("Total Trades", 0),
                "BOT Trades":     trade_row.get("BOT Trades", 0),
                "SLD Trades":     trade_row.get("SLD Trades", 0),
                "Net P&L ($)":    trade_row.get("Net P&L ($)", 0.0),
                "Trade Mode":     trade_row.get("Trade Mode", "—"),
                "Last Trade":     trade_row.get("Last Trade", None),
                "Halted":         risk.get("halted", False),
            })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    # Sort: active positions first (signal != 0), then by portfolio, then strategy
    df["_active"] = df["Signal"].abs()
    df = df.sort_values(["_active", "Portfolio", "Net P&L ($)"],
                        ascending=[False, True, False]).drop(columns=["_active"])
    df.insert(0, "#", range(1, len(df) + 1))
    return df.reset_index(drop=True)


# ── Active Strategy sheet builder ────────────────────────────────────────────

def _draw_portfolio_risk_cards(ws, portfolio_state: dict, start_row: int) -> int:
    """Draw one risk card per portfolio. Returns the next free row."""
    if not portfolio_state:
        return start_row

    card_cols = 5   # columns per card
    gap       = 1   # gap column between cards

    portfolios = list(portfolio_state.keys())
    col_cursor = 1

    for portfolio in portfolios:
        risk      = portfolio_state[portfolio].get("risk", {})
        positions = portfolio_state[portfolio].get("positions", {})
        active_ct = sum(1 for p in positions.values() if p.get("signal", 0) != 0)
        total_ct  = len(positions)

        peak_eq  = risk.get("peak_equity", 0.0)
        daily_l  = risk.get("daily_loss", 0.0)
        cum_pnl  = risk.get("cumulative_realized_pnl", 0.0)
        halted   = risk.get("halted", False)
        halt_rsn = risk.get("halt_reason", "")
        date_str = risk.get("date", "—")

        halt_color = C["red"] if halted else C["green"]
        halt_label = f"HALTED: {halt_rsn}" if halted else "RUNNING"

        metrics = [
            ("Portfolio",    portfolio,           C["gold"]),
            ("Date",         date_str,            C["silver"]),
            ("Peak Equity",  f"${peak_eq:,.2f}",  C["teal"]),
            ("Daily Loss",   f"${daily_l:,.2f}",  C["red"] if daily_l < 0 else C["silver"]),
            ("Cum PnL",      f"${cum_pnl:,.2f}",  C["green"] if cum_pnl >= 0 else C["red"]),
            ("Status",       halt_label,          halt_color),
            (f"Active ({active_ct}/{total_ct})", f"{active_ct} positions", C["teal"]),
        ]

        # Card header (portfolio name)
        ws.merge_cells(start_row=start_row, start_column=col_cursor,
                       end_row=start_row, end_column=col_cursor + card_cols - 1)
        hdr = ws.cell(row=start_row, column=col_cursor, value=f"  {portfolio} PORTFOLIO")
        hdr.font      = Font(name="Calibri", size=12, bold=True, color=C["white"])
        hdr.fill      = _fill(C["midnight"])
        hdr.alignment = _left()
        ws.row_dimensions[start_row].height = 22

        for mi, (label, value, color) in enumerate(metrics):
            r_lbl = start_row + 1 + mi * 2
            r_val = r_lbl + 1

            ws.merge_cells(start_row=r_lbl, start_column=col_cursor,
                           end_row=r_lbl, end_column=col_cursor + card_cols - 1)
            lc = ws.cell(row=r_lbl, column=col_cursor, value=label)
            lc.font      = _font(bold=True, size=8, color=C["silver"])
            lc.fill      = _fill(C["panel"])
            lc.alignment = _center()
            ws.row_dimensions[r_lbl].height = 13

            ws.merge_cells(start_row=r_val, start_column=col_cursor,
                           end_row=r_val, end_column=col_cursor + card_cols - 1)
            vc = ws.cell(row=r_val, column=col_cursor, value=value)
            vc.font      = Font(name="Calibri", size=11, bold=True, color=color)
            vc.fill      = _fill(C["charcoal"])
            vc.alignment = _center()
            vc.border    = _border()
            ws.row_dimensions[r_val].height = 16

        col_cursor += card_cols + gap

    # Return row after all cards (start_row + 1 header + 7 metrics × 2 rows)
    return start_row + 1 + len(metrics) * 2 + 2


def _draw_position_table(ws, pos_df: pd.DataFrame, start_row: int) -> int:
    """Draw current positions table. Returns next free row."""
    if pos_df.empty:
        ws.cell(row=start_row, column=1, value="No position data available.")
        return start_row + 2

    n_cols = len(pos_df.columns)

    # Header row
    for ci, col_name in enumerate(pos_df.columns, start=1):
        cell = ws.cell(row=start_row, column=ci, value=col_name)
        cell.font      = _font(bold=True, size=10, color=C["white"])
        cell.fill      = _fill(C["midnight"])
        cell.alignment = _center(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[start_row].height = 30

    signal_colors = {1: C["green"], -1: C["red"], 0: C["dim"]}
    direction_colors = {"LONG": C["green"], "SHORT": C["red"], "FLAT": C["dim"]}

    for ri, (_, row) in enumerate(pos_df.iterrows(), start=1):
        r = start_row + ri
        base_fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])
        signal_val = row.get("Signal", 0)

        for ci, col_name in enumerate(pos_df.columns, start=1):
            val  = row[col_name]
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = _border()

            if col_name == "#":
                cell.fill      = _fill(C["slate"])
                cell.font      = _font(bold=True, size=9, color=C["gold"])
                cell.alignment = _center()

            elif col_name == "Direction":
                color = direction_colors.get(str(val), C["silver"])
                cell.fill      = _fill(C["panel"])
                cell.font      = _font(bold=True, size=9, color=color)
                cell.alignment = _center()

            elif col_name == "Signal":
                color = signal_colors.get(int(val) if pd.notna(val) else 0, C["silver"])
                cell.fill      = base_fill
                cell.font      = _font(bold=True, size=10, color=color)
                cell.alignment = _center()

            elif col_name == "Halted":
                cell.fill      = _fill(C["red"] if val else C["panel"])
                cell.font      = _font(bold=True, size=9,
                                       color=C["white"] if val else C["dim"])
                cell.value     = "YES" if val else "no"
                cell.alignment = _center()

            elif col_name == "Trade Mode":
                color = C["teal"] if str(val) == "Live" else \
                        C["amber"] if str(val) == "Dry Run" else C["silver"]
                cell.fill      = base_fill
                cell.font      = _font(bold=True, size=9, color=color)
                cell.alignment = _center()

            elif col_name == "Net P&L ($)":
                color = C["green"] if (pd.notna(val) and val >= 0) else C["red"]
                cell.fill      = base_fill
                cell.font      = _font(bold=True, size=10, color=color)
                cell.alignment = _center()
                cell.number_format = '#,##0.00'

            elif col_name in ("Avg Cost", "Est. Value ($)", "BOT Value ($)",
                               "SLD Value ($)", "Commission ($)"):
                cell.fill      = base_fill
                cell.font      = _font(size=10, color=C["light_gray"])
                cell.alignment = _center()
                cell.number_format = '#,##0.00'

            elif col_name in ("First Trade", "Last Trade"):
                _fmt_ts(cell, val)
                cell.fill      = base_fill
                cell.font      = _font(size=9, color=C["silver"])
                cell.alignment = _center()

            elif col_name == "Portfolio":
                cell.fill      = _fill(C["slate"])
                cell.font      = _font(bold=True, size=9, color=C["gold"])
                cell.alignment = _center()

            elif col_name == "Strategy":
                cell.fill      = base_fill
                cell.font      = _font(size=9, color=C["light_gray"])
                cell.alignment = _left(wrap=False)

            else:
                cell.fill      = base_fill
                cell.font      = _font(size=10, color=C["silver"])
                cell.alignment = _center()

        ws.row_dimensions[r].height = 18

    return start_row + len(pos_df) + 2


def _draw_trade_log(ws, trades: pd.DataFrame, start_row: int,
                    max_rows: int = 60) -> int:
    """Draw the recent trade log table. Returns next free row."""
    if trades.empty:
        ws.cell(row=start_row, column=1, value="No trade data available.")
        return start_row + 2

    # Show most-recent trades first, capped at max_rows
    display = trades.sort_values("datetime", ascending=False).head(max_rows).copy()

    cols = ["datetime", "symbol", "orderRef", "side", "quantity",
            "price", "commission", "trade_mode", "execId"]
    cols = [c for c in cols if c in display.columns]
    display = display[cols].copy()
    display.columns = ["DateTime", "Symbol", "Strategy (orderRef)", "Side",
                       "Qty", "Price", "Commission", "Trade Mode",
                       "execId"][:len(cols)]

    display.insert(0, "#", range(1, len(display) + 1))
    n_cols = len(display.columns)

    side_color  = {"BOT": C["green"], "SLD": C["red"]}
    mode_color  = {"Live": C["teal"], "Dry Run": C["amber"], "Hybrid": C["silver"]}

    # Header
    for ci, col_name in enumerate(display.columns, start=1):
        cell = ws.cell(row=start_row, column=ci, value=col_name)
        cell.font      = _font(bold=True, size=10, color=C["white"])
        cell.fill      = _fill(C["midnight"])
        cell.alignment = _center(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[start_row].height = 28

    for ri, (_, row) in enumerate(display.iterrows(), start=1):
        r = start_row + ri
        base_fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])

        for ci, col_name in enumerate(display.columns, start=1):
            val  = row[col_name]
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = _border()

            if col_name == "#":
                cell.fill = _fill(C["slate"])
                cell.font = _font(bold=True, size=9, color=C["gold"])
                cell.alignment = _center()

            elif col_name == "Side":
                color = side_color.get(str(val), C["silver"])
                cell.fill      = base_fill
                cell.font      = _font(bold=True, size=10, color=color)
                cell.alignment = _center()

            elif col_name == "Trade Mode":
                color = mode_color.get(str(val), C["silver"])
                cell.fill      = base_fill
                cell.font      = _font(bold=True, size=9, color=color)
                cell.alignment = _center()

            elif col_name in ("Price", "Commission"):
                cell.fill          = base_fill
                cell.font          = _font(size=10, color=C["light_gray"])
                cell.alignment     = _center()
                cell.number_format = '#,##0.00'

            elif col_name == "DateTime":
                _fmt_ts(cell, val)
                cell.fill      = base_fill
                cell.font      = _font(size=9, color=C["silver"])
                cell.alignment = _center()

            elif col_name == "Strategy (orderRef)":
                cell.fill      = base_fill
                cell.font      = _font(size=9, color=C["light_gray"])
                cell.alignment = _left()

            elif col_name == "execId":
                cell.fill      = base_fill
                cell.font      = _font(size=8, color=C["dim"])
                cell.alignment = _left()

            else:
                cell.fill      = base_fill
                cell.font      = _font(size=10, color=C["silver"])
                cell.alignment = _center()

        ws.row_dimensions[r].height = 17

    return start_row + len(display) + 2


def build_active_strategy(ws, trades: pd.DataFrame, portfolio_state: dict):
    """Build the Active Strategy worksheet.

    Sections:
      A – Banner
      B – Portfolio Risk Cards (one card per portfolio in state/)
      C – Current Positions Table
      D – Strategy Trade P&L Summary
      E – Recent Trade Log (last 60 trades)
    """
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["green"]
    ws.freeze_panes = "A3"

    # ── A: Banner ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:V1")
    banner = ws.cell(row=1, column=1,
                     value="  ACTIVE STRATEGY MONITOR  —  Live Deployed Strategy Performance")
    banner.font      = Font(name="Calibri", size=14, bold=True, color=C["gold"])
    banner.fill      = _fill(C["navy"])
    banner.alignment = _left()
    ws.row_dimensions[1].height = 30

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    n_live  = (trades["trade_mode"] == "Live").sum() if not trades.empty else 0
    n_dry   = (trades["trade_mode"] == "Dry Run").sum() if not trades.empty else 0
    ws.merge_cells("A2:V2")
    sub = ws.cell(row=2, column=1,
                  value=f"  Last updated: {ts}   |   "
                        f"Total Trades: {len(trades)}   |   "
                        f"Live: {n_live}   |   Dry-Run: {n_dry}")
    sub.font      = _font(size=10, color=C["silver"])
    sub.fill      = _fill(C["midnight"])
    sub.alignment = _left()
    ws.row_dimensions[2].height = 18

    next_row = 4

    # ── B: Portfolio Risk Cards ────────────────────────────────────────────────
    if portfolio_state:
        _draw_section_header(ws, row=next_row, col=1, width=22,
                             text="PORTFOLIO STATUS")
        next_row += 1
        next_row = _draw_portfolio_risk_cards(ws, portfolio_state, next_row)
    else:
        ws.cell(row=next_row, column=1,
                value="  [INFO] No portfolio state files found.").font = _font(
                    size=9, color=C["dim"])
        next_row += 2

    # ── C: Current Positions ──────────────────────────────────────────────────
    strategy_pnl = build_strategy_pnl(trades)
    pos_df       = build_position_table(portfolio_state, strategy_pnl)

    _draw_section_header(ws, row=next_row, col=1, width=22,
                         text="CURRENT POSITIONS")
    next_row += 1
    next_row = _draw_position_table(ws, pos_df, next_row)

    # ── D: Strategy Trade P&L Summary ─────────────────────────────────────────
    _draw_section_header(ws, row=next_row, col=1, width=22,
                         text="STRATEGY TRADE P&L SUMMARY")
    next_row += 1

    if not strategy_pnl.empty:
        n_cols_pnl = len(strategy_pnl.columns)
        # Header
        for ci, col_name in enumerate(strategy_pnl.columns, start=1):
            cell = ws.cell(row=next_row, column=ci, value=col_name)
            cell.font      = _font(bold=True, size=10, color=C["white"])
            cell.fill      = _fill(C["slate"])
            cell.alignment = _center(wrap=True)
            cell.border    = _border()
        ws.row_dimensions[next_row].height = 28
        next_row += 1

        trade_mode_colors = {"Live": C["teal"], "Dry Run": C["amber"]}
        for ri, (_, row) in enumerate(strategy_pnl.iterrows(), start=1):
            r = next_row
            base_fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])
            for ci, col_name in enumerate(strategy_pnl.columns, start=1):
                val  = row[col_name]
                cell = ws.cell(row=r, column=ci, value=val)
                cell.border = _border()
                if col_name == "#":
                    cell.fill = _fill(C["slate"])
                    cell.font = _font(bold=True, size=9, color=C["gold"])
                    cell.alignment = _center()
                elif col_name == "Net P&L ($)":
                    color = C["green"] if (pd.notna(val) and val >= 0) else C["red"]
                    cell.fill = base_fill
                    cell.font = _font(bold=True, size=10, color=color)
                    cell.alignment = _center()
                    cell.number_format = '#,##0.00'
                elif col_name == "Trade Mode":
                    c = trade_mode_colors.get(str(val), C["silver"])
                    cell.fill = base_fill
                    cell.font = _font(bold=True, size=9, color=c)
                    cell.alignment = _center()
                elif col_name in ("BOT Value ($)", "SLD Value ($)", "Commission ($)"):
                    cell.fill = base_fill
                    cell.font = _font(size=10, color=C["light_gray"])
                    cell.alignment = _center()
                    cell.number_format = '#,##0.00'
                elif col_name in ("First Trade", "Last Trade"):
                    _fmt_ts(cell, val)
                    cell.fill = base_fill
                    cell.font = _font(size=9, color=C["silver"])
                    cell.alignment = _center()
                elif col_name == "Strategy":
                    cell.fill = base_fill
                    cell.font = _font(size=9, color=C["light_gray"])
                    cell.alignment = _left()
                else:
                    cell.fill = base_fill
                    cell.font = _font(size=10, color=C["silver"])
                    cell.alignment = _center()
            ws.row_dimensions[r].height = 18
            next_row += 1

        next_row += 1  # gap

    # ── E: Recent Trade Log ────────────────────────────────────────────────────
    _draw_section_header(ws, row=next_row, col=1, width=22,
                         text="RECENT TRADES  (last 60)")
    next_row += 1
    next_row = _draw_trade_log(ws, trades, next_row, max_rows=60)

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = {
        1: 5, 2: 13, 3: 38, 4: 9, 5: 8, 6: 8, 7: 10, 8: 12,
        9: 11, 10: 10, 11: 10, 12: 13, 13: 11, 14: 19, 15: 19,
        16: 14, 17: 14, 18: 12, 19: 10, 20: 20, 21: 20, 22: 36,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ============================================================================
# STEP 5c – BACKTEST vs EXECUTION COMPARISON  (three-way)
# ============================================================================
#
# Methodology (documented for auditability)
# ------------------------------------------
# Comparison shows THREE execution contexts side-by-side per deployed strategy:
#   1. BACKTEST  — metrics from Risk_Optimized_Compilation via Strategy Registry
#   2. PAPER     — metrics computed from Dry Run (paper account) trade records
#   3. LIVE      — metrics computed from Live account trade records
#
# Backtest source    : per-symbol averages of Sharpe / MDD / Annual Return
#                      from Strategy Registry (compiled from Risk_Optimized_Compilation).
#                      Symbols are case-normalised (upper) before joining.
#                      If only one strategy exists for a symbol in the registry,
#                      the value is exact (not averaged).
#
# Capital base proxy : notional of the FIRST BOT trade across ALL trade modes
#                      for a given strategy (qty × price).  For short-only
#                      strategies the first SLD trade is used.  The same
#                      capital base is applied to BOTH paper and live metric
#                      calculations to ensure comparability.
#
# Daily P&L series   : same average-cost-basis method as _avg_cost_pnl, but
#                      realised P&L is emitted trade-by-trade then summed by
#                      calendar date.  Non-trading calendar days = 0 P&L.
#
# Sharpe (annualised): mean(daily_returns) / std(daily_returns, ddof=1) × √252
#                      daily_returns = daily_realised_pnl / capital_base
#                      Risk-free rate = 0.  Requires ≥ 2 calendar days.
#
# MDD                : min( (equity − running_peak) / capital_base )
#                      equity = capital_base + cumsum(daily_pnl)
#                      Starting capital is prepended so day-one losses
#                      are captured.  Stored as a negative decimal.
#                      Denominator is capital_base (not running_peak) matching
#                      the backtesting convention in util.py.
#
# CAGR               : mean(daily_realised_pnl / capital_base) × 252
#                      Arithmetic annualisation matching backtesting formula
#                      mean(bar_pnl) × period in util.py.
#                      Requires ≥ 2 calendar days.
#
# Alpha decay        : Δ = Live/Paper metric − Backtest metric.
#                      Positive Δ Sharpe/CAGR = execution exceeds backtest.
#                      Positive Δ MDD = execution has smaller drawdown (better).
#
# ⚠  Execution metrics are approximations.  The capital base (first-trade
#    notional) may not match the actual capital allocated by the portfolio
#    manager.  Always review the "Capital Base ($)" column before drawing
#    conclusions.
# ============================================================================

def _build_daily_pnl_series(grp: pd.DataFrame) -> pd.Series:
    """Decompose a strategy's trade group into daily realised P&L increments.

    Mirrors the average-cost-basis logic of _avg_cost_pnl exactly, but emits
    the incremental realised P&L for *each trade* so it can be summed by date.

    Returns a pd.Series indexed by datetime.date with daily realised P&L.
    Missing calendar days are NOT filled here; call .reindex() on the caller
    side to expand to a full date range.
    """
    position   = 0.0
    avg_cost   = 0.0
    total_cost = 0.0
    records    = []          # list of (date, trade_pnl)

    for _, row in grp.sort_values("datetime").iterrows():
        qty   = abs(float(row.get("quantity",   0)))
        price =     float(row.get("price",      0))
        comm  = abs(float(row.get("commission", 0)))
        side  =  str(row.get("side", ""))
        date  = pd.Timestamp(row["datetime"]).date()

        trade_pnl = 0.0

        if side == "BOT":
            if position < 0:                              # covering a short
                close_qty  = min(qty, abs(position))
                close_comm = comm * (close_qty / qty) if qty > 0 else comm
                trade_pnl += close_qty * (avg_cost - price) - close_comm
                position  += close_qty
                remaining  = qty - close_qty
                total_cost = abs(position) * avg_cost if position != 0 else 0.0
                avg_cost   = 0.0 if position == 0 else avg_cost
                if remaining > 0:                         # crossed zero: new long
                    open_comm  = comm - close_comm
                    total_cost = remaining * price + open_comm
                    position   = remaining
                    avg_cost   = total_cost / position
            else:                                         # adding to / opening long
                total_cost += qty * price
                position   += qty
                avg_cost    = total_cost / position if position > 0 else 0.0
                trade_pnl  -= comm                        # opening commission is a cost

        elif side == "SLD":
            if position > 0:                              # closing / reducing long
                close_qty  = min(qty, position)
                close_comm = comm * (close_qty / qty) if qty > 0 else comm
                trade_pnl += close_qty * (price - avg_cost) - close_comm
                position  -= close_qty
                remaining  = qty - close_qty
                total_cost = position * avg_cost if position != 0 else 0.0
                avg_cost   = 0.0 if position == 0 else avg_cost
                if remaining > 0:                         # crossed zero: new short
                    open_comm  = comm - close_comm
                    total_cost = remaining * price + open_comm
                    position   = -remaining
                    avg_cost   = total_cost / remaining
            else:                                         # opening / adding to short
                trade_pnl  -= comm
                position   -= qty
                total_cost += qty * price
                avg_cost    = abs(total_cost / position) if position != 0 else 0.0

        records.append((date, trade_pnl))

    if not records:
        return pd.Series(dtype=float)

    df_rec = pd.DataFrame(records, columns=["date", "pnl"])
    return df_rec.groupby("date")["pnl"].sum()


def _compute_metrics_from_trade_group(
    grp: pd.DataFrame,
    capital_base: float,
    label: str = "",
) -> tuple:
    """Compute (sharpe, mdd, cagr, net_pnl) for a single trade sub-group.

    Parameters
    ----------
    grp          : Trade rows for ONE strategy and ONE execution mode.
    capital_base : First-trade notional shared across paper and live modes
                   so that metrics are comparable.
    label        : Strategy name for WARN messages (cosmetic only).

    Returns
    -------
    (sharpe, mdd, cagr, net_pnl) — floats.

    Data requirements by metric
    ---------------------------
    Net P&L : any number of trades (always computed)
    MDD     : ≥ 1 calendar day  (equity curve vs starting capital)
    CAGR    : ≥ 2 calendar days  (needs a meaningful time span)
    Sharpe  : ≥ 2 calendar days AND non-zero daily return variance
    """
    if grp.empty:
        return np.nan, np.nan, np.nan, 0.0

    if capital_base is None or capital_base <= 0:
        net_pnl = round(float(_avg_cost_pnl(grp)), 2)
        return np.nan, np.nan, np.nan, net_pnl

    try:
        daily_pnl_raw = _build_daily_pnl_series(grp)
    except Exception as e:
        print(f"  [WARN] P&L series error for {label!r}: {e}")
        return np.nan, np.nan, np.nan, 0.0

    net_pnl = round(float(daily_pnl_raw.sum()), 2) if not daily_pnl_raw.empty else 0.0

    if daily_pnl_raw.empty:
        return np.nan, np.nan, np.nan, net_pnl

    # Expand to full calendar date range so non-trading days contribute 0 P&L.
    # This is applied before ALL metric calculations for consistency.
    if len(daily_pnl_raw) > 1:
        date_range = pd.date_range(
            start=pd.Timestamp(daily_pnl_raw.index[0]),
            end=pd.Timestamp(daily_pnl_raw.index[-1]),
            freq="D",
        )
        daily_pnl = daily_pnl_raw.reindex(
            [d.date() for d in date_range], fill_value=0.0
        )
    else:
        daily_pnl = daily_pnl_raw   # single trading day — kept as-is

    n_days = len(daily_pnl)

    def _safe(v, ndigits):
        """Round v; return NaN for NaN/inf."""
        try:
            return round(float(v), ndigits) if np.isfinite(float(v)) else np.nan
        except (TypeError, ValueError):
            return np.nan

    try:
        # ── MDD: always compute (≥ 1 day) ─────────────────────────────────────
        # Prepend the starting capital so a loss on the very first day is captured.
        equity       = capital_base + daily_pnl.cumsum()
        full_equity  = pd.concat(
            [pd.Series([capital_base]), equity]
        ).reset_index(drop=True)
        running_peak = full_equity.cummax()
        # Denominator is capital_base (not running_peak) to match the backtesting
        # convention in util.py: dd = cumulative_pnl - cumulative_pnl.cummax()
        drawdowns    = (full_equity - running_peak) / capital_base
        mdd          = _safe(drawdowns.min(), 6)   # negative decimal, e.g. −0.12

        # ── Sharpe and CAGR: require ≥ 2 calendar days ────────────────────────
        if n_days >= 2:
            daily_ret = daily_pnl / capital_base
            ret_std   = float(daily_ret.std(ddof=1))
            sharpe    = _safe(
                (float(daily_ret.mean()) / ret_std) * np.sqrt(252)
                if ret_std > 1e-10 else np.nan,
                4
            )
            # Arithmetic annualisation: mean(daily_return) × 252
            # Matches backtesting formula: mean(bar_pnl) × period in util.py
            cagr = _safe(float(daily_ret.mean()) * 252, 6)
        else:
            # Single trading day: Sharpe and CAGR are not meaningful
            sharpe = np.nan
            cagr   = np.nan

        return sharpe, mdd, cagr, net_pnl

    except Exception as e:
        print(f"  [WARN] Metric computation error for {label!r}: {e}")
        return np.nan, np.nan, np.nan, net_pnl


def build_backtest_vs_execution_data(
    registry: pd.DataFrame,
    trades:   pd.DataFrame,
) -> pd.DataFrame:
    """Build the three-way Backtest / Paper / Live comparison DataFrame.

    One row per deployed strategy (orderRef).  Columns are grouped as:
      Identity  : #, Strategy, Symbol, Trade Mode, dates, days, capital base
      Backtest  : BT Sharpe, BT MDD, BT CAGR  (strategy's own backtest metrics)
      Paper     : Paper Sharpe, Paper MDD, Paper CAGR, Paper P&L ($)
      Live      : Live Sharpe, Live MDD, Live CAGR, Live P&L ($)
      Δ Decay   : Δ Sharpe/MDD/CAGR for Paper−BT and Live−BT

    Backtest metrics are matched to the EXACT deployed strategy by constructing
    a registry key "{Symbol}_{DataPoint}_{Model}_{Entry/Exit Model}" and checking
    whether the orderRef starts with that key.  Longest key wins to avoid
    prefix collisions.  Falls back to NaN if no match is found.
    """
    if trades.empty:
        return pd.DataFrame()

    tagged = (trades[~trades["orderRef"].isin(EXCLUDED_ORDER_REFS)]
              .dropna(subset=["orderRef", "datetime"]))
    if tagged.empty:
        return pd.DataFrame()

    # ── Build backtest lookup: strategy key → {BT Sharpe, BT MDD, BT CAGR} ──
    # Key format: "{Symbol}_{DataPoint}_{Model}_{Entry/Exit Model}"
    # This matches the leading segment of the orderRef (before username/timestamp).
    # Keys are sorted longest-first so that a more-specific key always wins over
    # a shorter prefix that would also satisfy startswith().
    bt_lookup: dict = {}   # key → {BT Sharpe, BT MDD, BT CAGR}
    _KEY_PARTS = ["Symbol", "Interval", "Data Point", "Model", "Entry / Exit Model"]
    _METRIC_MAP = {
        "BT Sharpe": "Sharpe",
        "BT MDD":    "MDD",
        "BT CAGR":   "Annual Return",   # Backtest Annual Return = mean(bar_pnl) × period (arithmetic)
    }
    if not registry.empty:
        available_parts = [c for c in _KEY_PARTS if c in registry.columns]
        if available_parts:
            for _, row in registry.iterrows():
                reg_key = "_".join(
                    str(row[c]).strip() for c in available_parts
                )
                entry = {}
                for out_col, src_col in _METRIC_MAP.items():
                    v = row.get(src_col, np.nan)
                    entry[out_col] = float(v) if pd.notna(v) else np.nan
                bt_lookup[reg_key] = entry
            # Sort keys longest-first for prefix matching (avoids short-key false hits)
            bt_lookup = dict(
                sorted(bt_lookup.items(), key=lambda kv: len(kv[0]), reverse=True)
            )
            n_found = sum(
                1 for v in bt_lookup.values()
                if not np.isnan(v.get("BT Sharpe", np.nan))
            )
            print(f"[COMPARE] Backtest lookup built: {len(bt_lookup)} strategy key(s), "
                  f"{n_found} with Sharpe data.")
        else:
            print("[COMPARE] Registry missing key columns — backtest lookup unavailable.")
    else:
        print("[COMPARE] Registry is empty — backtest columns will be N/A.")

    def _bt_match(order_ref: str) -> dict:
        """Return BT metrics for the registry entry whose key is a prefix of order_ref."""
        # Pass 1: direct prefix match (works when registry has Interval column)
        for reg_key, entry in bt_lookup.items():
            if order_ref.startswith(reg_key + "_") or order_ref == reg_key:
                return entry
        # Pass 2: strip 2nd token (Interval, e.g. "1h") from orderRef and retry.
        # Handles Stage-8 registries (Risk_Optimized_Compilation) that omit Interval.
        tokens = order_ref.split("_")
        if len(tokens) >= 3:
            stripped = tokens[0] + "_" + "_".join(tokens[2:])
            for reg_key, entry in bt_lookup.items():
                if stripped.startswith(reg_key + "_") or stripped == reg_key:
                    return entry
        return {}

    results = []

    for order_ref, grp in tagged.groupby("orderRef"):
        grp      = grp.dropna(subset=["datetime"]).sort_values("datetime").copy()
        symbol   = str(grp["symbol"].iloc[0]).strip().upper()
        modes    = set(grp["trade_mode"].dropna().unique())
        first_dt = grp["datetime"].min()
        last_dt  = grp["datetime"].max()
        days_traded = (
            int((pd.Timestamp(last_dt) - pd.Timestamp(first_dt)).days + 1)
            if pd.notna(first_dt) and pd.notna(last_dt) else 0
        )

        # Display mode: Live / Dry Run / Hybrid
        if len(modes) > 1:
            display_mode = "Hybrid"
        elif modes:
            display_mode = modes.pop()
        else:
            display_mode = "—"

        # ── Capital base: first BOT trade across ALL modes (shared denominator) ─
        bot_trades = grp[grp["side"] == "BOT"]
        if not bot_trades.empty:
            fr = bot_trades.iloc[0]
        else:
            sld_trades = grp[grp["side"] == "SLD"]
            fr = sld_trades.iloc[0] if not sld_trades.empty else None

        if fr is not None:
            capital_base = abs(float(fr["quantity"])) * float(fr["price"])
            if capital_base <= 0:
                capital_base = None
        else:
            capital_base = None

        # ── Backtest metrics: exact match to this strategy's registry entry ──────
        bt        = _bt_match(str(order_ref))
        bt_sharpe = bt.get("BT Sharpe", np.nan)
        bt_mdd    = bt.get("BT MDD",    np.nan)
        bt_cagr   = bt.get("BT CAGR",   np.nan)

        # ── Paper (dry-run) metrics ───────────────────────────────────────────
        paper_grp = grp[grp["trade_mode"] == "Dry Run"].copy()
        p_n_trades = len(paper_grp)
        if not paper_grp.empty:
            p_sh, p_mdd, p_cagr, p_pnl = _compute_metrics_from_trade_group(
                paper_grp, capital_base, label=f"{order_ref}[paper]"
            )
        else:
            p_sh = p_mdd = p_cagr = np.nan
            p_pnl = np.nan      # NaN signals "no paper trades", not zero profit

        # ── Live metrics ─────────────────────────────────────────────────────
        live_grp = grp[grp["trade_mode"] == "Live"].copy()
        l_n_trades = len(live_grp)
        if not live_grp.empty:
            l_sh, l_mdd, l_cagr, l_pnl = _compute_metrics_from_trade_group(
                live_grp, capital_base, label=f"{order_ref}[live]"
            )
        else:
            l_sh = l_mdd = l_cagr = np.nan
            l_pnl = np.nan      # NaN signals "no live trades", not zero profit

        # ── Alpha decay deltas ────────────────────────────────────────────────
        def _delta(a, b):
            if pd.notna(a) and pd.notna(b):
                return round(float(a) - float(b), 6)
            return np.nan

        results.append({
            "orderRef":           order_ref,
            "Symbol":             symbol,
            "Trade Mode":         display_mode,
            "First Trade":        first_dt,
            "Last Trade":         last_dt,
            "Days Traded":        days_traded,
            "Capital Base ($)":   round(capital_base, 2) if capital_base else np.nan,
            # ── Backtest (from Risk_Optimized_Compilation, strategy-exact match) ─
            "BT Sharpe":          round(bt_sharpe, 4) if pd.notna(bt_sharpe) else np.nan,
            "BT MDD":             round(bt_mdd,    6) if pd.notna(bt_mdd)    else np.nan,
            "BT CAGR":            round(bt_cagr,   6) if pd.notna(bt_cagr)   else np.nan,
            # ── Paper / dry-run execution ──────────────────────────────────────
            "Paper Trades":       p_n_trades if p_n_trades > 0 else np.nan,
            "Paper Sharpe":       p_sh,
            "Paper MDD":          p_mdd,
            "Paper CAGR":         p_cagr,
            "Paper P&L ($)":      p_pnl,
            # ── Live execution ─────────────────────────────────────────────────
            "Live Trades":        l_n_trades if l_n_trades > 0 else np.nan,
            "Live Sharpe":        l_sh,
            "Live MDD":           l_mdd,
            "Live CAGR":          l_cagr,
            "Live P&L ($)":       l_pnl,
            # ── Alpha decay: Paper vs Backtest ─────────────────────────────────
            "Δ Sharpe (P−BT)":    _delta(p_sh,   bt_sharpe),
            "Δ MDD (P−BT)":       _delta(p_mdd,  bt_mdd),
            "Δ CAGR (P−BT)":      _delta(p_cagr, bt_cagr),
            # ── Alpha decay: Live vs Backtest ──────────────────────────────────
            "Δ Sharpe (L−BT)":    _delta(l_sh,   bt_sharpe),
            "Δ MDD (L−BT)":       _delta(l_mdd,  bt_mdd),
            "Δ CAGR (L−BT)":      _delta(l_cagr, bt_cagr),
        })

    if not results:
        return pd.DataFrame()

    comp = pd.DataFrame(results)
    comp["Notes"] = comp.apply(_comparison_notes, axis=1)

    col_order = [
        "orderRef", "Symbol", "Trade Mode",
        "First Trade", "Last Trade", "Days Traded", "Capital Base ($)",
        "BT Sharpe",    "BT MDD",    "BT CAGR",
        "Paper Trades", "Paper Sharpe", "Paper MDD", "Paper CAGR", "Paper P&L ($)",
        "Live Trades",  "Live Sharpe",  "Live MDD",  "Live CAGR",  "Live P&L ($)",
        "Δ Sharpe (P−BT)", "Δ MDD (P−BT)", "Δ CAGR (P−BT)",
        "Δ Sharpe (L−BT)", "Δ MDD (L−BT)", "Δ CAGR (L−BT)",
        "Notes",
    ]
    col_order = [c for c in col_order if c in comp.columns]
    comp = (comp[col_order]
            .rename(columns={"orderRef": "Strategy"})
            .sort_values("Symbol")
            .reset_index(drop=True))
    comp.insert(0, "#", range(1, len(comp) + 1))
    return comp


def _comparison_notes(row: pd.Series) -> str:
    """Generate a concise alpha-decay quality note for a three-way comparison row."""
    notes = []
    days = row.get("Days Traded", 0)

    if pd.isna(row.get("Capital Base ($)")):
        return "Capital base undetermined — % metrics unavailable"

    # Use trade count as the primary "has trades" signal so that strategies
    # with fills but insufficient history for Sharpe/CAGR are not misreported.
    p_n = row.get("Paper Trades", np.nan)
    l_n = row.get("Live Trades",  np.nan)
    has_paper = pd.notna(p_n) and int(p_n) > 0
    has_live  = pd.notna(l_n) and int(l_n) > 0
    has_bt    = pd.notna(row.get("BT Sharpe"))

    if not has_paper and not has_live:
        return "No trades executed yet"

    if pd.notna(days) and int(days) < 30:
        notes.append(f"Short history ({int(days)}d)")

    # Flag small sample sizes — metrics are unreliable with few fills
    if pd.notna(p_n) and int(p_n) < 10:
        notes.append(f"Paper sample small (N={int(p_n)} fills)")
    if pd.notna(l_n) and int(l_n) < 10:
        notes.append(f"Live sample small (N={int(l_n)} fills)")

    if not has_bt:
        notes.append("No backtest data matched — verify strategy key in registry")
        return "; ".join(notes) if notes else "No backtest data matched"

    # Check alpha decay relative to backtest for both live and paper modes.
    # For all three metrics, a HIGHER numeric value is BETTER:
    #   Sharpe : larger positive = better risk-adjusted return
    #   MDD    : less negative (e.g. -0.05 > -0.20) = smaller drawdown = better
    #   CAGR   : higher = better annualised return
    _bt_col_map = {"Sharpe": "BT Sharpe", "MDD": "BT MDD", "CAGR": "BT CAGR"}
    for mode_label, sh_col, mdd_col, cagr_col in [
        ("Paper", "Paper Sharpe", "Paper MDD", "Paper CAGR"),
        ("Live",  "Live Sharpe",  "Live MDD",  "Live CAGR"),
    ]:
        for metric, col in [("Sharpe", sh_col), ("MDD", mdd_col), ("CAGR", cagr_col)]:
            bt_col = _bt_col_map[metric]
            lv = row.get(col)
            bt = row.get(bt_col)
            if pd.notna(lv) and pd.notna(bt) and abs(bt) > 1e-10:
                pct = (float(lv) - float(bt)) / abs(float(bt)) * 100
                if abs(pct) > 20:
                    qual = "better" if pct > 0 else "ALPHA DECAY"
                    notes.append(
                        f"{mode_label} {metric} {abs(pct):.0f}% "
                        f"{'above' if pct > 0 else 'below'} BT ({qual})"
                    )

    if not notes:
        notes.append("All modes within 20% of backtest")
    return "; ".join(notes)


# ---------- COMPARISON SHEET -------------------------------------------------
#
# Column groups and their Excel column positions (1-based):
#   Group 1  IDENTITY  (8 cols) : #, Strategy, Symbol, Trade Mode,
#                                   First Trade, Last Trade, Days Traded, Capital Base ($)
#   Group 2  BACKTEST  (3 cols) : BT Sharpe, BT MDD, BT CAGR
#   Group 3  PAPER     (4 cols) : Paper Sharpe, Paper MDD, Paper CAGR, Paper P&L ($)
#   Group 4  LIVE      (4 cols) : Live Sharpe, Live MDD, Live CAGR, Live P&L ($)
#   Group 5  Δ PAPER−BT(3 cols) : Δ Sharpe (P−BT), Δ MDD (P−BT), Δ CAGR (P−BT)
#   Group 6  Δ LIVE−BT (3 cols) : Δ Sharpe (L−BT), Δ MDD (L−BT), Δ CAGR (L−BT)
#   Group 7  NOTES     (1 col)  : Notes
#   Total: 26 columns

_COMP_GROUPS = [
    # (group_label, fill_hex, col_names)
    ("IDENTITY",
     "0D1B2A",
     ["#", "Strategy", "Symbol", "Trade Mode",
      "First Trade", "Last Trade", "Days Traded", "Capital Base ($)"]),
    ("BACKTEST  (Risk-Optimized Compilation — strategy-exact match)",
     "1A2744",
     ["BT Sharpe", "BT MDD", "BT CAGR"]),
    ("PAPER / DRY-RUN EXECUTION",
     "1E3A5F",
     ["Paper Trades", "Paper Sharpe", "Paper MDD", "Paper CAGR", "Paper P&L ($)"]),
    ("LIVE ACCOUNT EXECUTION",
     "0F3D2E",
     ["Live Trades", "Live Sharpe", "Live MDD", "Live CAGR", "Live P&L ($)"]),
    ("ALPHA DECAY  —  Paper vs Backtest  (positive = better than backtest)",
     "3B2A00",
     ["Δ Sharpe (P−BT)", "Δ MDD (P−BT)", "Δ CAGR (P−BT)"]),
    ("ALPHA DECAY  —  Live vs Backtest  (positive = better than backtest)",
     "3B1400",
     ["Δ Sharpe (L−BT)", "Δ MDD (L−BT)", "Δ CAGR (L−BT)"]),
    ("NOTES",
     "141F2E",
     ["Notes"]),
]

# Per-column number format and font colour logic (for data rows)
_COMP_COL_FMT = {
    "BT Sharpe":    "0.0000", "Paper Sharpe":    "0.0000", "Live Sharpe":    "0.0000",
    "BT MDD":       "0.00%",  "Paper MDD":       "0.00%",  "Live MDD":       "0.00%",
    "BT CAGR":      "0.00%",  "Paper CAGR":      "0.00%",  "Live CAGR":      "0.00%",
    "Paper P&L ($)":"#,##0.00","Live P&L ($)":   "#,##0.00","Capital Base ($)":"#,##0.00",
    "Days Traded":  "#,##0",  "Paper Trades":    "#,##0",  "Live Trades":    "#,##0",
    "Δ Sharpe (P−BT)":"0.0000","Δ Sharpe (L−BT)":"0.0000",
    "Δ MDD (P−BT)": "0.00%",  "Δ MDD (L−BT)":   "0.00%",
    "Δ CAGR (P−BT)":"0.00%",  "Δ CAGR (L−BT)":  "0.00%",
}


def _comp_cell_color(col_name: str, val) -> str:
    """Return hex font color for a comparison data cell."""
    if pd.isna(val):
        return C["dim"]
    fval = float(val)
    if "Sharpe" in col_name and "Δ" not in col_name:
        # BT/Paper/Live Sharpe: ≥1 teal, 0–1 amber, <0 red
        return C["teal"] if fval >= 1.0 else (C["amber"] if fval >= 0 else C["red"])
    if "MDD" in col_name and "Δ" not in col_name:
        # BT/Paper/Live MDD: <-20% red, <-10% amber, else silver
        return C["red"] if fval < -0.20 else (C["amber"] if fval < -0.10 else C["silver"])
    if "CAGR" in col_name and "Δ" not in col_name:
        return C["teal"] if fval >= 0 else C["red"]
    if "P&L" in col_name:
        return C["green"] if fval >= 0 else C["red"]
    if col_name.startswith("Δ"):
        # For all Δ columns: positive = execution ≥ backtest = green; negative = red
        # Note: Δ MDD positive means less-severe drawdown in execution = good
        return C["green"] if fval > 0 else (C["red"] if fval < 0 else C["dim"])
    return C["light_gray"]


def build_comparison_sheet(ws, comp: pd.DataFrame):
    """Build the Backtest vs Execution three-way comparison worksheet.

    Layout
    ------
    Row 1 : Title banner
    Row 2 : Methodology note (auditability)
    Row 3 : Group headers  (merged cells per group, colour-coded)
    Row 4 : Individual column headers
    Row 5+: Data rows
    """
    ws.sheet_view.showGridLines = False
    ws.tab_color    = "8B5CF6"    # purple — visually distinct
    ws.freeze_panes = "A5"        # freeze rows 1-4 (banner, note, groups, headers)

    if comp.empty:
        ws.merge_cells("A1:J2")
        cell = ws.cell(
            row=1, column=1,
            value="  No deployed strategies found — execute live or dry-run trades first."
        )
        cell.font      = _font(size=11, color=C["silver"])
        cell.fill      = _fill(C["navy"])
        cell.alignment = _left()
        ws.row_dimensions[1].height = 28
        return

    # Build column-index lookup from comp.columns
    col_idx = {c: i + 1 for i, c in enumerate(comp.columns)}
    n_cols  = len(comp.columns)

    # ── Row 1: Title banner ────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    banner = ws.cell(
        row=1, column=1,
        value="  BACKTEST vs EXECUTION  —  Alpha Decay Monitor: Backtest / Paper / Live"
    )
    banner.font      = Font(name="Calibri", size=14, bold=True, color=C["gold"])
    banner.fill      = _fill(C["navy"])
    banner.alignment = _left()
    ws.row_dimensions[1].height = 30

    # ── Row 2: Methodology note ────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    note = ws.cell(
        row=2, column=1,
        value=(
            "  BT metrics: Sharpe/MDD/CAGR from Strategy Registry (Risk_Optimized_Compilation) — "
            "strategy-exact match on Symbol+DataPoint+Model+Entry/Exit Model  |  "
            "Paper/Live metrics: avg-cost-basis realised P&L from trade logs (open positions excluded)  |  "
            "Capital Base: first BOT trade notional — same for Paper and Live to ensure comparability  |  "
            "Sharpe: annualised (x sqrt(252)), RFR=0, ddof=1  |  "
            "MDD: (equity-peak)/capital_base, negative decimal (-0.12 = -12%)  |  "
            "CAGR: mean(daily_pnl/capital) x 252 (arithmetic, matches backtest)  |  "
            "Trade Counts: number of fills (not round-trips) — low counts indicate unreliable metrics  |  "
            "Delta positive = execution exceeds backtest"
        )
    )
    note.font      = _font(size=8, color=C["silver"])
    note.fill      = _fill(C["midnight"])
    note.alignment = _left(wrap=False)
    ws.row_dimensions[2].height = 13

    # ── Row 3: Group headers (merged, colour-coded) ────────────────────────────
    for group_label, fill_hex, group_cols in _COMP_GROUPS:
        present = [c for c in group_cols if c in col_idx]
        if not present:
            continue
        c_start = col_idx[present[0]]
        c_end   = col_idx[present[-1]]
        if c_start == c_end:
            cell = ws.cell(row=3, column=c_start, value=f"  {group_label}")
        else:
            ws.merge_cells(start_row=3, start_column=c_start,
                           end_row=3,   end_column=c_end)
            cell = ws.cell(row=3, column=c_start, value=f"  {group_label}")
        cell.font      = Font(name="Calibri", size=9, bold=True, color=C["gold"])
        cell.fill      = _fill(fill_hex)
        cell.alignment = _left()
        cell.border    = _border()
    ws.row_dimensions[3].height = 18

    # ── Row 4: Individual column headers ──────────────────────────────────────
    # Build a fill map: column name → group fill colour
    col_group_fill = {}
    for _, fill_hex, group_cols in _COMP_GROUPS:
        for c in group_cols:
            col_group_fill[c] = fill_hex

    for ci, col_name in enumerate(comp.columns, start=1):
        fill_hex = col_group_fill.get(col_name, C["slate"])
        # Slightly lighten sub-headers vs group headers by using slate for identity
        if col_name in ("#", "Strategy", "Symbol", "Trade Mode",
                        "First Trade", "Last Trade", "Days Traded", "Capital Base ($)"):
            fill_hex = C["slate"]
        cell = ws.cell(row=4, column=ci, value=col_name)
        cell.font      = _font(bold=True, size=9, color=C["white"])
        cell.fill      = _fill(fill_hex)
        cell.alignment = _center(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[4].height = 30

    # ── Rows 5+: Data ─────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(comp.iterrows(), start=1):
        r         = ri + 4
        base_fill = _fill(C["charcoal"] if ri % 2 == 0 else C["dark_row"])
        delta_fill = _fill(C["panel"])   # subtle background for Δ columns

        for ci, col_name in enumerate(comp.columns, start=1):
            val  = row[col_name]
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = _border()

            is_delta = col_name.startswith("Δ")

            # Number format
            if col_name in _COMP_COL_FMT:
                cell.number_format = _COMP_COL_FMT[col_name]

            # Background
            cell.fill = delta_fill if is_delta else base_fill

            # ── Per-column styling ─────────────────────────────────────────────
            if col_name == "#":
                cell.fill      = _fill(C["slate"])
                cell.font      = _font(bold=True, size=9, color=C["gold"])
                cell.alignment = _center()

            elif col_name == "Strategy":
                cell.font      = _font(size=9, color=C["light_gray"])
                cell.alignment = _left(wrap=False)

            elif col_name == "Symbol":
                cell.font      = _font(bold=True, size=10, color=C["gold"])
                cell.alignment = _center()

            elif col_name == "Trade Mode":
                color = (C["teal"]  if str(val) == "Live"    else
                         C["amber"] if str(val) == "Dry Run"  else
                         C["silver"])
                cell.font      = _font(bold=True, size=9, color=color)
                cell.alignment = _center()

            elif col_name in ("First Trade", "Last Trade"):
                _fmt_ts(cell, val)
                cell.font      = _font(size=9, color=C["silver"])
                cell.alignment = _center()

            elif col_name == "Notes":
                cell.font      = _font(size=8, color=C["silver"])
                cell.alignment = _left(wrap=True)

            elif is_delta:
                color = _comp_cell_color(col_name, val)
                cell.font      = _font(bold=True, size=10, color=color)
                cell.alignment = _center()

            else:
                color = _comp_cell_color(col_name, val)
                bold  = "P&L" in col_name
                cell.font      = _font(bold=bold, size=10, color=color)
                cell.alignment = _center()

        ws.row_dimensions[r].height = 20

    # ── Column widths ──────────────────────────────────────────────────────────
    fixed_w = {
        "#":               4,   "Strategy":          36,  "Symbol":         9,
        "Trade Mode":     10,   "First Trade":       17,  "Last Trade":     17,
        "Days Traded":     9,   "Capital Base ($)":  14,
        "BT Sharpe":      11,   "BT MDD":            10,  "BT CAGR":        11,
        "Paper Trades":    9,   "Paper Sharpe":      12,  "Paper MDD":      10,
        "Paper CAGR":     11,   "Paper P&L ($)":     13,
        "Live Trades":     9,   "Live Sharpe":       12,  "Live MDD":       10,
        "Live CAGR":      11,   "Live P&L ($)":      13,
        "Δ Sharpe (P−BT)":12,  "Δ MDD (P−BT)":      10,  "Δ CAGR (P−BT)":  11,
        "Δ Sharpe (L−BT)":12,  "Δ MDD (L−BT)":      10,  "Δ CAGR (L−BT)":  11,
        "Notes":          52,
    }
    for ci, col_name in enumerate(comp.columns, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = fixed_w.get(col_name, 12)

    # ── Autofilter on header row ───────────────────────────────────────────────
    ws.auto_filter.ref = f"A4:{get_column_letter(n_cols)}{len(comp) + 4}"


# ============================================================================
# STEP 6 – ORCHESTRATOR
# ============================================================================

def write_journal(registry: pd.DataFrame, perf: pd.DataFrame,
                  scan_files: list[str], new_count: int,
                  total_before: int, journal_path: str,
                  trades=None,
                  portfolio_state=None,
                  dupes_removed: int = 0,
                  comp=None):
    """Write (or overwrite) the journal with fresh sheets."""
    if trades is None:
        trades = pd.DataFrame()
    if portfolio_state is None:
        portfolio_state = {}
    if comp is None:
        comp = pd.DataFrame()

    if os.path.exists(journal_path):
        try:
            wb = load_workbook(journal_path)
        except Exception as e:
            print(f"[WARN] Could not open existing journal ({e}); creating fresh workbook.")
            wb = openpyxl.Workbook()
        # Remove sheets we'll regenerate
        for sheet_name in [DASHBOARD_SHEET, ACTIVE_SHEET, REGISTRY_SHEET,
                           PERF_SHEET, COMPARISON_SHEET, CHANGE_SHEET]:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
    else:
        wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create sheets in order
    ws_dash   = wb.create_sheet(DASHBOARD_SHEET,  0)
    ws_active = wb.create_sheet(ACTIVE_SHEET,     1)
    ws_reg    = wb.create_sheet(REGISTRY_SHEET,   2)
    ws_perf   = wb.create_sheet(PERF_SHEET,       3)
    ws_comp   = wb.create_sheet(COMPARISON_SHEET, 4)
    ws_change = wb.create_sheet(CHANGE_SHEET,     5)

    print("\n[FORMAT] Building Dashboard ...")
    build_dashboard(ws_dash, registry, perf, scan_files, new_count, trades=trades)

    print("[FORMAT] Building Active Strategy ...")
    build_active_strategy(ws_active, trades, portfolio_state)

    print("[FORMAT] Building Strategy Registry ...")
    build_strategy_registry(ws_reg, registry)

    print("[FORMAT] Building Performance Summary ...")
    build_perf_summary(ws_perf, perf)

    print("[FORMAT] Building Backtest vs Execution comparison ...")
    build_comparison_sheet(ws_comp, comp)

    print("[FORMAT] Building Changelog ...")
    build_changelog(ws_change, new_count, scan_files, total_before,
                    dupes_removed=dupes_removed)

    # Set active sheet to Dashboard
    wb.active = ws_dash

    temp_path = journal_path + ".tmp"
    try:
        wb.save(temp_path)
        os.replace(temp_path, journal_path)
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise RuntimeError(f"[SAVE] Failed to write journal: {e}") from e
    print(f"\n[SAVE]  Journal saved -> {journal_path}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 70)
    print("  AQS TRADING JOURNAL GENERATOR")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # Pre-flight: verify required directories exist
    if not os.path.isdir(WF_RESULTS_DIR):
        print(f"[ERROR] WFAlphaResults directory not found: {WF_RESULTS_DIR}")
        print("[ERROR] Cannot proceed without source files. Exiting.")
        sys.exit(1)
    if not os.path.isdir(TRADES_DIR):
        print(f"[WARN] Trades directory not found: {TRADES_DIR} — trade data will be empty.")

    # 1. Scan for source files (prefer Risk_Optimized_Compilation, fall back to Final_Compilation)
    files, sheet_name = find_source_files(WF_RESULTS_DIR)

    # 2. Extract and compile strategies (if any found)
    if files:
        incoming = compile_all_strategies(files, sheet_name)
    else:
        print("[INFO] No compilation files found — will update journal from trade logs only.")
        incoming = pd.DataFrame()

    # 3. Load existing journal registry (if any)
    existing = load_existing_registry(JOURNAL_PATH)
    total_before = len(existing)

    # 4. Merge: append only new strategies (skip if no incoming alphas)
    if not incoming.empty:
        registry, new_count = merge_new_strategies(existing, incoming)
        registry = clean_registry(registry)

        # 4b. Final-pass duplicate audit (safety net for any edge cases)
        registry, dupes_removed = audit_and_dedupe_registry(registry)
        if dupes_removed > 0:
            registry = clean_registry(registry)   # re-number after removal
    else:
        # No new alphas — retain existing registry as-is
        registry = clean_registry(existing) if not existing.empty else pd.DataFrame()
        new_count = 0
        dupes_removed = 0

    print(f"\n[MERGE]  Existing: {total_before}  |  "
          f"New: {new_count}  |  Dupes removed: {dupes_removed}  |  "
          f"Total: {len(registry)}")

    # 5. Load live trade data & portfolio state (needed before perf summary)
    print("\n[TRADES] Scanning ibkr_deployment/trades/ ...")
    trades = load_trades(TRADES_DIR)

    print("[STATE]  Loading portfolio state ...")
    portfolio_state = load_portfolio_state(PORTFOLIO_STATE_DIR)

    # Guard: nothing to do if there are no strategies and no trade data
    if registry.empty and (trades is None or trades.empty):
        print("[ABORT] No strategies and no trade data found. Exiting.")
        sys.exit(0)

    # 6. Build Performance Summary (cross-referenced with trade logs)
    perf = build_performance_summary(registry, trades=trades)

    # 6b. Build Backtest vs Execution comparison
    print("\n[COMPARE] Computing backtest vs execution metrics ...")
    comp = build_backtest_vs_execution_data(registry, trades)
    print(f"[COMPARE] {len(comp)} deployed strategy row(s) in comparison sheet.")

    # 7. Backup existing journal before overwriting
    backup_journal(JOURNAL_PATH)

    # 8. Write journal
    write_journal(registry, perf, files, new_count, total_before, JOURNAL_PATH,
                  trades=trades, portfolio_state=portfolio_state,
                  dupes_removed=dupes_removed, comp=comp)

    print("\n" + "=" * 70)
    print("  DONE -- AQS_Trading_Journal.xlsx is ready.")
    print("=" * 70)


# ============================================================================
# DELIVERY — Summary Extraction, Email, Telegram
# ============================================================================

_EMAIL_REQUIRED_KEYS = {"smtp_server", "smtp_port", "sender", "app_password", "recipients"}

def load_email_config():
    """Load and validate email configuration from ibkr_deployment/email_config.json."""
    if not os.path.exists(EMAIL_CONFIG_PATH):
        return None
    try:
        with open(EMAIL_CONFIG_PATH, "r") as f:
            cfg = json.load(f)
    except Exception as e:
        print(f"[WARN] Could not load email config: {e}")
        return None
    missing = _EMAIL_REQUIRED_KEYS - set(cfg.keys())
    if missing:
        print(f"[WARN] Email config missing required keys: {', '.join(sorted(missing))} — skipping email.")
        return None
    return cfg


def load_telegram_config():
    """Load Telegram configuration from ibkr_deployment/telegram_config.json."""
    if not os.path.exists(TELEGRAM_CONFIG_PATH):
        return None
    try:
        with open(TELEGRAM_CONFIG_PATH, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"[WARN] Could not load Telegram config: {e}")
        return None


def extract_summary(journal_path):
    """Extract key stats from the journal for delivery message body.

    All sheets have a banner in row 1 and column headers in row 2,
    so we use header=1 (0-indexed) to skip the banner.
    """
    xls = pd.ExcelFile(journal_path)
    summary = {}

    # Strategy Registry stats
    if "Strategy Registry" in xls.sheet_names:
        reg = xls.parse("Strategy Registry", header=1)
        summary["total_strategies"] = len(reg)
        summary["symbols"] = (
            sorted(reg["Symbol"].dropna().unique().tolist())
            if "Symbol" in reg.columns else []
        )
        if "Status" in reg.columns:
            summary["status_counts"] = reg["Status"].value_counts().to_dict()

    # Performance Summary stats
    if "Performance Summary" in xls.sheet_names:
        perf = xls.parse("Performance Summary", header=1)
        if "Avg Sharpe" in perf.columns:
            val = perf["Avg Sharpe"].mean()
            if pd.notna(val):
                summary["avg_sharpe"] = val
        if "Avg Trade Count" in perf.columns and perf["Avg Trade Count"].notna().any():
            summary["total_trades"] = int(perf["Avg Trade Count"].sum())

    # Active Strategy stats
    if "Active Strategy" in xls.sheet_names:
        active = xls.parse("Active Strategy", header=1)
        if not active.empty:
            summary["active_rows"] = len(active)

    # Changelog — latest entry
    if "Changelog" in xls.sheet_names:
        cl = xls.parse("Changelog", header=1)
        if not cl.empty:
            summary["last_import"] = str(cl.iloc[0].get("Timestamp", ""))
            summary["new_strategies"] = cl.iloc[0].get("New Strategies", 0)

    return summary


def format_summary_text(summary, timestamp):
    """Format summary dict into a readable text message."""
    lines = [
        f"AQS Trading Journal — {timestamp}",
        "",
        f"Strategies: {summary.get('total_strategies', 0)}",
        f"Symbols: {', '.join(summary.get('symbols', []))}",
    ]

    if "status_counts" in summary:
        for status, count in summary["status_counts"].items():
            lines.append(f"  - {status}: {count}")

    if "avg_sharpe" in summary:
        lines.append(f"Avg Sharpe: {summary['avg_sharpe']:.2f}")

    if "total_trades" in summary:
        lines.append(f"Total Trades: {summary['total_trades']}")

    if "new_strategies" in summary:
        lines.append(f"New This Run: {summary['new_strategies']}")

    return "\n".join(lines)


def send_email(journal_path, summary_text, config):
    """Send journal via Gmail SMTP with .xlsx attachment."""
    msg = MIMEMultipart()
    msg["From"] = config["sender"]
    msg["To"] = ", ".join(config["recipients"])
    msg["Subject"] = (
        f"{config.get('subject_prefix', 'AQS Trading Journal')} "
        f"— {datetime.now().strftime('%Y-%m-%d')}"
    )

    msg.attach(MIMEText(summary_text, "plain"))

    with open(journal_path, "rb") as f:
        part = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={os.path.basename(journal_path)}",
        )
        msg.attach(part)

    with smtplib.SMTP(config["smtp_server"], config["smtp_port"]) as server:
        server.starttls()
        server.login(config["sender"], config["app_password"])
        server.send_message(msg)


_TG_RETRY_ATTEMPTS = 3
_TG_RETRY_DELAYS   = [5, 10, 20]   # seconds between attempts

def _send_telegram_doc(token, chat_id, journal_path, caption, topic_id=None):
    """Low-level: send one document to one Telegram destination with retry."""
    import time
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    data = {"chat_id": chat_id, "caption": caption[:1024]}
    if topic_id:
        data["message_thread_id"] = topic_id

    for attempt in range(1, _TG_RETRY_ATTEMPTS + 1):
        try:
            with open(journal_path, "rb") as f:
                resp = requests.post(
                    url,
                    data=data,
                    files={"document": (os.path.basename(journal_path), f)},
                    timeout=30,
                )
        except requests.RequestException as e:
            print(f"[WARN] Telegram attempt {attempt}/{_TG_RETRY_ATTEMPTS} network error: {e}")
        else:
            if resp.status_code == 200:
                print(f"[OK] Journal sent to Telegram (chat_id={chat_id})")
                return True
            print(
                f"[WARN] Telegram attempt {attempt}/{_TG_RETRY_ATTEMPTS} failed "
                f"({chat_id}): {resp.status_code} {resp.text[:200]}"
            )

        if attempt < _TG_RETRY_ATTEMPTS:
            time.sleep(_TG_RETRY_DELAYS[attempt - 1])

    print(f"[ERROR] Telegram sendDocument failed after {_TG_RETRY_ATTEMPTS} attempts ({chat_id}).")
    return False


def send_telegram_document(journal_path, summary_text, config):
    """Send journal via Telegram Bot API sendDocument to primary bot only.
    The secondary destination is reserved for trade execution alerts."""
    token = (config.get("bot_token") or "").strip()
    chat_id = (config.get("chat_id") or "").strip()
    if not token or not chat_id:
        print("[WARN] Telegram enabled but missing bot_token/chat_id.")
        return False

    return _send_telegram_doc(token, chat_id, journal_path, summary_text)


def deliver_journal(journal_path, email_only=False, telegram_only=False,
                    dry_run=False):
    """Deliver journal via configured channels."""
    if not os.path.exists(journal_path):
        print(f"[DELIVER] Journal not found: {journal_path}")
        return False

    summary = extract_summary(journal_path)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary_text = format_summary_text(summary, timestamp)

    print(f"\n[DELIVER] Summary:\n{summary_text}\n")

    if dry_run:
        print("[DELIVER] Dry run — not sending.")
        return True

    send_email_flag = not telegram_only
    send_tg_flag = not email_only
    success = True

    if send_email_flag:
        email_cfg = load_email_config()
        if email_cfg and email_cfg.get("enabled"):
            try:
                send_email(journal_path, summary_text, email_cfg)
                print("[OK] Email sent.")
            except Exception as e:
                print(f"[ERROR] Email delivery failed: {e}")
                success = False
        else:
            print("[SKIP] Email not configured or disabled.")

    if send_tg_flag:
        tg_cfg = load_telegram_config()
        if tg_cfg and tg_cfg.get("enabled"):
            try:
                ok = send_telegram_document(journal_path, summary_text, tg_cfg)
                if not ok:
                    success = False
            except Exception as e:
                print(f"[ERROR] Telegram delivery failed: {e}")
                success = False
        else:
            print("[SKIP] Telegram not configured or disabled.")

    return success


# ============================================================================
# CLI ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="AQS Trading Journal Generator & Delivery"
    )
    parser.add_argument(
        "--deliver", action="store_true",
        help="Generate + deliver via email/Telegram",
    )
    parser.add_argument(
        "--email-only", action="store_true",
        help="Generate + deliver via email only",
    )
    parser.add_argument(
        "--telegram-only", action="store_true",
        help="Generate + deliver via Telegram only",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Generate + print delivery summary, don't send",
    )
    parser.add_argument(
        "--deliver-only", action="store_true",
        help="Skip generation, deliver existing journal",
    )
    args = parser.parse_args()

    wants_delivery = (
        args.deliver or args.email_only or args.telegram_only
        or args.dry_run or args.deliver_only
    )

    # Step 1: Generate (unless --deliver-only)
    if not args.deliver_only:
        main()

    # Step 2: Deliver (if any delivery flag is set)
    if wants_delivery:
        ok = deliver_journal(
            JOURNAL_PATH,
            email_only=args.email_only,
            telegram_only=args.telegram_only,
            dry_run=args.dry_run,
        )
        if not ok:
            sys.exit(1)
